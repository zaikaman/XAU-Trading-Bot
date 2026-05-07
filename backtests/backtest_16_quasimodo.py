"""
Backtest #16 — SMC + RTM Quasimodo (QM) Pattern
==================================================================
Base: SMC-Only v4 (#1 Baseline)
Added: Quasimodo pattern detection from RTM methodology

QM = 5-point reversal pattern:
  Bearish: H(A) → L(B) → HH(C) → LL(D) → entry at QML (A level)
  Bullish: L(A) → H(B) → LL(C) → HH(D) → entry at QML (A level)

Two improvements:
  1. QM-enhanced SL: When SMC signal + QM align → use tighter SL from QM head
  2. QM-only entries: When price retraces to QML zone without SMC signal

Usage:
    python backtests/backtest_16_quasimodo.py
"""

import polars as pl
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, field
from enum import Enum
import sys
import os
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.mt5_connector import MT5Connector
from src.smc_polars import SMCAnalyzer, SMCSignal
from src.feature_eng import FeatureEngineer
from src.regime_detector import MarketRegimeDetector, MarketRegime
from src.ml_model import TradingModel
from src.config import get_config
from src.dynamic_confidence import DynamicConfidenceManager, create_dynamic_confidence, MarketQuality
from loguru import logger

logger.remove()
logger.add(sys.stderr, level="WARNING")

WIB = ZoneInfo("Asia/Jakarta")


# ─── Enums & Dataclasses ──────────────────────────────────────

class TradeResult(Enum):
    WIN = "WIN"
    LOSS = "LOSS"
    BREAKEVEN = "BREAKEVEN"

class ExitReason(Enum):
    TAKE_PROFIT = "take_profit"
    SMART_TP = "smart_tp"
    PEAK_PROTECT = "peak_protect"
    EARLY_EXIT = "early_exit"
    EARLY_CUT = "early_cut"
    MAX_LOSS = "max_loss"
    STALL = "stall"
    TREND_REVERSAL = "trend_reversal"
    TIMEOUT = "timeout"
    WEEKEND_CLOSE = "weekend_close"
    TRAILING_SL = "trailing_sl"
    BREAKEVEN_EXIT = "breakeven_exit"
    DAILY_LIMIT = "daily_limit"
    REGIME_DANGER = "regime_danger"
    MARKET_SIGNAL = "market_signal"

class TradingMode(Enum):
    NORMAL = "normal"
    RECOVERY = "recovery"
    PROTECTED = "protected"
    STOPPED = "stopped"

@dataclass
class SimulatedTrade:
    ticket: int
    entry_time: datetime
    exit_time: datetime
    direction: str
    entry_price: float
    exit_price: float
    stop_loss: float
    take_profit: float
    lot_size: float
    profit_usd: float
    profit_pips: float
    result: TradeResult
    exit_reason: ExitReason
    smc_confidence: float
    regime: str
    session: str
    signal_reason: str
    has_bos: bool = False
    has_choch: bool = False
    has_fvg: bool = False
    has_ob: bool = False
    atr_at_entry: float = 0.0
    rr_ratio: float = 0.0
    trading_mode: str = "normal"
    entry_source: str = "SMC"  # "SMC", "QM+SMC", "QM-only"

@dataclass
class BacktestStats:
    total_trades: int = 0
    wins: int = 0
    losses: int = 0
    total_profit: float = 0.0
    total_loss: float = 0.0
    max_drawdown: float = 0.0
    max_drawdown_usd: float = 0.0
    win_rate: float = 0.0
    profit_factor: float = 0.0
    avg_win: float = 0.0
    avg_loss: float = 0.0
    avg_trade: float = 0.0
    expectancy: float = 0.0
    sharpe_ratio: float = 0.0
    trades: List[SimulatedTrade] = field(default_factory=list)
    equity_curve: List[float] = field(default_factory=list)
    avoided_signals: int = 0
    daily_limit_stops: int = 0
    recovery_mode_trades: int = 0
    # QM stats
    qm_enhanced: int = 0     # SMC + QM aligned → QM SL used
    qm_only: int = 0         # QM-only entries (no SMC signal)
    standard_smc: int = 0    # Standard SMC entries (no QM)
    qm_patterns_found: int = 0


# ─── QM + SMC Backtest ──────────────────────────

class QuasimodoBacktest:
    """SMC-Only v4 + RTM Quasimodo pattern detection."""

    def __init__(
        self,
        capital: float = 5000.0,
        max_daily_loss_percent: float = 5.0,
        max_loss_per_trade_percent: float = 1.0,
        base_lot_size: float = 0.01,
        max_lot_size: float = 0.02,
        recovery_lot_size: float = 0.01,
        trend_reversal_threshold: float = 0.75,
        max_concurrent_positions: int = 2,
        breakeven_pips: float = 30.0,
        trail_start_pips: float = 50.0,
        trail_step_pips: float = 30.0,
        min_profit_to_protect: float = 5.0,
        max_drawdown_from_peak: float = 50.0,
        trade_cooldown_bars: int = 10,
        trend_reversal_mult: float = 0.6,
        # QM params
        qm_lookback: int = 60,          # Bars to scan for QM patterns
        qm_max_age: int = 30,           # Max bars since D-point for valid QM
        qm_zone_tolerance_pct: float = 0.004,  # 0.4% = ~$11 at $2800
        qm_rr_ratio: float = 2.0,       # RR for QM entries
    ):
        self.capital = capital
        self.max_daily_loss_usd = capital * (max_daily_loss_percent / 100)
        self.max_loss_per_trade = capital * (max_loss_per_trade_percent / 100)
        self.base_lot_size = base_lot_size
        self.max_lot_size = max_lot_size
        self.recovery_lot_size = recovery_lot_size
        self.trend_reversal_threshold = trend_reversal_threshold
        self.max_concurrent_positions = max_concurrent_positions
        self.breakeven_pips = breakeven_pips
        self.trail_start_pips = trail_start_pips
        self.trail_step_pips = trail_step_pips
        self.min_profit_to_protect = min_profit_to_protect
        self.max_drawdown_from_peak = max_drawdown_from_peak
        self.trade_cooldown_bars = trade_cooldown_bars
        self.trend_reversal_mult = trend_reversal_mult
        self.qm_lookback = qm_lookback
        self.qm_max_age = qm_max_age
        self.qm_zone_tolerance_pct = qm_zone_tolerance_pct
        self.qm_rr_ratio = qm_rr_ratio

        config = get_config()
        self.smc = SMCAnalyzer(
            swing_length=config.smc.swing_length,
            ob_lookback=config.smc.ob_lookback,
        )
        self.features = FeatureEngineer()
        self.dynamic_confidence = create_dynamic_confidence()

        self.ml_model = TradingModel(model_path="models/xgboost_model.pkl")
        try:
            self.ml_model.load()
            print("  ML model loaded (for exit evaluation)")
        except Exception:
            print("  [WARN] ML model not loaded")

        self.regime_detector = MarketRegimeDetector(model_path="models/hmm_regime.pkl")
        try:
            self.regime_detector.load()
        except Exception:
            print("  [WARN] HMM model not loaded")

        self._ticket_counter = 2000000

    # ═══ QUASIMODO PATTERN DETECTION ═══

    def _detect_qm_patterns(self, df_slice: pl.DataFrame) -> List[dict]:
        """
        Detect Quasimodo patterns from swing points.

        Bearish QM: H(A) → L(B) → HH(C) → LL(D)
          - C > A (higher high = head)
          - D < B (lower low = CHoCH / structure break)
          - QML = A level (entry zone for SELL)
          - SL = above C (head)

        Bullish QM: L(A) → H(B) → LL(C) → HH(D)
          - C < A (lower low = head)
          - D > B (higher high = CHoCH / structure break)
          - QML = A level (entry zone for BUY)
          - SL = below C (head)
        """
        n = len(df_slice)
        if n < self.qm_lookback:
            return []

        sh_col = df_slice["swing_high"].to_list()
        sl_col = df_slice["swing_low"].to_list()

        # Get swing high/low levels (actual prices at swing point bars)
        sh_level = df_slice["swing_high_level"].to_list()
        sl_level = df_slice["swing_low_level"].to_list()

        # Collect recent swing points as (index, price, type)
        swings = []
        scan_start = max(0, n - self.qm_lookback)
        for i in range(scan_start, n):
            if sh_col[i] == 1 and sh_level[i] is not None:
                swings.append((i, float(sh_level[i]), "H"))
            if sl_col[i] == -1 and sl_level[i] is not None:
                swings.append((i, float(sl_level[i]), "L"))

        # Sort by index (should already be, but ensure)
        swings.sort(key=lambda x: x[0])

        if len(swings) < 4:
            return []

        patterns = []

        # Scan for QM patterns in consecutive swing points
        for i in range(len(swings) - 3):
            a = swings[i]
            b = swings[i + 1]
            c = swings[i + 2]
            d = swings[i + 3]

            # Bearish QM: H(A) - L(B) - H(C) - L(D)
            # where C > A (higher high) and D < B (lower low)
            if a[2] == "H" and b[2] == "L" and c[2] == "H" and d[2] == "L":
                if c[1] > a[1] and d[1] < b[1]:
                    # D must be recent enough
                    if n - d[0] <= self.qm_max_age:
                        sl_price = c[1] + 2.0  # $2 above head
                        patterns.append({
                            "direction": "SELL",
                            "qml_level": a[1],
                            "head_level": c[1],
                            "sl_price": sl_price,
                            "d_idx": d[0],
                            "freshness": n - d[0],
                        })

            # Bullish QM: L(A) - H(B) - L(C) - H(D)
            # where C < A (lower low) and D > B (higher high)
            if a[2] == "L" and b[2] == "H" and c[2] == "L" and d[2] == "H":
                if c[1] < a[1] and d[1] > b[1]:
                    if n - d[0] <= self.qm_max_age:
                        sl_price = c[1] - 2.0  # $2 below head
                        patterns.append({
                            "direction": "BUY",
                            "qml_level": a[1],
                            "head_level": c[1],
                            "sl_price": sl_price,
                            "d_idx": d[0],
                            "freshness": n - d[0],
                        })

        return patterns

    def _find_matching_qm(self, patterns: List[dict], current_price: float, direction: str = None) -> Optional[dict]:
        """Find QM pattern where current price is near QML level."""
        tolerance = current_price * self.qm_zone_tolerance_pct

        best = None
        for qm in patterns:
            if direction and qm["direction"] != direction:
                continue

            dist = abs(current_price - qm["qml_level"])
            if dist <= tolerance:
                # For SELL: price should be AT or ABOVE QML
                # For BUY: price should be AT or BELOW QML
                if qm["direction"] == "SELL" and current_price >= qm["qml_level"] - tolerance:
                    if best is None or qm["freshness"] < best["freshness"]:
                        best = qm
                elif qm["direction"] == "BUY" and current_price <= qm["qml_level"] + tolerance:
                    if best is None or qm["freshness"] < best["freshness"]:
                        best = qm

        return best

    # ── Session filter (synced) ──

    def _get_session_from_time(self, dt):
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib_time = dt.astimezone(WIB)
        hour = wib_time.hour
        if 6 <= hour < 15:
            return "Sydney-Tokyo", True, 0.5
        elif 15 <= hour < 16:
            return "Tokyo-London Overlap", True, 0.75
        elif 16 <= hour < 19:
            return "London Early", True, 0.8
        elif 19 <= hour < 24:
            return "London-NY Overlap (Golden)", True, 1.0
        elif 0 <= hour < 4:
            return "NY Session", True, 0.9
        else:
            return "Off Hours", False, 0.0

    def _hours_to_golden(self, dt):
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib = dt.astimezone(WIB)
        if 19 <= wib.hour < 24:
            return 0
        target = wib.replace(hour=19, minute=0, second=0, microsecond=0)
        if wib.hour >= 19:
            target += timedelta(days=1)
        return max(0, (target - wib).total_seconds() / 3600)

    def _is_near_weekend_close(self, dt):
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib = dt.astimezone(WIB)
        return wib.weekday() == 5 and wib.hour >= 4 and wib.minute >= 30

    def _calculate_lot_size(self, confidence, regime, trading_mode, session_mult):
        if trading_mode == TradingMode.STOPPED:
            return 0
        lot = self.base_lot_size
        if trading_mode in (TradingMode.RECOVERY, TradingMode.PROTECTED):
            lot = self.recovery_lot_size
        else:
            if confidence >= 0.65:
                lot = self.max_lot_size
            elif confidence >= 0.55:
                lot = self.base_lot_size
            else:
                lot = self.recovery_lot_size
        if regime.lower() in ["high_volatility", "crisis"]:
            lot = self.recovery_lot_size
        lot = max(0.01, lot * session_mult)
        return round(lot, 2)

    # ── Full exit simulation (all 3 systems — same as baseline) ──

    def _simulate_trade_exit(
        self, df, entry_idx, direction, entry_price, take_profit, stop_loss,
        lot_size, daily_loss_so_far, feature_cols, max_bars=100,
    ):
        pip_value = 10
        highs = df["high"].to_list()
        lows = df["low"].to_list()
        closes = df["close"].to_list()
        times = df["time"].to_list()

        atr = 12.0
        if "atr" in df.columns:
            atr_list = df["atr"].to_list()
            if entry_idx < len(atr_list) and atr_list[entry_idx] is not None:
                atr = atr_list[entry_idx]

        reversal_momentum_threshold = atr * self.trend_reversal_mult
        min_loss_for_reversal_exit = atr * 0.8

        profit_history = []
        peak_profit = 0.0
        stall_count = 0
        reversal_warnings = 0
        current_sl = stop_loss
        breakeven_moved = False

        if direction == "BUY":
            target_tp_profit = (take_profit - entry_price) / 0.1 * pip_value * lot_size
        else:
            target_tp_profit = (entry_price - take_profit) / 0.1 * pip_value * lot_size

        cached_ml_signal = ""
        cached_ml_confidence = 0.5

        for i in range(entry_idx + 1, min(entry_idx + max_bars, len(df))):
            high = highs[i]
            low = lows[i]
            close = closes[i]
            current_time = times[i]

            if direction == "BUY":
                current_pips = (close - entry_price) / 0.1
                pip_profit_from_entry = current_pips
            else:
                current_pips = (entry_price - close) / 0.1
                pip_profit_from_entry = current_pips
            current_profit = current_pips * pip_value * lot_size

            profit_history.append(current_profit)
            if current_profit > peak_profit:
                peak_profit = current_profit

            bars_since_entry = i - entry_idx

            if bars_since_entry % 4 == 0 and self.ml_model.fitted:
                try:
                    df_s = df.head(i + 1)
                    ml_pred = self.ml_model.predict(df_s, feature_cols)
                    cached_ml_signal = ml_pred.signal
                    cached_ml_confidence = ml_pred.confidence
                except Exception:
                    pass

            momentum = 0.0
            if len(profit_history) >= 3:
                recent = profit_history[-5:] if len(profit_history) >= 5 else profit_history
                momentum = max(-100, min(100, ((recent[-1] - recent[0]) / 10) * 50))

            profit_growing = momentum > 0

            # A) TP hit
            if direction == "BUY" and high >= take_profit:
                pips = (take_profit - entry_price) / 0.1
                return pips * pip_value * lot_size, pips, ExitReason.TAKE_PROFIT, i, take_profit
            elif direction == "SELL" and low <= take_profit:
                pips = (entry_price - take_profit) / 0.1
                return pips * pip_value * lot_size, pips, ExitReason.TAKE_PROFIT, i, take_profit

            # Trailing/breakeven SL hit
            if breakeven_moved and current_sl > 0:
                if direction == "BUY" and low <= current_sl:
                    pips = (current_sl - entry_price) / 0.1
                    reason = ExitReason.TRAILING_SL if pip_profit_from_entry >= self.trail_start_pips else ExitReason.BREAKEVEN_EXIT
                    return pips * pip_value * lot_size, pips, reason, i, current_sl
                elif direction == "SELL" and high >= current_sl:
                    pips = (entry_price - current_sl) / 0.1
                    reason = ExitReason.TRAILING_SL if pip_profit_from_entry >= self.trail_start_pips else ExitReason.BREAKEVEN_EXIT
                    return pips * pip_value * lot_size, pips, reason, i, current_sl

            # Breakeven move
            if pip_profit_from_entry >= self.breakeven_pips and not breakeven_moved:
                current_sl = entry_price + 2 if direction == "BUY" else entry_price - 2
                breakeven_moved = True

            # Trailing SL
            if pip_profit_from_entry >= self.trail_start_pips:
                trail_dist = self.trail_step_pips * 0.1
                if direction == "BUY":
                    new_sl = close - trail_dist
                    if new_sl > current_sl: current_sl = new_sl
                else:
                    new_sl = close + trail_dist
                    if current_sl == 0 or new_sl < current_sl: current_sl = new_sl

            # Peak protect
            if peak_profit > self.min_profit_to_protect:
                dd_pct = ((peak_profit - current_profit) / peak_profit) * 100 if peak_profit > 0 else 0
                if dd_pct > self.max_drawdown_from_peak:
                    return current_profit, current_pips, ExitReason.PEAK_PROTECT, i, close

            # Market analysis
            if bars_since_entry % 5 == 0 and bars_since_entry >= 5 and i >= 20:
                ma_fast = np.mean(closes[i-4:i+1])
                ma_slow = np.mean(closes[i-19:i+1])
                trend = "NEUTRAL"
                if ma_fast > ma_slow * 1.001: trend = "BULLISH"
                elif ma_fast < ma_slow * 0.999: trend = "BEARISH"
                roc = (closes[i] / closes[max(0,i-4)] - 1) * 100
                mom_dir = "BULLISH" if roc > 0.3 else ("BEARISH" if roc < -0.3 else "NEUTRAL")

                rsi_val = None
                if "rsi" in df.columns:
                    rsi_list = df["rsi"].to_list()
                    if i < len(rsi_list): rsi_val = rsi_list[i]

                urgency = 0
                should_exit = False
                if cached_ml_confidence > 0.75:
                    if (direction == "BUY" and cached_ml_signal == "SELL") or \
                       (direction == "SELL" and cached_ml_signal == "BUY"):
                        should_exit = True; urgency += 2
                if rsi_val:
                    if (rsi_val > 75 and direction == "BUY") or (rsi_val < 25 and direction == "SELL"):
                        should_exit = True; urgency += 2
                if (direction == "BUY" and trend == "BEARISH" and mom_dir == "BEARISH") or \
                   (direction == "SELL" and trend == "BULLISH" and mom_dir == "BULLISH"):
                    should_exit = True; urgency += 3

                if should_exit and current_profit > self.min_profit_to_protect / 2:
                    return current_profit, current_pips, ExitReason.MARKET_SIGNAL, i, close
                if urgency >= 7 and current_profit > 0:
                    return current_profit, current_pips, ExitReason.MARKET_SIGNAL, i, close

            # Weekend
            if self._is_near_weekend_close(current_time):
                if current_profit > -10:
                    return current_profit, current_pips, ExitReason.WEEKEND_CLOSE, i, close

            # B) SmartRiskManager
            if current_profit >= 15:
                if current_profit >= 40:
                    return current_profit, current_pips, ExitReason.SMART_TP, i, close
                if current_profit >= 25 and momentum < -30:
                    return current_profit, current_pips, ExitReason.SMART_TP, i, close
                if peak_profit > 30 and current_profit < peak_profit * 0.6:
                    return current_profit, current_pips, ExitReason.PEAK_PROTECT, i, close
                if current_profit >= 20:
                    progress = (current_profit / target_tp_profit) * 100 if target_tp_profit > 0 else 0
                    tp_prob = min(40, max(0, progress * 0.4)) + ((momentum + 100) / 200) * 30 + 10 - min(10, bars_since_entry / 4 * 2)
                    if tp_prob < 25:
                        return current_profit, current_pips, ExitReason.SMART_TP, i, close

            if 5 <= current_profit < 15:
                if momentum < -50 and cached_ml_confidence >= 0.65:
                    is_rev = (direction == "BUY" and cached_ml_signal == "SELL") or \
                             (direction == "SELL" and cached_ml_signal == "BUY")
                    if is_rev:
                        return current_profit, current_pips, ExitReason.EARLY_EXIT, i, close

            if current_profit < 0:
                loss_pct = abs(current_profit) / self.max_loss_per_trade * 100
                if momentum < -30 and loss_pct >= 30:
                    return current_profit, current_pips, ExitReason.EARLY_CUT, i, close

            is_ml_rev = False
            if (direction == "BUY" and cached_ml_signal == "SELL" and cached_ml_confidence >= self.trend_reversal_threshold) or \
               (direction == "SELL" and cached_ml_signal == "BUY" and cached_ml_confidence >= self.trend_reversal_threshold):
                is_ml_rev = True
                reversal_warnings += 1

            if is_ml_rev and current_profit < -8 and abs(current_profit) > self.max_loss_per_trade * 0.4:
                return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close
            if reversal_warnings >= 3 and current_profit < -10:
                return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close

            if current_profit <= -(self.max_loss_per_trade * 0.50):
                htg = self._hours_to_golden(current_time)
                if not (htg <= 1 and htg > 0 and momentum > -40):
                    return current_profit, current_pips, ExitReason.MAX_LOSS, i, close

            if len(profit_history) >= 10:
                r = max(profit_history[-10:]) - min(profit_history[-10:])
                if r < 3 and current_profit < -15:
                    stall_count += 1
                    if stall_count >= 5:
                        return current_profit, current_pips, ExitReason.STALL, i, close

            pot_daily = daily_loss_so_far + abs(min(0, current_profit))
            if pot_daily >= self.max_daily_loss_usd:
                return current_profit, current_pips, ExitReason.DAILY_LIMIT, i, close

            # C) Time exit
            if bars_since_entry >= 16:
                if current_profit < 5 and not profit_growing:
                    if current_profit > -15:
                        return current_profit, current_pips, ExitReason.TIMEOUT, i, close
            if bars_since_entry >= 24:
                if current_profit < 10 or not profit_growing:
                    return current_profit, current_pips, ExitReason.TIMEOUT, i, close
            if bars_since_entry >= 32:
                return current_profit, current_pips, ExitReason.TIMEOUT, i, close

            if bars_since_entry > 10:
                rc = closes[i-5:i+1]
                mom_val = rc[-1] - rc[0]
                if direction == "BUY" and mom_val < -reversal_momentum_threshold and current_profit < -min_loss_for_reversal_exit:
                    return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close
                elif direction == "SELL" and mom_val > reversal_momentum_threshold and current_profit < -min_loss_for_reversal_exit:
                    return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close

        final_idx = min(entry_idx + max_bars - 1, len(df) - 1)
        fp = closes[final_idx]
        pips = (fp - entry_price) / 0.1 if direction == "BUY" else (entry_price - fp) / 0.1
        return pips * 10 * lot_size, pips, ExitReason.TIMEOUT, final_idx, fp

    # ── Main backtest run ──

    def run(self, df, start_date=None, end_date=None, initial_capital=5000.0):
        stats = BacktestStats()
        capital = initial_capital
        peak_capital = initial_capital
        stats.equity_curve.append(capital)

        daily_loss = 0.0
        daily_profit = 0.0
        consecutive_losses = 0
        trading_mode = TradingMode.NORMAL
        current_date = None

        feature_cols = []
        if self.ml_model.fitted and self.ml_model.feature_names:
            feature_cols = [f for f in self.ml_model.feature_names if f in df.columns]

        times = df["time"].to_list()
        closes = df["close"].to_list()
        start_idx = next((i for i, t in enumerate(times) if t >= start_date), 100) if start_date else 100
        end_idx = next((i for i, t in enumerate(times) if t > end_date), len(df) - 100) if end_date else len(df) - 100

        last_trade_idx = -self.trade_cooldown_bars * 2

        print(f"\n  Running SMC + Quasimodo backtest...")
        print(f"  QM: lookback={self.qm_lookback}, max_age={self.qm_max_age}, tolerance={self.qm_zone_tolerance_pct*100:.1f}%")
        print(f"  QM RR: 1:{self.qm_rr_ratio}")
        print(f"  Date range: {times[start_idx]} to {times[end_idx - 1]}")
        print(f"  Total bars: {end_idx - start_idx}")

        for i in range(start_idx, end_idx):
            if i - last_trade_idx < self.trade_cooldown_bars:
                continue

            current_time = times[i]
            current_price = closes[i]

            # Daily reset
            trade_date = current_time.date() if hasattr(current_time, 'date') else current_time
            if current_date is None or trade_date != current_date:
                daily_loss = 0.0
                daily_profit = 0.0
                current_date = trade_date
                if consecutive_losses < 2:
                    trading_mode = TradingMode.NORMAL

            if trading_mode == TradingMode.STOPPED:
                continue

            session_name, can_trade, lot_mult = self._get_session_from_time(current_time)
            if not can_trade:
                continue

            if hasattr(current_time, 'weekday') and current_time.weekday() >= 5:
                continue

            df_slice = df.head(i + 1)

            # Regime check
            regime = "normal"
            try:
                if self.regime_detector.fitted:
                    regime_state = self.regime_detector.get_current_state(df_slice)
                    if regime_state:
                        regime = regime_state.regime.value
                        if regime_state.regime == MarketRegime.CRISIS:
                            continue
                        if regime_state.recommendation == "SLEEP":
                            continue
            except Exception:
                pass

            # Dynamic confidence AVOID
            ml_signal = ""
            ml_confidence = 0.5
            try:
                if self.ml_model.fitted and feature_cols:
                    ml_pred = self.ml_model.predict(df_slice, feature_cols)
                    ml_signal = ml_pred.signal
                    ml_confidence = ml_pred.confidence

                market_analysis = self.dynamic_confidence.analyze_market(
                    session=session_name, regime=regime, volatility="medium",
                    trend_direction=regime, has_smc_signal=True,
                    ml_signal=ml_signal, ml_confidence=ml_confidence,
                )
                if market_analysis.quality == MarketQuality.AVOID:
                    stats.avoided_signals += 1
                    continue
            except Exception:
                pass

            # ═══ SMC Signal ═══
            smc_signal = None
            try:
                smc_signal = self.smc.generate_signal(df_slice)
            except Exception:
                pass

            # ═══ QM Pattern Detection ═══
            qm_patterns = self._detect_qm_patterns(df_slice)
            if qm_patterns:
                stats.qm_patterns_found += len(qm_patterns)

            # ═══ DETERMINE ENTRY ═══
            entry_source = None
            direction = None
            entry_price = None
            stop_loss = None
            take_profit = None
            confidence = 0.5
            signal_reason = ""

            if smc_signal:
                # Check if any QM pattern aligns with SMC direction
                qm_match = self._find_matching_qm(qm_patterns, current_price, smc_signal.signal_type)

                if qm_match:
                    # QM + SMC aligned: use QM's tighter SL
                    entry_source = "QM+SMC"
                    direction = smc_signal.signal_type
                    entry_price = smc_signal.entry_price

                    # Use QM head as SL (typically tighter than ATR-based)
                    qm_sl = qm_match["sl_price"]
                    smc_sl = smc_signal.stop_loss

                    # Choose tighter SL (closer to entry) but minimum $5 distance
                    if direction == "BUY":
                        qm_risk = entry_price - qm_sl
                        smc_risk = entry_price - smc_sl
                        if qm_risk > 5 and qm_risk < smc_risk:
                            stop_loss = qm_sl
                        else:
                            stop_loss = smc_sl
                    else:
                        qm_risk = qm_sl - entry_price
                        smc_risk = smc_sl - entry_price
                        if qm_risk > 5 and qm_risk < smc_risk:
                            stop_loss = qm_sl
                        else:
                            stop_loss = smc_sl

                    # Recalculate TP with QM RR ratio
                    risk = abs(entry_price - stop_loss)
                    if direction == "BUY":
                        take_profit = entry_price + risk * self.qm_rr_ratio
                    else:
                        take_profit = entry_price - risk * self.qm_rr_ratio

                    confidence = max(smc_signal.confidence, 0.65)
                    signal_reason = f"QM+SMC: {smc_signal.reason} | QML={qm_match['qml_level']:.2f}"
                    stats.qm_enhanced += 1
                else:
                    # Standard SMC entry (no QM)
                    entry_source = "SMC"
                    direction = smc_signal.signal_type
                    entry_price = smc_signal.entry_price
                    stop_loss = smc_signal.stop_loss
                    take_profit = smc_signal.take_profit
                    confidence = smc_signal.confidence
                    signal_reason = smc_signal.reason
                    stats.standard_smc += 1

            elif qm_patterns:
                # No SMC signal, but check for QM-only entry
                qm_match = self._find_matching_qm(qm_patterns, current_price)
                if qm_match:
                    entry_source = "QM-only"
                    direction = qm_match["direction"]
                    entry_price = current_price
                    stop_loss = qm_match["sl_price"]
                    risk = abs(entry_price - stop_loss)

                    # Minimum risk $5
                    if risk < 5:
                        continue

                    if direction == "BUY":
                        take_profit = entry_price + risk * self.qm_rr_ratio
                    else:
                        take_profit = entry_price - risk * self.qm_rr_ratio

                    confidence = 0.58  # Moderate confidence for QM-only
                    signal_reason = f"QM-only: QML={qm_match['qml_level']:.2f}, head={qm_match['head_level']:.2f}"
                    stats.qm_only += 1

            if entry_source is None:
                continue

            # SMC details (for tracking)
            recent_df = df_slice.tail(10)
            recent_bos = recent_df["bos"].to_list() if "bos" in df_slice.columns else []
            recent_choch = recent_df["choch"].to_list() if "choch" in df_slice.columns else []
            recent_fvg_bull = recent_df["is_fvg_bull"].to_list() if "is_fvg_bull" in df_slice.columns else []
            recent_fvg_bear = recent_df["is_fvg_bear"].to_list() if "is_fvg_bear" in df_slice.columns else []
            recent_obs = recent_df["ob"].to_list() if "ob" in df_slice.columns else []
            has_bos = 1 in recent_bos or -1 in recent_bos
            has_choch = 1 in recent_choch or -1 in recent_choch
            has_fvg = any(recent_fvg_bull) or any(recent_fvg_bear)
            has_ob = 1 in recent_obs or -1 in recent_obs

            atr_at_entry = 12.0
            if "atr" in df_slice.columns:
                atr_val = df_slice.tail(1)["atr"].item()
                if atr_val and atr_val > 0: atr_at_entry = atr_val

            # ML confidence adjustment
            ml_agrees = (direction == "BUY" and ml_signal == "BUY") or \
                        (direction == "SELL" and ml_signal == "SELL")
            if ml_agrees and entry_source != "QM-only":
                confidence = (confidence + ml_confidence) / 2
            if regime == "high_volatility":
                confidence *= 0.9

            lot_size = self._calculate_lot_size(confidence, regime, trading_mode, lot_mult)
            if lot_size <= 0:
                continue

            if trading_mode == TradingMode.RECOVERY:
                stats.recovery_mode_trades += 1

            risk = abs(entry_price - stop_loss)
            rr = abs(take_profit - entry_price) / risk if risk > 0 else 0

            profit, pips, exit_reason, exit_idx, exit_price = self._simulate_trade_exit(
                df=df, entry_idx=i, direction=direction,
                entry_price=entry_price, take_profit=take_profit,
                stop_loss=stop_loss, lot_size=lot_size,
                daily_loss_so_far=daily_loss, feature_cols=feature_cols,
            )

            self._ticket_counter += 1
            result = TradeResult.WIN if profit > 0 else (TradeResult.LOSS if profit < 0 else TradeResult.BREAKEVEN)

            trade = SimulatedTrade(
                ticket=self._ticket_counter, entry_time=current_time,
                exit_time=times[exit_idx] if exit_idx < len(times) else times[-1],
                direction=direction, entry_price=entry_price,
                exit_price=exit_price, stop_loss=stop_loss,
                take_profit=take_profit, lot_size=lot_size,
                profit_usd=profit, profit_pips=pips, result=result,
                exit_reason=exit_reason, smc_confidence=confidence,
                regime=regime, session=session_name, signal_reason=signal_reason,
                has_bos=has_bos, has_choch=has_choch, has_fvg=has_fvg,
                has_ob=has_ob, atr_at_entry=atr_at_entry, rr_ratio=rr,
                trading_mode=trading_mode.value, entry_source=entry_source,
            )
            stats.trades.append(trade)

            stats.total_trades += 1
            capital += profit

            if profit > 0:
                stats.wins += 1
                stats.total_profit += profit
                daily_profit += profit
                consecutive_losses = 0
                if trading_mode == TradingMode.RECOVERY:
                    trading_mode = TradingMode.NORMAL
            else:
                stats.losses += 1
                stats.total_loss += abs(profit)
                daily_loss += abs(profit)
                consecutive_losses += 1

            if daily_loss >= self.max_daily_loss_usd:
                trading_mode = TradingMode.STOPPED
                stats.daily_limit_stops += 1
            elif consecutive_losses >= 3 or daily_loss >= self.max_daily_loss_usd * 0.6:
                trading_mode = TradingMode.PROTECTED
            elif consecutive_losses >= 2:
                trading_mode = TradingMode.RECOVERY

            if capital > peak_capital: peak_capital = capital
            dd_pct = (peak_capital - capital) / peak_capital * 100
            dd_usd = peak_capital - capital
            if dd_pct > stats.max_drawdown:
                stats.max_drawdown = dd_pct
                stats.max_drawdown_usd = dd_usd

            stats.equity_curve.append(capital)
            last_trade_idx = exit_idx

            if stats.total_trades % 100 == 0:
                print(f"  {stats.total_trades} trades processed...")

        # Final stats
        if stats.total_trades > 0:
            stats.win_rate = stats.wins / stats.total_trades * 100
            stats.avg_win = stats.total_profit / stats.wins if stats.wins > 0 else 0
            stats.avg_loss = stats.total_loss / stats.losses if stats.losses > 0 else 0
            stats.avg_trade = (stats.total_profit - stats.total_loss) / stats.total_trades
            stats.profit_factor = stats.total_profit / stats.total_loss if stats.total_loss > 0 else float("inf")
            wp = stats.wins / stats.total_trades
            lp = stats.losses / stats.total_trades
            stats.expectancy = (wp * stats.avg_win) - (lp * stats.avg_loss)
            rets = [t.profit_usd for t in stats.trades]
            if len(rets) > 1:
                stats.sharpe_ratio = (np.mean(rets) / np.std(rets)) * np.sqrt(252) if np.std(rets) > 0 else 0

        return stats


# ─── Main ──────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("XAUBOT AI — #16 SMC + RTM Quasimodo Pattern")
    print("Base: SMC-Only v4 | Added: QM pattern for entry + SL improvement")
    print("=" * 70)

    config = get_config()
    mt5 = MT5Connector(
        login=config.mt5_login, password=config.mt5_password,
        server=config.mt5_server, path=config.mt5_path,
    )
    mt5.connect()
    print(f"\nConnected to MT5")

    print("Fetching XAUUSD M15 data...")
    df = mt5.get_market_data(symbol="XAUUSD", timeframe="M15", count=50000)
    if len(df) == 0:
        print("ERROR: No data")
        mt5.disconnect()
        return

    print(f"  Received {len(df)} bars")
    times = df["time"].to_list()
    print(f"  Data range: {times[0]} to {times[-1]}")

    end_date = datetime.now()
    start_date = datetime(2025, 8, 1)
    data_start = times[0]
    if hasattr(data_start, 'replace') and data_start.tzinfo:
        start_date = start_date.replace(tzinfo=data_start.tzinfo)
        end_date = end_date.replace(tzinfo=data_start.tzinfo)
    if data_start > start_date:
        start_date = data_start + timedelta(days=5)

    print(f"\n  Backtest period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

    print("\nCalculating indicators...")
    features = FeatureEngineer()
    smc = SMCAnalyzer(swing_length=config.smc.swing_length, ob_lookback=config.smc.ob_lookback)
    df = features.calculate_all(df, include_ml_features=True)
    df = smc.calculate_all(df)

    regime_detector = MarketRegimeDetector(model_path="models/hmm_regime.pkl")
    try:
        regime_detector.load()
        df = regime_detector.predict(df)
        print("  HMM regime loaded")
    except Exception:
        print("  [WARN] HMM not available")
    print("  Indicators calculated")

    bt = QuasimodoBacktest(
        capital=5000.0,
        max_daily_loss_percent=5.0,
        max_loss_per_trade_percent=1.0,
        base_lot_size=0.01,
        max_lot_size=0.02,
        recovery_lot_size=0.01,
        breakeven_pips=30.0,
        trail_start_pips=50.0,
        trail_step_pips=30.0,
        min_profit_to_protect=5.0,
        max_drawdown_from_peak=50.0,
        trade_cooldown_bars=10,
        trend_reversal_mult=0.6,
        qm_lookback=60,
        qm_max_age=30,
        qm_zone_tolerance_pct=0.004,
        qm_rr_ratio=2.0,
    )

    stats = bt.run(df=df, start_date=start_date, end_date=end_date, initial_capital=5000.0)
    net_pnl = stats.total_profit - stats.total_loss

    print("\n" + "=" * 70)
    print("#16 SMC + QUASIMODO — RESULTS")
    print("=" * 70)

    print(f"\n  QM Pattern Stats:")
    print(f"    QM patterns found:  {stats.qm_patterns_found}")
    print(f"    QM+SMC entries:     {stats.qm_enhanced} (QM SL used)")
    print(f"    QM-only entries:    {stats.qm_only}")
    print(f"    Standard SMC:       {stats.standard_smc}")

    # Per-source breakdown
    for src in ["SMC", "QM+SMC", "QM-only"]:
        src_trades = [t for t in stats.trades if t.entry_source == src]
        if src_trades:
            sw = sum(1 for t in src_trades if t.result == TradeResult.WIN)
            sp = sum(t.profit_usd for t in src_trades)
            swr = sw / len(src_trades) * 100
            print(f"    {src:10s}: {len(src_trades):3d} trades, {swr:.1f}% WR, ${sp:,.2f}")

    print(f"\n  Performance:")
    print(f"    Total Trades:     {stats.total_trades}")
    print(f"    Wins:             {stats.wins}")
    print(f"    Losses:           {stats.losses}")
    print(f"    Win Rate:         {stats.win_rate:.1f}%")

    print(f"\n  Profit/Loss:")
    print(f"    Total Profit:     ${stats.total_profit:,.2f}")
    print(f"    Total Loss:       ${stats.total_loss:,.2f}")
    print(f"    Net PnL:          ${net_pnl:,.2f}")
    print(f"    Profit Factor:    {stats.profit_factor:.2f}")

    print(f"\n  Risk Metrics:")
    print(f"    Max Drawdown:     {stats.max_drawdown:.1f}% (${stats.max_drawdown_usd:,.2f})")
    print(f"    Avg Win:          ${stats.avg_win:,.2f}")
    print(f"    Avg Loss:         ${stats.avg_loss:,.2f}")
    print(f"    Expectancy:       ${stats.expectancy:,.2f}")
    print(f"    Sharpe Ratio:     {stats.sharpe_ratio:.2f}")

    print(f"\n  vs BASELINE (#1): ${net_pnl - 1449.86:+,.2f}")

    print(f"\n  Exit Reasons:")
    exit_counts = {}
    for t in stats.trades:
        r = t.exit_reason.value
        exit_counts[r] = exit_counts.get(r, 0) + 1
    for reason, count in sorted(exit_counts.items(), key=lambda x: -x[1]):
        pct = count / stats.total_trades * 100 if stats.total_trades > 0 else 0
        print(f"    {reason:20s}: {count} ({pct:.1f}%)")

    print(f"\n  Direction:")
    for d in ["BUY", "SELL"]:
        dt = [t for t in stats.trades if t.direction == d]
        dw = sum(1 for t in dt if t.result == TradeResult.WIN)
        dp = sum(t.profit_usd for t in dt)
        dwr = dw / len(dt) * 100 if dt else 0
        print(f"    {d}: {len(dt)} trades, {dwr:.1f}% WR, ${dp:,.2f}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "16_quasimodo_results")
    os.makedirs(output_dir, exist_ok=True)

    # Log
    log_path = os.path.join(output_dir, f"quasimodo_{timestamp}.log")
    lines = []
    lines.append("=" * 80)
    lines.append("#16 SMC + RTM Quasimodo — Backtest Log")
    lines.append("=" * 80)
    lines.append(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
    lines.append(f"QM params: lookback=60, max_age=30, tolerance=0.4%, RR=1:2")
    lines.append(f"")
    lines.append(f"QM patterns: {stats.qm_patterns_found} | QM+SMC: {stats.qm_enhanced} | QM-only: {stats.qm_only} | Standard: {stats.standard_smc}")
    lines.append(f"Trades: {stats.total_trades} | WR: {stats.win_rate:.1f}% | Net: ${net_pnl:,.2f}")
    lines.append(f"PF: {stats.profit_factor:.2f} | DD: {stats.max_drawdown:.1f}% | Sharpe: {stats.sharpe_ratio:.2f}")
    lines.append(f"vs Baseline: ${net_pnl - 1449.86:+,.2f}")
    lines.append("")
    lines.append("--- TRADE LOG ---")
    lines.append(f"{'#':>4} {'Entry Time':>16} {'Dir':>4} {'Entry':>10} {'Exit':>10} {'P/L($)':>8} {'Result':>6} {'Exit Reason':>18} {'Source':>10}")
    lines.append("-" * 110)
    for idx, t in enumerate(stats.trades, 1):
        lines.append(
            f"{idx:4d} {t.entry_time.strftime('%Y-%m-%d %H:%M'):>16} {t.direction:>4} "
            f"{t.entry_price:>10.2f} {t.exit_price:>10.2f} {t.profit_usd:>8.2f} "
            f"{t.result.value:>6} {t.exit_reason.value:>18} {t.entry_source:>10}"
        )
    with open(log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Log saved: {log_path}")

    # XLSX
    xlsx_path = os.path.join(output_dir, f"quasimodo_{timestamp}.xlsx")
    wb = Workbook()
    hf = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    hfl = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    wf = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    lf = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bd = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "#16 SMC + Quasimodo Report"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    r = 3
    for lbl, val in [
        ("Trades", stats.total_trades), ("WR", f"{stats.win_rate:.1f}%"),
        ("Net PnL", f"${net_pnl:,.2f}"), ("PF", f"{stats.profit_factor:.2f}"),
        ("Max DD", f"{stats.max_drawdown:.1f}%"), ("Sharpe", f"{stats.sharpe_ratio:.2f}"),
        ("Avg Win", f"${stats.avg_win:,.2f}"), ("Avg Loss", f"${stats.avg_loss:,.2f}"),
        ("QM+SMC", stats.qm_enhanced), ("QM-only", stats.qm_only), ("Standard SMC", stats.standard_smc),
    ]:
        ws.cell(row=r, column=1, value=lbl)
        ws.cell(row=r, column=2, value=val)
        r += 1

    ws2 = wb.create_sheet("Trade Log")
    headers = ["#", "Entry Time", "Dir", "Entry", "Exit", "SL", "TP", "Lot", "P/L", "Result", "Exit", "Source", "RR"]
    for c, h in enumerate(headers, 1):
        ws2.cell(row=1, column=c, value=h).font = hf
        ws2.cell(row=1, column=c).fill = hfl
    for ri, t in enumerate(stats.trades, 2):
        vals = [ri-1, t.entry_time.strftime("%Y-%m-%d %H:%M"), t.direction, t.entry_price,
                t.exit_price, t.stop_loss, t.take_profit, t.lot_size,
                round(t.profit_usd, 2), t.result.value, t.exit_reason.value,
                t.entry_source, round(t.rr_ratio, 2)]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.border = bd
            if ci == 9 and isinstance(v, (int, float)):
                cell.fill = wf if v > 0 else (lf if v < 0 else PatternFill())

    wb.save(xlsx_path)
    print(f"  Report saved: {xlsx_path}")

    mt5.disconnect()
    print("\n" + "=" * 70)
    print(f"Output: {output_dir}")
    print("=" * 70)
    print("Backtest complete!")


if __name__ == "__main__":
    main()
