"""
Backtest B: SMC + H4 Zone Filter + Tighter SL
===============================================
Base: SMC-Only v4 (100% synced with main_live.py)
Added: H4 Multi-Timeframe Zone Filter + Tighter SL using H4 zone boundary

Logic:
  - Same H4 zone filter as Backtest A
  - SL CHANGED: Use H4 zone boundary for tighter SL instead of swing low + 1.5x ATR
    * BUY: SL = H4 demand zone bottom - small buffer (instead of M15 swing low)
    * SELL: SL = H4 supply zone top + small buffer
  - Minimum SL: 0.5x ATR (prevent too-tight SL)
  - TP adjusted: RR 1:2 (instead of 1:1.5) since SL is tighter

Exit: ALL 3 systems unchanged

Usage:
    python backtests/backtest_h4_zone_tight_sl.py
"""

import polars as pl
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, field
from enum import Enum
import sys
import os
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.mt5_connector import MT5Connector
from src.smc_polars import SMCAnalyzer, SMCSignal
from src.feature_eng import FeatureEngineer
from src.regime_detector import MarketRegimeDetector, MarketRegime
from src.ml_model import TradingModel
from src.config import get_config
from src.dynamic_confidence import DynamicConfidenceManager, create_dynamic_confidence, MarketQuality
from loguru import logger

logger.remove()
logger.add(sys.stderr, level="WARNING")

WIB = ZoneInfo("Asia/Jakarta")

# H4 zone tolerance (±1.5% price deviation for zone matching ~$42 at $2800)
# H4 zones are narrow ($5-20 wide), need wider tolerance for practical matching
H4_ZONE_TOLERANCE = 0.015
# Tighter SL: minimum distance = 0.5x ATR
MIN_SL_ATR_MULT = 0.5
# Tighter SL target RR = 1:2 (instead of baseline 1:1.5)
TIGHT_SL_RR = 2.0
# SL buffer beyond zone boundary (in price points, ~$2)
SL_ZONE_BUFFER = 2.0


# ─── Enums & Dataclasses ──────────────────────────────────────

class TradeResult(Enum):
    WIN = "WIN"
    LOSS = "LOSS"
    BREAKEVEN = "BREAKEVEN"

class ExitReason(Enum):
    TAKE_PROFIT = "take_profit"
    SMART_TP = "smart_tp"
    PEAK_PROTECT = "peak_protect"
    EARLY_EXIT = "early_exit"
    EARLY_CUT = "early_cut"
    MAX_LOSS = "max_loss"
    STALL = "stall"
    TREND_REVERSAL = "trend_reversal"
    TIMEOUT = "timeout"
    WEEKEND_CLOSE = "weekend_close"
    TRAILING_SL = "trailing_sl"
    BREAKEVEN_EXIT = "breakeven_exit"
    DAILY_LIMIT = "daily_limit"
    REGIME_DANGER = "regime_danger"
    MARKET_SIGNAL = "market_signal"

class TradingMode(Enum):
    NORMAL = "normal"
    RECOVERY = "recovery"
    PROTECTED = "protected"
    STOPPED = "stopped"

@dataclass
class SimulatedTrade:
    ticket: int
    entry_time: datetime
    exit_time: datetime
    direction: str
    entry_price: float
    exit_price: float
    stop_loss: float
    take_profit: float
    lot_size: float
    profit_usd: float
    profit_pips: float
    result: TradeResult
    exit_reason: ExitReason
    smc_confidence: float
    regime: str
    session: str
    signal_reason: str
    has_bos: bool = False
    has_choch: bool = False
    has_fvg: bool = False
    has_ob: bool = False
    atr_at_entry: float = 0.0
    rr_ratio: float = 0.0
    trading_mode: str = "normal"
    h4_zone_type: str = "none"
    sl_type: str = "baseline"  # "baseline", "h4_zone", "m15_ob"
    original_sl: float = 0.0   # baseline SL for comparison

@dataclass
class BacktestStats:
    total_trades: int = 0
    wins: int = 0
    losses: int = 0
    total_profit: float = 0.0
    total_loss: float = 0.0
    max_drawdown: float = 0.0
    max_drawdown_usd: float = 0.0
    win_rate: float = 0.0
    profit_factor: float = 0.0
    avg_win: float = 0.0
    avg_loss: float = 0.0
    avg_trade: float = 0.0
    expectancy: float = 0.0
    sharpe_ratio: float = 0.0
    trades: List[SimulatedTrade] = field(default_factory=list)
    equity_curve: List[float] = field(default_factory=list)
    avoided_signals: int = 0
    daily_limit_stops: int = 0
    recovery_mode_trades: int = 0
    # H4 zone filter stats
    h4_filtered: int = 0
    h4_filtered_buy: int = 0
    h4_filtered_sell: int = 0
    h4_zone_ob_trades: int = 0
    h4_zone_fvg_trades: int = 0
    # Tight SL stats
    tight_sl_used: int = 0
    baseline_sl_used: int = 0
    avg_sl_distance_tight: float = 0.0
    avg_sl_distance_baseline: float = 0.0


# ─── H4 Zone Helper ──────────────────────────────────────────

def extract_h4_zones(df_h4: pl.DataFrame, current_m15_time) -> Dict:
    """
    Extract active H4 OB and FVG zones from H4 data.
    Only use H4 candles that have CLOSED before current M15 time.
    """
    zones = {
        "bullish_obs": [],
        "bearish_obs": [],
        "bullish_fvgs": [],
        "bearish_fvgs": [],
    }

    h4_times = df_h4["time"].to_list()
    h4_obs = df_h4["ob"].to_list()
    h4_ob_tops = df_h4["ob_top"].to_list()
    h4_ob_bottoms = df_h4["ob_bottom"].to_list()
    h4_fvg_bulls = df_h4["is_fvg_bull"].to_list()
    h4_fvg_bears = df_h4["is_fvg_bear"].to_list()
    h4_fvg_tops = df_h4["fvg_top"].to_list()
    h4_fvg_bottoms = df_h4["fvg_bottom"].to_list()
    h4_closes = df_h4["close"].to_list()
    h4_highs = df_h4["high"].to_list()
    h4_lows = df_h4["low"].to_list()

    n = len(df_h4)

    # Scan last 50 H4 candles (~8 days) for active zones
    start = max(0, n - 50)
    for i in range(start, n):
        if h4_times[i] >= current_m15_time:
            break

        # Order Blocks — zone invalid only if price BROKE THROUGH (not just touched)
        if h4_obs[i] == 1 and h4_ob_tops[i] is not None:
            invalidated = False
            for j in range(i + 1, min(i + 20, n)):
                if h4_times[j] >= current_m15_time:
                    break
                # Bullish OB invalid if price broke BELOW zone bottom
                if h4_closes[j] < h4_ob_bottoms[i]:
                    invalidated = True
                    break
            if not invalidated:
                zones["bullish_obs"].append({
                    "top": h4_ob_tops[i],
                    "bottom": h4_ob_bottoms[i],
                    "time": h4_times[i],
                })

        if h4_obs[i] == -1 and h4_ob_tops[i] is not None:
            invalidated = False
            for j in range(i + 1, min(i + 20, n)):
                if h4_times[j] >= current_m15_time:
                    break
                # Bearish OB invalid if price broke ABOVE zone top
                if h4_closes[j] > h4_ob_tops[i]:
                    invalidated = True
                    break
            if not invalidated:
                zones["bearish_obs"].append({
                    "top": h4_ob_tops[i],
                    "bottom": h4_ob_bottoms[i],
                    "time": h4_times[i],
                })

        # FVGs — invalid only if price CLOSED beyond the gap (fully filled)
        if h4_fvg_bulls[i] and h4_fvg_tops[i] is not None:
            filled = False
            for j in range(i + 1, min(i + 20, n)):
                if h4_times[j] >= current_m15_time:
                    break
                # Bullish FVG filled if price closed below gap bottom
                if h4_closes[j] < h4_fvg_bottoms[i]:
                    filled = True
                    break
            if not filled:
                zones["bullish_fvgs"].append({
                    "top": h4_fvg_tops[i],
                    "bottom": h4_fvg_bottoms[i],
                    "time": h4_times[i],
                })

        if h4_fvg_bears[i] and h4_fvg_tops[i] is not None:
            filled = False
            for j in range(i + 1, min(i + 20, n)):
                if h4_times[j] >= current_m15_time:
                    break
                # Bearish FVG filled if price closed above gap top
                if h4_closes[j] > h4_fvg_tops[i]:
                    filled = True
                    break
            if not filled:
                zones["bearish_fvgs"].append({
                    "top": h4_fvg_tops[i],
                    "bottom": h4_fvg_bottoms[i],
                    "time": h4_times[i],
                })

    return zones


def is_price_in_h4_zone(price: float, direction: str, h4_zones: Dict, tolerance: float = H4_ZONE_TOLERANCE) -> Tuple[bool, str, Optional[Dict]]:
    """
    Check if price is within an active H4 zone.
    Returns (is_in_zone, zone_type, matched_zone_dict).
    """
    price_tol = price * tolerance

    if direction == "BUY":
        for ob in h4_zones.get("bullish_obs", []):
            if ob["bottom"] - price_tol <= price <= ob["top"] + price_tol:
                return True, "OB", ob
        for fvg in h4_zones.get("bullish_fvgs", []):
            if fvg["bottom"] - price_tol <= price <= fvg["top"] + price_tol:
                return True, "FVG", fvg

    elif direction == "SELL":
        for ob in h4_zones.get("bearish_obs", []):
            if ob["bottom"] - price_tol <= price <= ob["top"] + price_tol:
                return True, "OB", ob
        for fvg in h4_zones.get("bearish_fvgs", []):
            if fvg["bottom"] - price_tol <= price <= fvg["top"] + price_tol:
                return True, "FVG", fvg

    return False, "none", None


def calculate_tight_sl(entry_price: float, direction: str, matched_zone: Dict,
                       baseline_sl: float, atr: float) -> Tuple[float, str]:
    """
    Calculate tighter SL using H4 zone boundary.

    BUY: SL = zone bottom - buffer (instead of swing low - 1.5x ATR)
    SELL: SL = zone top + buffer (instead of swing high + 1.5x ATR)

    Constraints:
      - Minimum SL distance = MIN_SL_ATR_MULT * ATR
      - If tight SL is WORSE than baseline, use baseline

    Returns (new_sl, sl_type)
    """
    min_sl_distance = atr * MIN_SL_ATR_MULT

    if direction == "BUY":
        # Tight SL = below H4 demand zone bottom
        zone_sl = matched_zone["bottom"] - SL_ZONE_BUFFER

        # Ensure minimum distance
        sl_distance = entry_price - zone_sl
        if sl_distance < min_sl_distance:
            zone_sl = entry_price - min_sl_distance

        # Use tight SL only if it's TIGHTER (higher) than baseline
        if zone_sl > baseline_sl:
            return zone_sl, "h4_zone"
        else:
            return baseline_sl, "baseline"

    else:  # SELL
        # Tight SL = above H4 supply zone top
        zone_sl = matched_zone["top"] + SL_ZONE_BUFFER

        sl_distance = zone_sl - entry_price
        if sl_distance < min_sl_distance:
            zone_sl = entry_price + min_sl_distance

        # Use tight SL only if it's TIGHTER (lower) than baseline
        if zone_sl < baseline_sl:
            return zone_sl, "h4_zone"
        else:
            return baseline_sl, "baseline"


# ─── SMC + H4 Zone + Tight SL Backtest ──────────────────────

class SMCH4ZoneTightSLBacktest:
    """SMC-Only v4 + H4 Zone Filter + Tighter SL from zone boundary. All exit systems unchanged."""

    def __init__(self, capital=5000.0, max_daily_loss_percent=5.0,
                 max_loss_per_trade_percent=1.0, base_lot_size=0.01,
                 max_lot_size=0.02, recovery_lot_size=0.01,
                 trend_reversal_threshold=0.75, max_concurrent_positions=2,
                 breakeven_pips=30.0, trail_start_pips=50.0, trail_step_pips=30.0,
                 min_profit_to_protect=5.0, max_drawdown_from_peak=50.0,
                 trade_cooldown_bars=10, trend_reversal_mult=0.6):
        self.capital = capital
        self.max_daily_loss_usd = capital * (max_daily_loss_percent / 100)
        self.max_loss_per_trade = capital * (max_loss_per_trade_percent / 100)
        self.base_lot_size = base_lot_size
        self.max_lot_size = max_lot_size
        self.recovery_lot_size = recovery_lot_size
        self.trend_reversal_threshold = trend_reversal_threshold
        self.breakeven_pips = breakeven_pips
        self.trail_start_pips = trail_start_pips
        self.trail_step_pips = trail_step_pips
        self.min_profit_to_protect = min_profit_to_protect
        self.max_drawdown_from_peak = max_drawdown_from_peak
        self.trade_cooldown_bars = trade_cooldown_bars
        self.trend_reversal_mult = trend_reversal_mult

        config = get_config()
        self.smc = SMCAnalyzer(swing_length=config.smc.swing_length, ob_lookback=config.smc.ob_lookback)
        self.features = FeatureEngineer()
        self.dynamic_confidence = create_dynamic_confidence()

        self.ml_model = TradingModel(model_path="models/xgboost_model.pkl")
        try:
            self.ml_model.load()
            print("  ML model loaded (for exit evaluation)")
        except Exception:
            print("  [WARN] ML model not loaded")

        self.regime_detector = MarketRegimeDetector(model_path="models/hmm_regime.pkl")
        try:
            self.regime_detector.load()
        except Exception:
            print("  [WARN] HMM model not loaded")

        self._ticket_counter = 3000000

    def _get_session_from_time(self, dt):
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib_time = dt.astimezone(WIB)
        hour = wib_time.hour
        if 6 <= hour < 15: return "Sydney-Tokyo", True, 0.5
        elif 15 <= hour < 16: return "Tokyo-London Overlap", True, 0.75
        elif 16 <= hour < 19: return "London Early", True, 0.8
        elif 19 <= hour < 24: return "London-NY Overlap (Golden)", True, 1.0
        elif 0 <= hour < 4: return "NY Session", True, 0.9
        else: return "Off Hours", False, 0.0

    def _hours_to_golden(self, dt):
        if dt.tzinfo is None: dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib = dt.astimezone(WIB)
        if 19 <= wib.hour < 24: return 0
        target = wib.replace(hour=19, minute=0, second=0, microsecond=0)
        if wib.hour >= 19: target += timedelta(days=1)
        return max(0, (target - wib).total_seconds() / 3600)

    def _is_near_weekend_close(self, dt):
        if dt.tzinfo is None: dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib = dt.astimezone(WIB)
        return wib.weekday() == 5 and wib.hour >= 4 and wib.minute >= 30

    def _calculate_lot_size(self, confidence, regime, trading_mode, session_mult):
        if trading_mode == TradingMode.STOPPED: return 0
        lot = self.base_lot_size
        if trading_mode in (TradingMode.RECOVERY, TradingMode.PROTECTED):
            lot = self.recovery_lot_size
        else:
            if confidence >= 0.65: lot = self.max_lot_size
            elif confidence >= 0.55: lot = self.base_lot_size
            else: lot = self.recovery_lot_size
        if regime.lower() in ["high_volatility", "crisis"]:
            lot = self.recovery_lot_size
        return round(max(0.01, lot * session_mult), 2)

    # ── Full exit simulation (ALL 3 systems — identical to baseline) ──
    def _simulate_trade_exit(self, df, entry_idx, direction, entry_price, take_profit, stop_loss, lot_size, daily_loss_so_far, feature_cols, max_bars=100):
        pip_value = 10
        highs = df["high"].to_list()
        lows = df["low"].to_list()
        closes = df["close"].to_list()
        times = df["time"].to_list()
        atr = 12.0
        if "atr" in df.columns:
            atr_list = df["atr"].to_list()
            if entry_idx < len(atr_list) and atr_list[entry_idx] is not None:
                atr = atr_list[entry_idx]
        reversal_momentum_threshold = atr * self.trend_reversal_mult
        min_loss_for_reversal_exit = atr * 0.8
        profit_history, price_history = [], []
        peak_profit, stall_count, reversal_warnings = 0.0, 0, 0
        current_sl, breakeven_moved = stop_loss, False
        if direction == "BUY":
            target_tp_profit = (take_profit - entry_price) / 0.1 * pip_value * lot_size
        else:
            target_tp_profit = (entry_price - take_profit) / 0.1 * pip_value * lot_size
        cached_ml_signal, cached_ml_confidence = "", 0.5

        for i in range(entry_idx + 1, min(entry_idx + max_bars, len(df))):
            high, low, close, current_time = highs[i], lows[i], closes[i], times[i]
            if direction == "BUY":
                current_pips = (close - entry_price) / 0.1
                pip_profit_from_entry = current_pips
            else:
                current_pips = (entry_price - close) / 0.1
                pip_profit_from_entry = current_pips
            current_profit = current_pips * pip_value * lot_size
            profit_history.append(current_profit)
            price_history.append(close)
            if current_profit > peak_profit: peak_profit = current_profit
            bars_since_entry = i - entry_idx

            if bars_since_entry % 4 == 0 and self.ml_model.fitted:
                try:
                    ml_pred = self.ml_model.predict(df.head(i + 1), feature_cols)
                    cached_ml_signal, cached_ml_confidence = ml_pred.signal, ml_pred.confidence
                except Exception: pass

            momentum = 0.0
            if len(profit_history) >= 3:
                recent = profit_history[-5:] if len(profit_history) >= 5 else profit_history
                momentum = max(-100, min(100, ((recent[-1] - recent[0]) / 10) * 50))
            profit_growing = momentum > 0

            # A) SmartPositionManager
            if direction == "BUY" and high >= take_profit:
                pips = (take_profit - entry_price) / 0.1
                return pips * pip_value * lot_size, pips, ExitReason.TAKE_PROFIT, i, take_profit
            elif direction == "SELL" and low <= take_profit:
                pips = (entry_price - take_profit) / 0.1
                return pips * pip_value * lot_size, pips, ExitReason.TAKE_PROFIT, i, take_profit

            if breakeven_moved and current_sl > 0:
                if direction == "BUY" and low <= current_sl:
                    pips = (current_sl - entry_price) / 0.1
                    reason = ExitReason.TRAILING_SL if pip_profit_from_entry >= self.trail_start_pips else ExitReason.BREAKEVEN_EXIT
                    return pips * pip_value * lot_size, pips, reason, i, current_sl
                elif direction == "SELL" and high >= current_sl:
                    pips = (entry_price - current_sl) / 0.1
                    reason = ExitReason.TRAILING_SL if pip_profit_from_entry >= self.trail_start_pips else ExitReason.BREAKEVEN_EXIT
                    return pips * pip_value * lot_size, pips, reason, i, current_sl

            if pip_profit_from_entry >= self.breakeven_pips and not breakeven_moved:
                current_sl = entry_price + 2 if direction == "BUY" else entry_price - 2
                breakeven_moved = True
            if pip_profit_from_entry >= self.trail_start_pips:
                trail_distance = self.trail_step_pips * 0.1
                if direction == "BUY":
                    new_sl = close - trail_distance
                    if new_sl > current_sl: current_sl = new_sl
                else:
                    new_sl = close + trail_distance
                    if current_sl == 0 or new_sl < current_sl: current_sl = new_sl

            if peak_profit > self.min_profit_to_protect:
                dd_pct = ((peak_profit - current_profit) / peak_profit) * 100 if peak_profit > 0 else 0
                if dd_pct > self.max_drawdown_from_peak:
                    return current_profit, current_pips, ExitReason.PEAK_PROTECT, i, close

            if bars_since_entry % 5 == 0 and bars_since_entry >= 5 and i >= 20:
                ma_fast = np.mean(closes[i-4:i+1])
                ma_slow = np.mean(closes[i-19:i+1])
                trend = "BULLISH" if ma_fast > ma_slow * 1.001 else ("BEARISH" if ma_fast < ma_slow * 0.999 else "NEUTRAL")
                roc = (closes[i] / closes[max(0,i-4)] - 1) * 100
                mom_dir = "BULLISH" if roc > 0.3 else ("BEARISH" if roc < -0.3 else "NEUTRAL")
                rsi_val = None
                if "rsi" in df.columns:
                    rsi_list = df["rsi"].to_list()
                    if i < len(rsi_list): rsi_val = rsi_list[i]
                urgency, should_exit = 0, False
                if cached_ml_confidence > 0.75:
                    if (direction == "BUY" and cached_ml_signal == "SELL") or (direction == "SELL" and cached_ml_signal == "BUY"):
                        should_exit, urgency = True, urgency + 2
                if rsi_val:
                    if (rsi_val > 75 and direction == "BUY") or (rsi_val < 25 and direction == "SELL"):
                        should_exit, urgency = True, urgency + 2
                if (direction == "BUY" and trend == "BEARISH" and mom_dir == "BEARISH") or \
                   (direction == "SELL" and trend == "BULLISH" and mom_dir == "BULLISH"):
                    should_exit, urgency = True, urgency + 3
                if should_exit and current_profit > self.min_profit_to_protect / 2:
                    return current_profit, current_pips, ExitReason.MARKET_SIGNAL, i, close
                if urgency >= 7 and current_profit > 0:
                    return current_profit, current_pips, ExitReason.MARKET_SIGNAL, i, close

            if self._is_near_weekend_close(current_time):
                if current_profit > -10:
                    return current_profit, current_pips, ExitReason.WEEKEND_CLOSE, i, close

            # B) SmartRiskManager
            if current_profit >= 15:
                if current_profit >= 40: return current_profit, current_pips, ExitReason.SMART_TP, i, close
                if current_profit >= 25 and momentum < -30: return current_profit, current_pips, ExitReason.SMART_TP, i, close
                if peak_profit > 30 and current_profit < peak_profit * 0.6: return current_profit, current_pips, ExitReason.PEAK_PROTECT, i, close
                if current_profit >= 20:
                    progress = (current_profit / target_tp_profit) * 100 if target_tp_profit > 0 else 0
                    tp_prob = min(40, max(0, progress * 0.4)) + ((momentum + 100) / 200) * 30 + 10 - min(10, bars_since_entry / 4 * 2)
                    if tp_prob < 25: return current_profit, current_pips, ExitReason.SMART_TP, i, close
            if 5 <= current_profit < 15:
                if momentum < -50 and cached_ml_confidence >= 0.65:
                    if (direction == "BUY" and cached_ml_signal == "SELL") or (direction == "SELL" and cached_ml_signal == "BUY"):
                        return current_profit, current_pips, ExitReason.EARLY_EXIT, i, close
            if current_profit < 0:
                loss_pct = abs(current_profit) / self.max_loss_per_trade * 100
                if momentum < -30 and loss_pct >= 30:
                    return current_profit, current_pips, ExitReason.EARLY_CUT, i, close

            is_ml_rev = False
            if (direction == "BUY" and cached_ml_signal == "SELL" and cached_ml_confidence >= self.trend_reversal_threshold) or \
               (direction == "SELL" and cached_ml_signal == "BUY" and cached_ml_confidence >= self.trend_reversal_threshold):
                is_ml_rev = True
                reversal_warnings += 1
            if is_ml_rev and current_profit < -8 and abs(current_profit) > self.max_loss_per_trade * 0.4:
                return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close
            if reversal_warnings >= 3 and current_profit < -10:
                return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close
            if current_profit <= -(self.max_loss_per_trade * 0.50):
                htg = self._hours_to_golden(current_time)
                if not (htg <= 1 and htg > 0 and momentum > -40):
                    return current_profit, current_pips, ExitReason.MAX_LOSS, i, close
            if len(profit_history) >= 10:
                if max(profit_history[-10:]) - min(profit_history[-10:]) < 3 and current_profit < -15:
                    stall_count += 1
                    if stall_count >= 5: return current_profit, current_pips, ExitReason.STALL, i, close
            if daily_loss_so_far + abs(min(0, current_profit)) >= self.max_daily_loss_usd:
                return current_profit, current_pips, ExitReason.DAILY_LIMIT, i, close

            # C) Time-based
            if bars_since_entry >= 16 and current_profit < 5 and not profit_growing and current_profit > -15:
                return current_profit, current_pips, ExitReason.TIMEOUT, i, close
            if bars_since_entry >= 24 and (current_profit < 10 or not profit_growing):
                return current_profit, current_pips, ExitReason.TIMEOUT, i, close
            if bars_since_entry >= 32:
                return current_profit, current_pips, ExitReason.TIMEOUT, i, close
            if bars_since_entry > 10:
                mom = closes[i] - closes[i-5]
                if direction == "BUY" and mom < -reversal_momentum_threshold and current_profit < -min_loss_for_reversal_exit:
                    return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close
                elif direction == "SELL" and mom > reversal_momentum_threshold and current_profit < -min_loss_for_reversal_exit:
                    return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close

        final_idx = min(entry_idx + max_bars - 1, len(df) - 1)
        fp = closes[final_idx]
        pips = (fp - entry_price) / 0.1 if direction == "BUY" else (entry_price - fp) / 0.1
        return pips * pip_value * lot_size, pips, ExitReason.TIMEOUT, final_idx, fp

    # ── Main backtest run ──
    def run(self, df_m15, df_h4, start_date=None, end_date=None, initial_capital=5000.0):
        stats = BacktestStats()
        capital = initial_capital
        peak_capital = initial_capital
        stats.equity_curve.append(capital)

        daily_loss, daily_profit, daily_trades = 0.0, 0.0, 0
        consecutive_losses = 0
        trading_mode = TradingMode.NORMAL
        current_date = None

        feature_cols = []
        if self.ml_model.fitted and self.ml_model.feature_names:
            feature_cols = [f for f in self.ml_model.feature_names if f in df_m15.columns]

        times = df_m15["time"].to_list()
        start_idx = next((i for i, t in enumerate(times) if t >= start_date), 100) if start_date else 100
        end_idx = next((i for i, t in enumerate(times) if t > end_date), len(df_m15) - 100) if end_date else len(df_m15) - 100

        last_trade_idx = -self.trade_cooldown_bars * 2

        # Cache H4 zones
        cached_h4_zones = None
        cached_h4_bar = -100

        # Track SL distances for stats
        tight_sl_distances = []
        baseline_sl_distances = []

        print(f"\n  Running SMC + H4 Zone + Tight SL backtest...")
        print(f"  H4 zones: OB + FVG (unmitigated/unfilled only)")
        print(f"  Zone tolerance: ±{H4_ZONE_TOLERANCE*100:.2f}%")
        print(f"  SL: H4 zone boundary (min {MIN_SL_ATR_MULT}x ATR)")
        print(f"  TP: RR 1:{TIGHT_SL_RR}")
        print(f"  SL buffer: ${SL_ZONE_BUFFER}")
        print(f"  Date range: {times[start_idx]} to {times[end_idx - 1]}")
        print(f"  Total bars: {end_idx - start_idx}")

        for i in range(start_idx, end_idx):
            if i - last_trade_idx < self.trade_cooldown_bars:
                continue

            current_time = times[i]

            # Daily reset
            trade_date = current_time.date() if hasattr(current_time, 'date') else current_time
            if current_date is None or trade_date != current_date:
                daily_loss, daily_profit, daily_trades = 0.0, 0.0, 0
                current_date = trade_date
                if consecutive_losses < 2: trading_mode = TradingMode.NORMAL
            if trading_mode == TradingMode.STOPPED: continue

            session_name, can_trade, lot_mult = self._get_session_from_time(current_time)
            if not can_trade: continue
            if hasattr(current_time, 'weekday') and current_time.weekday() >= 5: continue

            df_slice = df_m15.head(i + 1)

            # Regime check
            regime = "normal"
            try:
                if self.regime_detector.fitted:
                    regime_state = self.regime_detector.get_current_state(df_slice)
                    if regime_state:
                        regime = regime_state.regime.value
                        if regime_state.regime == MarketRegime.CRISIS: continue
                        if regime_state.recommendation == "SLEEP": continue
            except Exception: pass

            # DynamicConfidence AVOID
            ml_signal, ml_confidence = "", 0.5
            try:
                if self.ml_model.fitted and feature_cols:
                    ml_pred = self.ml_model.predict(df_slice, feature_cols)
                    ml_signal, ml_confidence = ml_pred.signal, ml_pred.confidence
                market_analysis = self.dynamic_confidence.analyze_market(
                    session=session_name, regime=regime, volatility="medium",
                    trend_direction=regime, has_smc_signal=True,
                    ml_signal=ml_signal, ml_confidence=ml_confidence)
                if market_analysis.quality == MarketQuality.AVOID:
                    stats.avoided_signals += 1
                    continue
            except Exception: pass

            # SMC Signal
            try:
                smc_signal = self.smc.generate_signal(df_slice)
            except Exception: continue
            if smc_signal is None: continue

            # ═══════════════════════════════════════════════════════
            # H4 ZONE FILTER — update zones every 16 bars (4h)
            # ═══════════════════════════════════════════════════════
            if i - cached_h4_bar >= 16 or cached_h4_zones is None:
                cached_h4_zones = extract_h4_zones(df_h4, current_time)
                cached_h4_bar = i

            in_zone, zone_type, matched_zone = is_price_in_h4_zone(
                smc_signal.entry_price, smc_signal.signal_type, cached_h4_zones
            )

            if not in_zone:
                stats.h4_filtered += 1
                if smc_signal.signal_type == "BUY":
                    stats.h4_filtered_buy += 1
                else:
                    stats.h4_filtered_sell += 1
                continue

            if zone_type == "OB":
                stats.h4_zone_ob_trades += 1
            elif zone_type == "FVG":
                stats.h4_zone_fvg_trades += 1
            # ═══════════════════════════════════════════════════════

            # SMC details
            recent_df = df_slice.tail(10)
            recent_bos = recent_df["bos"].to_list() if "bos" in df_slice.columns else []
            recent_choch = recent_df["choch"].to_list() if "choch" in df_slice.columns else []
            recent_fvg_bull = recent_df["is_fvg_bull"].to_list() if "is_fvg_bull" in df_slice.columns else []
            recent_fvg_bear = recent_df["is_fvg_bear"].to_list() if "is_fvg_bear" in df_slice.columns else []
            recent_obs = recent_df["ob"].to_list() if "ob" in df_slice.columns else []
            has_bos = 1 in recent_bos or -1 in recent_bos
            has_choch = 1 in recent_choch or -1 in recent_choch
            has_fvg = any(recent_fvg_bull) or any(recent_fvg_bear)
            has_ob = 1 in recent_obs or -1 in recent_obs

            atr_at_entry = 12.0
            if "atr" in df_slice.columns:
                v = df_slice.tail(1)["atr"].item()
                if v and v > 0: atr_at_entry = v

            confidence = smc_signal.confidence
            ml_agrees = (smc_signal.signal_type == "BUY" and ml_signal == "BUY") or \
                        (smc_signal.signal_type == "SELL" and ml_signal == "SELL")
            if ml_agrees: confidence = (smc_signal.confidence + ml_confidence) / 2
            if regime == "high_volatility": confidence *= 0.9

            lot_size = self._calculate_lot_size(confidence, regime, trading_mode, lot_mult)
            if lot_size <= 0: continue
            if trading_mode == TradingMode.RECOVERY: stats.recovery_mode_trades += 1

            entry_price = smc_signal.entry_price
            baseline_sl = smc_signal.stop_loss
            baseline_tp = smc_signal.take_profit

            # ═══════════════════════════════════════════════════════
            # TIGHT SL — Use H4 zone boundary for tighter SL
            # ═══════════════════════════════════════════════════════
            sl, sl_type = calculate_tight_sl(
                entry_price, smc_signal.signal_type, matched_zone,
                baseline_sl, atr_at_entry
            )

            # Recalculate TP based on new SL with better RR
            risk = abs(entry_price - sl)
            if smc_signal.signal_type == "BUY":
                tp = entry_price + (risk * TIGHT_SL_RR)
            else:
                tp = entry_price - (risk * TIGHT_SL_RR)

            rr = abs(tp - entry_price) / risk if risk > 0 else 0

            # Track SL distances
            sl_distance = abs(entry_price - sl)
            baseline_sl_distance = abs(entry_price - baseline_sl)
            if sl_type == "h4_zone":
                stats.tight_sl_used += 1
                tight_sl_distances.append(sl_distance)
            else:
                stats.baseline_sl_used += 1
                baseline_sl_distances.append(sl_distance)
            # ═══════════════════════════════════════════════════════

            profit, pips, exit_reason, exit_idx, exit_price = self._simulate_trade_exit(
                df=df_m15, entry_idx=i, direction=smc_signal.signal_type,
                entry_price=entry_price, take_profit=tp, stop_loss=sl,
                lot_size=lot_size, daily_loss_so_far=daily_loss, feature_cols=feature_cols)

            self._ticket_counter += 1
            result = TradeResult.WIN if profit > 0 else (TradeResult.LOSS if profit < 0 else TradeResult.BREAKEVEN)

            trade = SimulatedTrade(
                ticket=self._ticket_counter, entry_time=current_time,
                exit_time=times[exit_idx] if exit_idx < len(times) else times[-1],
                direction=smc_signal.signal_type, entry_price=entry_price, exit_price=exit_price,
                stop_loss=sl, take_profit=tp, lot_size=lot_size,
                profit_usd=profit, profit_pips=pips, result=result, exit_reason=exit_reason,
                smc_confidence=confidence, regime=regime, session=session_name,
                signal_reason=smc_signal.reason, has_bos=has_bos, has_choch=has_choch,
                has_fvg=has_fvg, has_ob=has_ob, atr_at_entry=atr_at_entry,
                rr_ratio=rr, trading_mode=trading_mode.value, h4_zone_type=zone_type,
                sl_type=sl_type, original_sl=baseline_sl)
            stats.trades.append(trade)

            stats.total_trades += 1
            daily_trades += 1
            capital += profit

            if profit > 0:
                stats.wins += 1
                stats.total_profit += profit
                daily_profit += profit
                consecutive_losses = 0
                if trading_mode == TradingMode.RECOVERY: trading_mode = TradingMode.NORMAL
            else:
                stats.losses += 1
                stats.total_loss += abs(profit)
                daily_loss += abs(profit)
                consecutive_losses += 1

            if daily_loss >= self.max_daily_loss_usd:
                trading_mode = TradingMode.STOPPED
                stats.daily_limit_stops += 1
            elif consecutive_losses >= 3 or daily_loss >= self.max_daily_loss_usd * 0.6:
                trading_mode = TradingMode.PROTECTED
            elif consecutive_losses >= 2:
                trading_mode = TradingMode.RECOVERY

            if capital > peak_capital: peak_capital = capital
            dd_pct = (peak_capital - capital) / peak_capital * 100
            dd_usd = peak_capital - capital
            if dd_pct > stats.max_drawdown:
                stats.max_drawdown = dd_pct
                stats.max_drawdown_usd = dd_usd
            stats.equity_curve.append(capital)
            last_trade_idx = exit_idx

            if stats.total_trades % 50 == 0:
                print(f"  {stats.total_trades} trades processed...")

        if stats.total_trades > 0:
            stats.win_rate = stats.wins / stats.total_trades * 100
            stats.avg_win = stats.total_profit / stats.wins if stats.wins > 0 else 0
            stats.avg_loss = stats.total_loss / stats.losses if stats.losses > 0 else 0
            stats.avg_trade = (stats.total_profit - stats.total_loss) / stats.total_trades
            stats.profit_factor = stats.total_profit / stats.total_loss if stats.total_loss > 0 else float("inf")
            wp = stats.wins / stats.total_trades
            lp = stats.losses / stats.total_trades
            stats.expectancy = (wp * stats.avg_win) - (lp * stats.avg_loss)
            returns = [t.profit_usd for t in stats.trades]
            if len(returns) > 1:
                stats.sharpe_ratio = (np.mean(returns) / np.std(returns)) * np.sqrt(252) if np.std(returns) > 0 else 0

        # Calculate avg SL distances
        if tight_sl_distances:
            stats.avg_sl_distance_tight = np.mean(tight_sl_distances)
        if baseline_sl_distances:
            stats.avg_sl_distance_baseline = np.mean(baseline_sl_distances)

        return stats


# ─── Report & Log generators ─────────────────────────────────

def generate_xlsx_report(stats, filepath, start_date, end_date, variant_name):
    wb = Workbook()
    hf = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    sf = Font(name="Calibri", bold=True, size=10)
    sfill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    wfill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    lfill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    net_pnl = stats.total_profit - stats.total_loss

    ws = wb.active
    ws.title = "Summary"
    ws.merge_cells("A1:F1")
    ws["A1"] = f"XAUBot AI — {variant_name}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws["A2"] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"

    data = [
        ("Performance", "", True), ("Total Trades", stats.total_trades, False),
        ("Wins", stats.wins, False), ("Losses", stats.losses, False),
        ("Win Rate", f"{stats.win_rate:.1f}%", False), ("", "", False),
        ("H4 Zone Filter", "", True),
        ("  Total filtered", stats.h4_filtered, False),
        ("  BUY filtered", stats.h4_filtered_buy, False),
        ("  SELL filtered", stats.h4_filtered_sell, False),
        ("  Trades in OB zone", stats.h4_zone_ob_trades, False),
        ("  Trades in FVG zone", stats.h4_zone_fvg_trades, False),
        ("", "", False),
        ("Tight SL Stats", "", True),
        ("  H4 zone SL used", stats.tight_sl_used, False),
        ("  Baseline SL used", stats.baseline_sl_used, False),
        ("  Avg tight SL dist", f"${stats.avg_sl_distance_tight:.2f}", False),
        ("  Avg baseline SL dist", f"${stats.avg_sl_distance_baseline:.2f}", False),
        ("", "", False), ("Profit - Loss", "", True),
        ("Total Profit", f"${stats.total_profit:,.2f}", False),
        ("Total Loss", f"${stats.total_loss:,.2f}", False),
        ("Net PnL", f"${net_pnl:,.2f}", False),
        ("Profit Factor", f"{stats.profit_factor:.2f}", False),
        ("", "", False), ("Risk Metrics", "", True),
        ("Max Drawdown", f"{stats.max_drawdown:.1f}%", False),
        ("Max DD ($)", f"${stats.max_drawdown_usd:,.2f}", False),
        ("Avg Win", f"${stats.avg_win:,.2f}", False),
        ("Avg Loss", f"${stats.avg_loss:,.2f}", False),
        ("Expectancy", f"${stats.expectancy:,.2f}", False),
        ("Sharpe Ratio", f"{stats.sharpe_ratio:.2f}", False),
    ]
    row = 5
    for lbl, val, hdr in data:
        ws.cell(row=row, column=1, value=lbl)
        ws.cell(row=row, column=2, value=val)
        if hdr:
            ws.cell(row=row, column=1).font = sf
            ws.cell(row=row, column=1).fill = sfill
            ws.cell(row=row, column=2).fill = sfill
        if lbl == "Net PnL":
            ws.cell(row=row, column=2).font = Font(bold=True, color="006100" if net_pnl > 0 else "9C0006")
        row += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    # Trade Log
    ws2 = wb.create_sheet("Trade Log")
    headers = ["Ticket","Entry Time","Exit Time","Dir","Entry","Exit","SL","TP",
               "Lot","Profit ($)","Pips","Result","Exit Reason","Conf","Regime","Session",
               "H4 Zone","SL Type","Orig SL","RR"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill
    for ri, t in enumerate(stats.trades, 2):
        vals = [t.ticket, t.entry_time.strftime("%Y-%m-%d %H:%M"), t.exit_time.strftime("%Y-%m-%d %H:%M"),
                t.direction, t.entry_price, t.exit_price, t.stop_loss, t.take_profit,
                t.lot_size, round(t.profit_usd,2), round(t.profit_pips,1), t.result.value,
                t.exit_reason.value, round(t.smc_confidence,2), t.regime, t.session,
                t.h4_zone_type, t.sl_type, t.original_sl, round(t.rr_ratio, 2)]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.border = border
            if ci == 10 and isinstance(v, (int,float)):
                cell.fill = wfill if v > 0 else (lfill if v < 0 else PatternFill())

    # Equity Curve
    ws3 = wb.create_sheet("Equity Curve")
    for c, h in enumerate(["Trade #","Equity"], 1):
        ws3.cell(row=1, column=c, value=h).font = hf; ws3.cell(row=1, column=c).fill = hfill
    for idx, eq in enumerate(stats.equity_curve):
        ws3.cell(row=idx+2, column=1, value=idx)
        ws3.cell(row=idx+2, column=2, value=round(eq,2))
    if len(stats.equity_curve) > 1:
        chart = LineChart(); chart.title = "Equity Curve"; chart.width = 30; chart.height = 15
        chart.add_data(Reference(ws3, min_col=2, min_row=1, max_row=len(stats.equity_curve)+1), titles_from_data=True)
        ws3.add_chart(chart, "D2")

    # Daily PnL
    ws4 = wb.create_sheet("Daily PnL")
    for c, h in enumerate(["Date","Trades","Wins","WR","Net PnL","Cumulative"], 1):
        ws4.cell(row=1, column=c, value=h).font = hf; ws4.cell(row=1, column=c).fill = hfill
    dpnl = {}
    for t in stats.trades:
        d = t.entry_time.strftime("%Y-%m-%d")
        if d not in dpnl: dpnl[d] = {"t":0,"w":0,"p":0.0}
        dpnl[d]["t"] += 1
        if t.result == TradeResult.WIN: dpnl[d]["w"] += 1
        dpnl[d]["p"] += t.profit_usd
    cum = 0.0
    for ri, (d, v) in enumerate(sorted(dpnl.items()), 2):
        wr = v["w"]/v["t"]*100 if v["t"]>0 else 0
        cum += v["p"]
        for ci, val in enumerate([d, v["t"], v["w"], f"{wr:.0f}%", round(v["p"],2), round(cum,2)], 1):
            ws4.cell(row=ri, column=ci, value=val)

    wb.save(filepath)
    print(f"\n  Report saved: {filepath}")


def generate_log(stats, filepath, start_date, end_date, variant_name):
    net_pnl = stats.total_profit - stats.total_loss
    lines = [
        "=" * 80, f"XAUBOT AI — {variant_name}", "=" * 80,
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", "",
        "--- H4 ZONE FILTER STATS ---",
        f"  Total filtered:     {stats.h4_filtered}",
        f"  BUY filtered:       {stats.h4_filtered_buy}",
        f"  SELL filtered:      {stats.h4_filtered_sell}",
        f"  Trades in OB zone:  {stats.h4_zone_ob_trades}",
        f"  Trades in FVG zone: {stats.h4_zone_fvg_trades}", "",
        "--- TIGHT SL STATS ---",
        f"  H4 zone SL used:    {stats.tight_sl_used}",
        f"  Baseline SL used:   {stats.baseline_sl_used}",
        f"  Avg tight SL dist:  ${stats.avg_sl_distance_tight:.2f}",
        f"  Avg baseline SL dist: ${stats.avg_sl_distance_baseline:.2f}", "",
        "--- PERFORMANCE ---",
        f"  Trades: {stats.total_trades} | Wins: {stats.wins} | Losses: {stats.losses}",
        f"  Win Rate: {stats.win_rate:.1f}% | PF: {stats.profit_factor:.2f}",
        f"  Net PnL: ${net_pnl:,.2f} | Sharpe: {stats.sharpe_ratio:.2f}",
        f"  Max DD: {stats.max_drawdown:.1f}% (${stats.max_drawdown_usd:,.2f})",
        f"  Avg Win: ${stats.avg_win:,.2f} | Avg Loss: ${stats.avg_loss:,.2f}", "",
        "--- EXIT REASONS ---",
    ]
    ec = {}
    for t in stats.trades:
        ec[t.exit_reason.value] = ec.get(t.exit_reason.value, 0) + 1
    for r, c in sorted(ec.items(), key=lambda x: -x[1]):
        lines.append(f"  {r:20s}: {c:4d} ({c/stats.total_trades*100:.1f}%)")
    lines.append("")
    lines.append("--- DIRECTION ---")
    for d in ["BUY", "SELL"]:
        dt = [t for t in stats.trades if t.direction == d]
        dw = sum(1 for t in dt if t.result == TradeResult.WIN)
        dp = sum(t.profit_usd for t in dt)
        lines.append(f"  {d}: {len(dt)} trades, {dw/len(dt)*100:.1f}% WR, ${dp:,.2f}" if dt else f"  {d}: 0 trades")
    lines.append("")
    lines.append("--- H4 ZONE TYPE ---")
    for zt in ["OB", "FVG"]:
        zt_trades = [t for t in stats.trades if t.h4_zone_type == zt]
        zt_w = sum(1 for t in zt_trades if t.result == TradeResult.WIN)
        zt_p = sum(t.profit_usd for t in zt_trades)
        zt_wr = zt_w / len(zt_trades) * 100 if zt_trades else 0
        lines.append(f"  {zt:4s}: {len(zt_trades)} trades, {zt_wr:.1f}% WR, ${zt_p:,.2f}")
    lines.append("")
    lines.append("--- SL TYPE BREAKDOWN ---")
    for slt in ["h4_zone", "baseline"]:
        slt_trades = [t for t in stats.trades if t.sl_type == slt]
        slt_w = sum(1 for t in slt_trades if t.result == TradeResult.WIN)
        slt_p = sum(t.profit_usd for t in slt_trades)
        slt_wr = slt_w / len(slt_trades) * 100 if slt_trades else 0
        lines.append(f"  {slt:12s}: {len(slt_trades)} trades, {slt_wr:.1f}% WR, ${slt_p:,.2f}")
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Log saved: {filepath}")


# ─── Main ──────────────────────────────────────────────────────

def main():
    VARIANT = "SMC + H4 Zone + Tight SL (RR 1:2)"
    print("=" * 70)
    print(f"XAUBOT AI — {VARIANT}")
    print("Base: SMC-Only v4 | Added: H4 zone filter + tighter SL from zone boundary")
    print("=" * 70)

    config = get_config()
    mt5 = MT5Connector(login=config.mt5_login, password=config.mt5_password,
                       server=config.mt5_server, path=config.mt5_path)
    mt5.connect()
    print(f"\nConnected to MT5")

    print("Fetching M15 data...")
    df_m15 = mt5.get_market_data(symbol="XAUUSD", timeframe="M15", count=50000)
    print(f"  M15: {len(df_m15)} bars")

    print("Fetching H4 data...")
    df_h4 = mt5.get_market_data(symbol="XAUUSD", timeframe="H4", count=3000)
    print(f"  H4: {len(df_h4)} bars")

    times = df_m15["time"].to_list()
    print(f"  M15 range: {times[0]} to {times[-1]}")

    end_date = datetime.now()
    start_date = datetime(2025, 8, 1)
    data_start = times[0]
    if hasattr(data_start, 'replace') and data_start.tzinfo:
        start_date = start_date.replace(tzinfo=data_start.tzinfo)
        end_date = end_date.replace(tzinfo=data_start.tzinfo)
    if data_start > start_date:
        start_date = data_start + timedelta(days=5)

    print(f"\n  Backtest period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

    print("\nCalculating M15 indicators...")
    features = FeatureEngineer()
    smc_m15 = SMCAnalyzer(swing_length=config.smc.swing_length, ob_lookback=config.smc.ob_lookback)
    df_m15 = features.calculate_all(df_m15, include_ml_features=True)
    df_m15 = smc_m15.calculate_all(df_m15)

    print("Calculating H4 SMC zones...")
    smc_h4 = SMCAnalyzer(swing_length=5, fvg_min_gap_pips=5.0, ob_lookback=10)
    df_h4 = smc_h4.calculate_all(df_h4)
    h4_bull_obs = (df_h4["ob"] == 1).sum()
    h4_bear_obs = (df_h4["ob"] == -1).sum()
    h4_bull_fvg = df_h4["is_fvg_bull"].sum()
    h4_bear_fvg = df_h4["is_fvg_bear"].sum()
    print(f"  H4 OBs: {h4_bull_obs} bullish, {h4_bear_obs} bearish")
    print(f"  H4 FVGs: {h4_bull_fvg} bullish, {h4_bear_fvg} bearish")

    regime_detector = MarketRegimeDetector(model_path="models/hmm_regime.pkl")
    try:
        regime_detector.load()
        df_m15 = regime_detector.predict(df_m15)
        print("  HMM regime loaded")
    except Exception:
        print("  [WARN] HMM not available")

    backtest = SMCH4ZoneTightSLBacktest(capital=5000.0, max_daily_loss_percent=5.0,
        max_loss_per_trade_percent=1.0, base_lot_size=0.01, max_lot_size=0.02,
        recovery_lot_size=0.01, breakeven_pips=30.0, trail_start_pips=50.0,
        trail_step_pips=30.0, min_profit_to_protect=5.0, max_drawdown_from_peak=50.0,
        trade_cooldown_bars=10, trend_reversal_mult=0.6)

    stats = backtest.run(df_m15=df_m15, df_h4=df_h4, start_date=start_date, end_date=end_date, initial_capital=5000.0)

    net_pnl = stats.total_profit - stats.total_loss
    baseline = 1449.86

    print("\n" + "=" * 70)
    print(f"{VARIANT} — RESULTS")
    print("=" * 70)
    print(f"\n  H4 Zone Filter:")
    print(f"    Filtered:   {stats.h4_filtered} (BUY: {stats.h4_filtered_buy}, SELL: {stats.h4_filtered_sell})")
    print(f"    OB trades:  {stats.h4_zone_ob_trades}")
    print(f"    FVG trades: {stats.h4_zone_fvg_trades}")
    print(f"\n  Tight SL:")
    print(f"    H4 zone SL: {stats.tight_sl_used} trades (avg dist ${stats.avg_sl_distance_tight:.2f})")
    print(f"    Baseline SL: {stats.baseline_sl_used} trades (avg dist ${stats.avg_sl_distance_baseline:.2f})")
    print(f"\n  Performance:")
    print(f"    Trades: {stats.total_trades} | WR: {stats.win_rate:.1f}%")
    print(f"    Net PnL: ${net_pnl:,.2f} | PF: {stats.profit_factor:.2f}")
    print(f"    Max DD: {stats.max_drawdown:.1f}% | Sharpe: {stats.sharpe_ratio:.2f}")
    print(f"    Avg Win: ${stats.avg_win:,.2f} | Avg Loss: ${stats.avg_loss:,.2f}")
    print(f"\n  vs BASELINE: ${net_pnl - baseline:,.2f}")
    print(f"\n  Direction:")
    for d in ["BUY", "SELL"]:
        dt = [t for t in stats.trades if t.direction == d]
        dw = sum(1 for t in dt if t.result == TradeResult.WIN)
        dp = sum(t.profit_usd for t in dt)
        print(f"    {d}: {len(dt)} trades, {dw/len(dt)*100:.1f}% WR, ${dp:,.2f}" if dt else f"    {d}: 0 trades")
    print(f"\n  Exit Reasons:")
    ec = {}
    for t in stats.trades:
        ec[t.exit_reason.value] = ec.get(t.exit_reason.value, 0) + 1
    for r, c in sorted(ec.items(), key=lambda x: -x[1]):
        print(f"    {r:20s}: {c} ({c/stats.total_trades*100:.1f}%)" if stats.total_trades > 0 else "")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "10_h4_zone_tight_sl_results")
    os.makedirs(out_dir, exist_ok=True)
    generate_log(stats, os.path.join(out_dir, f"h4_zone_tight_sl_{ts}.log"), start_date, end_date, VARIANT)
    generate_xlsx_report(stats, os.path.join(out_dir, f"h4_zone_tight_sl_{ts}.xlsx"), start_date, end_date, VARIANT)

    mt5.disconnect()
    print(f"\n{'='*70}\nOutput: {out_dir}\n{'='*70}\nBacktest complete!")


if __name__ == "__main__":
    main()
