"""
Backtest: SMC + Stoch + Sell + Broker SL Only Exit
=====================================================
Base: SMC-Only v4 + Stochastic Filter + Sell Filter Strict
Changed: EXIT SYSTEM stripped down — trust broker SL/TP

Entry filters (from backtest_stoch_sell.py):
  1. Stochastic: BUY blocked if K > 75, SELL blocked if K < 25
  2. Sell Filter: SELL requires ML agree + conf >= 55%

Exit: Same simplified Broker SL Only as backtest_broker_sl.py

KEEP:
  - Broker SL hit (SMC swing low + 1.5x ATR)
  - Broker TP hit (RR 1:1.5)
  - Trailing SL (WIDER: start at $10/100 pips, trail $7/70 pips behind)
  - Weekend close
  - Daily loss limit
  - Hard timeout at 12 hours (48 bars)

REMOVED:
  - Breakeven move
  - Early cut
  - Trend reversal exit
  - Peak protect
  - Stall detection
  - Market signal exit
  - Smart TP
  - Early exit
  - 4h/6h timeout (replaced by 12h hard max)

Usage:
    python backtests/backtest_stoch_sell_broker_sl.py
"""

import polars as pl
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, field
from enum import Enum
import sys
import os
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.mt5_connector import MT5Connector
from src.smc_polars import SMCAnalyzer, SMCSignal
from src.feature_eng import FeatureEngineer
from src.regime_detector import MarketRegimeDetector, MarketRegime
from src.ml_model import TradingModel
from src.config import get_config
from src.dynamic_confidence import DynamicConfidenceManager, create_dynamic_confidence, MarketQuality
from loguru import logger

logger.remove()
logger.add(sys.stderr, level="WARNING")

WIB = ZoneInfo("Asia/Jakarta")

# Stochastic parameters
STOCH_K_PERIOD = 14
STOCH_D_PERIOD = 3
STOCH_OVERBOUGHT = 75
STOCH_OVERSOLD = 25

# Sell filter parameters
SELL_FILTER_MIN_ML_CONF = 0.55


# ─── Enums & Dataclasses ──────────────────────────────────────

class TradeResult(Enum):
    WIN = "WIN"
    LOSS = "LOSS"
    BREAKEVEN = "BREAKEVEN"


class ExitReason(Enum):
    TAKE_PROFIT = "take_profit"
    MAX_LOSS = "max_loss"
    TRAILING_SL = "trailing_sl"
    WEEKEND_CLOSE = "weekend_close"
    DAILY_LIMIT = "daily_limit"
    TIMEOUT = "timeout"


class TradingMode(Enum):
    NORMAL = "normal"
    RECOVERY = "recovery"
    PROTECTED = "protected"
    STOPPED = "stopped"


@dataclass
class SimulatedTrade:
    ticket: int
    entry_time: datetime
    exit_time: datetime
    direction: str
    entry_price: float
    exit_price: float
    stop_loss: float
    take_profit: float
    lot_size: float
    profit_usd: float
    profit_pips: float
    result: TradeResult
    exit_reason: ExitReason
    smc_confidence: float
    regime: str
    session: str
    signal_reason: str
    has_bos: bool = False
    has_choch: bool = False
    has_fvg: bool = False
    has_ob: bool = False
    atr_at_entry: float = 0.0
    rr_ratio: float = 0.0
    trading_mode: str = "normal"
    stoch_k: float = 0.0
    stoch_d: float = 0.0


@dataclass
class BacktestStats:
    total_trades: int = 0
    wins: int = 0
    losses: int = 0
    total_profit: float = 0.0
    total_loss: float = 0.0
    max_drawdown: float = 0.0
    max_drawdown_usd: float = 0.0
    win_rate: float = 0.0
    profit_factor: float = 0.0
    avg_win: float = 0.0
    avg_loss: float = 0.0
    avg_trade: float = 0.0
    expectancy: float = 0.0
    sharpe_ratio: float = 0.0
    trades: List[SimulatedTrade] = field(default_factory=list)
    equity_curve: List[float] = field(default_factory=list)
    avoided_signals: int = 0
    daily_limit_stops: int = 0
    recovery_mode_trades: int = 0
    # Stochastic filter stats
    stoch_filtered: int = 0
    stoch_filtered_buy_overbought: int = 0
    stoch_filtered_sell_oversold: int = 0
    # Sell filter stats
    sell_filtered: int = 0
    sell_filtered_no_ml_agree: int = 0
    sell_filtered_low_conf: int = 0


# ─── Stochastic Calculation ──────────────────────────────────

def calculate_stochastic(df: pl.DataFrame, k_period: int = 14, d_period: int = 3) -> pl.DataFrame:
    """Calculate Stochastic Oscillator %K and %D."""
    highs = df["high"].to_list()
    lows = df["low"].to_list()
    closes = df["close"].to_list()
    n = len(closes)

    stoch_k = [50.0] * n
    stoch_d = [50.0] * n

    for i in range(k_period - 1, n):
        high_max = max(highs[i - k_period + 1 : i + 1])
        low_min = min(lows[i - k_period + 1 : i + 1])
        if high_max - low_min > 0:
            stoch_k[i] = ((closes[i] - low_min) / (high_max - low_min)) * 100
        else:
            stoch_k[i] = 50.0

    for i in range(k_period - 1 + d_period - 1, n):
        stoch_d[i] = np.mean(stoch_k[i - d_period + 1 : i + 1])

    df = df.with_columns([
        pl.Series("stoch_k", stoch_k),
        pl.Series("stoch_d", stoch_d),
    ])
    return df


# ─── SMC + Stoch + Sell + Broker SL Only Backtest ────────────

class StochSellBrokerSLBacktest:
    """SMC-Only v4 + Stochastic Filter + Sell Filter Strict + Broker SL Only exit."""

    def __init__(
        self,
        capital: float = 5000.0,
        # SmartRiskManager params (synced)
        max_daily_loss_percent: float = 5.0,
        max_loss_per_trade_percent: float = 1.0,
        base_lot_size: float = 0.01,
        max_lot_size: float = 0.02,
        recovery_lot_size: float = 0.01,
        trend_reversal_threshold: float = 0.75,
        max_concurrent_positions: int = 2,
        # Other
        trade_cooldown_bars: int = 10,
    ):
        self.capital = capital
        self.max_daily_loss_usd = capital * (max_daily_loss_percent / 100)
        self.max_loss_per_trade = capital * (max_loss_per_trade_percent / 100)
        self.base_lot_size = base_lot_size
        self.max_lot_size = max_lot_size
        self.recovery_lot_size = recovery_lot_size
        self.trend_reversal_threshold = trend_reversal_threshold
        self.max_concurrent_positions = max_concurrent_positions
        self.trade_cooldown_bars = trade_cooldown_bars

        config = get_config()
        self.smc = SMCAnalyzer(
            swing_length=config.smc.swing_length,
            ob_lookback=config.smc.ob_lookback,
        )
        self.features = FeatureEngineer()
        self.dynamic_confidence = create_dynamic_confidence()

        # ML model for entry evaluation + stoch/sell filter
        self.ml_model = TradingModel(model_path="models/xgboost_model.pkl")
        try:
            self.ml_model.load()
            print("  ML model loaded (for entry + stoch filter + sell filter)")
        except Exception:
            print("  [WARN] ML model not loaded — ML checks disabled")

        self.regime_detector = MarketRegimeDetector(model_path="models/hmm_regime.pkl")
        try:
            self.regime_detector.load()
        except Exception:
            print("  [WARN] HMM model not loaded")

        self._ticket_counter = 2000000

    # ── Session filter (synced) ──

    def _get_session_from_time(self, dt: datetime) -> Tuple[str, bool, float]:
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib_time = dt.astimezone(WIB)
        hour = wib_time.hour
        if 6 <= hour < 15:
            return "Sydney-Tokyo", True, 0.5
        elif 15 <= hour < 16:
            return "Tokyo-London Overlap", True, 0.75
        elif 16 <= hour < 19:
            return "London Early", True, 0.8
        elif 19 <= hour < 24:
            return "London-NY Overlap (Golden)", True, 1.0
        elif 0 <= hour < 4:
            return "NY Session", True, 0.9
        else:
            return "Off Hours", False, 0.0

    def _hours_to_golden(self, dt: datetime) -> float:
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib = dt.astimezone(WIB)
        if 19 <= wib.hour < 24:
            return 0
        target = wib.replace(hour=19, minute=0, second=0, microsecond=0)
        if wib.hour >= 19:
            target += timedelta(days=1)
        return max(0, (target - wib).total_seconds() / 3600)

    def _is_near_weekend_close(self, dt: datetime) -> bool:
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib = dt.astimezone(WIB)
        if wib.weekday() == 5 and wib.hour >= 4 and wib.minute >= 30:
            return True
        return False

    # ── Lot sizing (synced) ──

    def _calculate_lot_size(self, confidence, regime, trading_mode, session_mult):
        if trading_mode == TradingMode.STOPPED:
            return 0
        lot = self.base_lot_size
        if trading_mode in (TradingMode.RECOVERY, TradingMode.PROTECTED):
            lot = self.recovery_lot_size
        else:
            if confidence >= 0.65:
                lot = self.max_lot_size
            elif confidence >= 0.55:
                lot = self.base_lot_size
            else:
                lot = self.recovery_lot_size
        if regime.lower() in ["high_volatility", "crisis"]:
            lot = self.recovery_lot_size
        lot = max(0.01, lot * session_mult)
        return round(lot, 2)

    # ── Simplified exit simulation (Broker SL Only) ──

    def _simulate_trade_exit(
        self, df, entry_idx, direction, entry_price, take_profit, stop_loss,
        lot_size, daily_loss_so_far, feature_cols, max_bars=100,
    ):
        pip_value = 10
        highs = df["high"].to_list()
        lows = df["low"].to_list()
        closes = df["close"].to_list()
        times = df["time"].to_list()

        # Wide trailing params (NO breakeven, just trailing)
        trail_start_pips = 100.0  # Start trail after $10 profit
        trail_step_pips = 70.0    # Trail $7 behind price
        current_sl = stop_loss
        trailing_active = False

        for i in range(entry_idx + 1, min(entry_idx + max_bars, len(df))):
            high, low, close, current_time = highs[i], lows[i], closes[i], times[i]

            if direction == "BUY":
                current_pips = (close - entry_price) / 0.1
                pip_profit_from_entry = current_pips
            else:
                current_pips = (entry_price - close) / 0.1
                pip_profit_from_entry = current_pips
            current_profit = current_pips * pip_value * lot_size
            bars_since_entry = i - entry_idx

            # 1. BROKER TP HIT
            if direction == "BUY" and high >= take_profit:
                pips = (take_profit - entry_price) / 0.1
                return pips * pip_value * lot_size, pips, ExitReason.TAKE_PROFIT, i, take_profit
            elif direction == "SELL" and low <= take_profit:
                pips = (entry_price - take_profit) / 0.1
                return pips * pip_value * lot_size, pips, ExitReason.TAKE_PROFIT, i, take_profit

            # 2. BROKER SL HIT (original SMC SL or trailing SL)
            if direction == "BUY" and low <= current_sl:
                pips = (current_sl - entry_price) / 0.1
                reason = ExitReason.TRAILING_SL if trailing_active else ExitReason.MAX_LOSS
                return pips * pip_value * lot_size, pips, reason, i, current_sl
            elif direction == "SELL" and high >= current_sl:
                pips = (entry_price - current_sl) / 0.1
                reason = ExitReason.TRAILING_SL if trailing_active else ExitReason.MAX_LOSS
                return pips * pip_value * lot_size, pips, reason, i, current_sl

            # 3. WIDE TRAILING SL (no breakeven, start at $10 profit)
            if pip_profit_from_entry >= trail_start_pips:
                trail_distance = trail_step_pips * 0.1
                if direction == "BUY":
                    new_trail_sl = close - trail_distance
                    if new_trail_sl > current_sl:
                        current_sl = new_trail_sl
                        trailing_active = True
                else:
                    new_trail_sl = close + trail_distance
                    if current_sl == 0 or new_trail_sl < current_sl:
                        current_sl = new_trail_sl
                        trailing_active = True

            # 4. WEEKEND CLOSE
            if self._is_near_weekend_close(current_time):
                if current_profit > -10:
                    return current_profit, current_pips, ExitReason.WEEKEND_CLOSE, i, close

            # 5. DAILY LOSS LIMIT
            if daily_loss_so_far + abs(min(0, current_profit)) >= self.max_daily_loss_usd:
                return current_profit, current_pips, ExitReason.DAILY_LIMIT, i, close

            # 6. HARD TIMEOUT (12 hours = 48 bars on M15)
            if bars_since_entry >= 48:
                return current_profit, current_pips, ExitReason.TIMEOUT, i, close

        # End of data
        final_idx = min(entry_idx + max_bars - 1, len(df) - 1)
        fp = closes[final_idx]
        pips = (fp - entry_price) / 0.1 if direction == "BUY" else (entry_price - fp) / 0.1
        return pips * pip_value * lot_size, pips, ExitReason.TIMEOUT, final_idx, fp

    # ── Main backtest run ──

    def run(self, df, start_date=None, end_date=None, initial_capital=5000.0):
        stats = BacktestStats()
        capital = initial_capital
        peak_capital = initial_capital
        stats.equity_curve.append(capital)

        daily_loss = 0.0
        daily_profit = 0.0
        daily_trades = 0
        consecutive_losses = 0
        trading_mode = TradingMode.NORMAL
        current_date = None

        feature_cols = []
        if self.ml_model.fitted and self.ml_model.feature_names:
            feature_cols = [f for f in self.ml_model.feature_names if f in df.columns]

        stoch_k_list = df["stoch_k"].to_list()
        stoch_d_list = df["stoch_d"].to_list()

        times = df["time"].to_list()
        start_idx = next((i for i, t in enumerate(times) if t >= start_date), 100) if start_date else 100
        end_idx = next((i for i, t in enumerate(times) if t > end_date), len(df) - 100) if end_date else len(df) - 100

        last_trade_idx = -self.trade_cooldown_bars * 2

        print(f"\n  Running SMC + Stoch + Sell + Broker SL Only backtest...")
        print(f"  Stochastic: BUY blocked if K > {STOCH_OVERBOUGHT}, SELL blocked if K < {STOCH_OVERSOLD}")
        print(f"  Sell Filter: SELL requires ML agree + conf >= {SELL_FILTER_MIN_ML_CONF:.0%}")
        print(f"  Exit: Broker SL Only (simplified)")
        print(f"  Date range: {times[start_idx]} to {times[end_idx - 1]}")
        print(f"  Total bars: {end_idx - start_idx}")

        for i in range(start_idx, end_idx):
            if i - last_trade_idx < self.trade_cooldown_bars:
                continue

            current_time = times[i]

            # Daily reset
            trade_date = current_time.date() if hasattr(current_time, 'date') else current_time
            if current_date is None or trade_date != current_date:
                daily_loss = 0.0
                daily_profit = 0.0
                daily_trades = 0
                current_date = trade_date
                if consecutive_losses < 2:
                    trading_mode = TradingMode.NORMAL

            if trading_mode == TradingMode.STOPPED:
                continue

            session_name, can_trade, lot_mult = self._get_session_from_time(current_time)
            if not can_trade:
                continue

            if hasattr(current_time, 'weekday') and current_time.weekday() >= 5:
                continue

            df_slice = df.head(i + 1)

            # Regime check
            regime = "normal"
            regime_state = None
            try:
                if self.regime_detector.fitted:
                    regime_state = self.regime_detector.get_current_state(df_slice)
                    if regime_state:
                        regime = regime_state.regime.value
                        if regime_state.regime == MarketRegime.CRISIS:
                            continue
                        if regime_state.recommendation == "SLEEP":
                            continue
            except Exception:
                pass

            # DynamicConfidence AVOID filter
            try:
                ml_signal = ""
                ml_confidence = 0.5
                if self.ml_model.fitted and feature_cols:
                    ml_pred = self.ml_model.predict(df_slice, feature_cols)
                    ml_signal = ml_pred.signal
                    ml_confidence = ml_pred.confidence

                market_analysis = self.dynamic_confidence.analyze_market(
                    session=session_name, regime=regime, volatility="medium",
                    trend_direction=regime, has_smc_signal=True,
                    ml_signal=ml_signal, ml_confidence=ml_confidence,
                )
                if market_analysis.quality == MarketQuality.AVOID:
                    stats.avoided_signals += 1
                    continue
            except Exception:
                pass

            # SMC Signal
            try:
                smc_signal = self.smc.generate_signal(df_slice)
            except Exception:
                continue

            if smc_signal is None:
                continue

            # ═══════════════════════════════════════════════════════
            # FILTER 1: STOCHASTIC
            # ═══════════════════════════════════════════════════════
            current_stoch_k = stoch_k_list[i] if i < len(stoch_k_list) else 50.0
            current_stoch_d = stoch_d_list[i] if i < len(stoch_d_list) else 50.0

            if smc_signal.signal_type == "BUY":
                if current_stoch_k > STOCH_OVERBOUGHT:
                    stats.stoch_filtered += 1
                    stats.stoch_filtered_buy_overbought += 1
                    continue

            if smc_signal.signal_type == "SELL":
                if current_stoch_k < STOCH_OVERSOLD:
                    stats.stoch_filtered += 1
                    stats.stoch_filtered_sell_oversold += 1
                    continue

            # ═══════════════════════════════════════════════════════
            # FILTER 2: SELL FILTER STRICT (ML agree + conf >= 55%)
            # ═══════════════════════════════════════════════════════
            if smc_signal.signal_type == "SELL":
                if ml_signal != "SELL":
                    stats.sell_filtered += 1
                    stats.sell_filtered_no_ml_agree += 1
                    continue
                if ml_confidence < SELL_FILTER_MIN_ML_CONF:
                    stats.sell_filtered += 1
                    stats.sell_filtered_low_conf += 1
                    continue

            # ═══════════════════════════════════════════════════════

            # SMC details
            recent_df = df_slice.tail(10)
            recent_bos = recent_df["bos"].to_list() if "bos" in df_slice.columns else []
            recent_choch = recent_df["choch"].to_list() if "choch" in df_slice.columns else []
            recent_fvg_bull = recent_df["is_fvg_bull"].to_list() if "is_fvg_bull" in df_slice.columns else []
            recent_fvg_bear = recent_df["is_fvg_bear"].to_list() if "is_fvg_bear" in df_slice.columns else []
            recent_obs = recent_df["ob"].to_list() if "ob" in df_slice.columns else []

            has_bos = 1 in recent_bos or -1 in recent_bos
            has_choch = 1 in recent_choch or -1 in recent_choch
            has_fvg = any(recent_fvg_bull) or any(recent_fvg_bear)
            has_ob = 1 in recent_obs or -1 in recent_obs

            atr_at_entry = 12.0
            if "atr" in df_slice.columns:
                atr_val = df_slice.tail(1)["atr"].item()
                if atr_val is not None and atr_val > 0:
                    atr_at_entry = atr_val

            # Confidence (synced)
            confidence = smc_signal.confidence
            ml_agrees = (
                (smc_signal.signal_type == "BUY" and ml_signal == "BUY") or
                (smc_signal.signal_type == "SELL" and ml_signal == "SELL")
            )
            if ml_agrees:
                confidence = (smc_signal.confidence + ml_confidence) / 2
            if regime == "high_volatility":
                confidence *= 0.9

            # Lot size
            lot_size = self._calculate_lot_size(confidence, regime, trading_mode, lot_mult)
            if lot_size <= 0:
                continue

            if trading_mode == TradingMode.RECOVERY:
                stats.recovery_mode_trades += 1

            # Execute trade
            entry_price = smc_signal.entry_price
            take_profit_price = smc_signal.take_profit
            stop_loss_price = smc_signal.stop_loss
            risk = abs(entry_price - stop_loss_price)
            rr = abs(take_profit_price - entry_price) / risk if risk > 0 else 0

            profit, pips, exit_reason, exit_idx, exit_price = self._simulate_trade_exit(
                df=df, entry_idx=i, direction=smc_signal.signal_type,
                entry_price=entry_price, take_profit=take_profit_price,
                stop_loss=stop_loss_price, lot_size=lot_size,
                daily_loss_so_far=daily_loss, feature_cols=feature_cols,
            )

            self._ticket_counter += 1
            result = TradeResult.WIN if profit > 0 else (TradeResult.LOSS if profit < 0 else TradeResult.BREAKEVEN)

            trade = SimulatedTrade(
                ticket=self._ticket_counter,
                entry_time=current_time,
                exit_time=times[exit_idx] if exit_idx < len(times) else times[-1],
                direction=smc_signal.signal_type,
                entry_price=entry_price, exit_price=exit_price,
                stop_loss=stop_loss_price, take_profit=take_profit_price,
                lot_size=lot_size, profit_usd=profit, profit_pips=pips,
                result=result, exit_reason=exit_reason,
                smc_confidence=confidence, regime=regime,
                session=session_name, signal_reason=smc_signal.reason,
                has_bos=has_bos, has_choch=has_choch, has_fvg=has_fvg, has_ob=has_ob,
                atr_at_entry=atr_at_entry, rr_ratio=rr,
                trading_mode=trading_mode.value,
                stoch_k=current_stoch_k, stoch_d=current_stoch_d,
            )
            stats.trades.append(trade)

            # Update state
            stats.total_trades += 1
            daily_trades += 1
            capital += profit

            if profit > 0:
                stats.wins += 1
                stats.total_profit += profit
                daily_profit += profit
                consecutive_losses = 0
                if trading_mode == TradingMode.RECOVERY:
                    trading_mode = TradingMode.NORMAL
            else:
                stats.losses += 1
                stats.total_loss += abs(profit)
                daily_loss += abs(profit)
                consecutive_losses += 1

            if daily_loss >= self.max_daily_loss_usd:
                trading_mode = TradingMode.STOPPED
                stats.daily_limit_stops += 1
            elif consecutive_losses >= 3 or daily_loss >= self.max_daily_loss_usd * 0.6:
                trading_mode = TradingMode.PROTECTED
            elif consecutive_losses >= 2:
                trading_mode = TradingMode.RECOVERY

            if capital > peak_capital:
                peak_capital = capital
            drawdown_pct = (peak_capital - capital) / peak_capital * 100
            drawdown_usd = peak_capital - capital
            if drawdown_pct > stats.max_drawdown:
                stats.max_drawdown = drawdown_pct
                stats.max_drawdown_usd = drawdown_usd

            stats.equity_curve.append(capital)
            last_trade_idx = exit_idx

            if stats.total_trades % 100 == 0:
                print(f"  {stats.total_trades} trades processed...")

        # Final statistics
        if stats.total_trades > 0:
            stats.win_rate = stats.wins / stats.total_trades * 100
            stats.avg_win = stats.total_profit / stats.wins if stats.wins > 0 else 0
            stats.avg_loss = stats.total_loss / stats.losses if stats.losses > 0 else 0
            stats.avg_trade = (stats.total_profit - stats.total_loss) / stats.total_trades
            stats.profit_factor = stats.total_profit / stats.total_loss if stats.total_loss > 0 else float("inf")
            win_prob = stats.wins / stats.total_trades
            loss_prob = stats.losses / stats.total_trades
            stats.expectancy = (win_prob * stats.avg_win) - (loss_prob * stats.avg_loss)
            returns = [t.profit_usd for t in stats.trades]
            if len(returns) > 1:
                avg_return = np.mean(returns)
                std_return = np.std(returns)
                stats.sharpe_ratio = (avg_return / std_return) * np.sqrt(252) if std_return > 0 else 0

        return stats


# ─── XLSX Report ───────────────────────────────────────────────

def generate_xlsx_report(stats: BacktestStats, filepath: str, start_date, end_date):
    wb = Workbook()
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_font = Font(name="Calibri", bold=True, size=10)
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    net_pnl = stats.total_profit - stats.total_loss

    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1F4E79"
    ws.merge_cells("A1:F1")
    ws["A1"] = "XAUBot AI — SMC + Stoch + Sell + Broker SL Only Backtest"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws["A2"] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"].font = Font(name="Calibri", size=10, italic=True)

    summary_data = [
        ("Performance Metrics", "", True),
        ("Total Trades", stats.total_trades, False),
        ("Wins", stats.wins, False),
        ("Losses", stats.losses, False),
        ("Win Rate", f"{stats.win_rate:.1f}%", False),
        ("", "", False),
        ("Stochastic Filter", "", True),
        ("  Total blocked", stats.stoch_filtered, False),
        ("  BUY blocked (overbought)", stats.stoch_filtered_buy_overbought, False),
        ("  SELL blocked (oversold)", stats.stoch_filtered_sell_oversold, False),
        ("Sell Filter", "", True),
        ("  Total blocked", stats.sell_filtered, False),
        ("  ML disagree", stats.sell_filtered_no_ml_agree, False),
        ("  Low ML conf", stats.sell_filtered_low_conf, False),
        ("", "", False),
        ("Other Filters", "", True),
        ("Avoided (AVOID filter)", stats.avoided_signals, False),
        ("Recovery Mode Trades", stats.recovery_mode_trades, False),
        ("Daily Limit Stops", stats.daily_limit_stops, False),
        ("", "", False),
        ("Profit - Loss", "", True),
        ("Total Profit", f"${stats.total_profit:,.2f}", False),
        ("Total Loss", f"${stats.total_loss:,.2f}", False),
        ("Net PnL", f"${net_pnl:,.2f}", False),
        ("Profit Factor", f"{stats.profit_factor:.2f}", False),
        ("", "", False),
        ("Risk Metrics", "", True),
        ("Max Drawdown", f"{stats.max_drawdown:.1f}%", False),
        ("Max Drawdown ($)", f"${stats.max_drawdown_usd:,.2f}", False),
        ("Avg Win", f"${stats.avg_win:,.2f}", False),
        ("Avg Loss", f"${stats.avg_loss:,.2f}", False),
        ("Avg Trade", f"${stats.avg_trade:,.2f}", False),
        ("Expectancy", f"${stats.expectancy:,.2f}", False),
        ("Sharpe Ratio", f"{stats.sharpe_ratio:.2f}", False),
    ]

    row = 5
    for label, value, is_header in summary_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if is_header:
            ws.cell(row=row, column=1).font = subheader_font
            ws.cell(row=row, column=1).fill = subheader_fill
            ws.cell(row=row, column=2).fill = subheader_fill
        if label == "Net PnL":
            ws.cell(row=row, column=2).font = Font(bold=True, color="006100" if net_pnl > 0 else "9C0006")
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    # Exit reasons
    exit_counts = {}
    for t in stats.trades:
        reason = t.exit_reason.value
        exit_counts[reason] = exit_counts.get(reason, 0) + 1
    ws.cell(row=5, column=4, value="Exit Reasons")
    ws.cell(row=5, column=4).font = subheader_font
    ws.cell(row=5, column=4).fill = subheader_fill
    ws.cell(row=5, column=5).fill = subheader_fill
    ws.cell(row=5, column=6).fill = subheader_fill
    row = 6
    for reason, count in sorted(exit_counts.items(), key=lambda x: -x[1]):
        pct = count / stats.total_trades * 100 if stats.total_trades > 0 else 0
        ws.cell(row=row, column=4, value=reason)
        ws.cell(row=row, column=5, value=count)
        ws.cell(row=row, column=6, value=f"{pct:.1f}%")
        row += 1

    # Session breakdown
    row += 1
    ws.cell(row=row, column=4, value="Session Performance")
    ws.cell(row=row, column=4).font = subheader_font
    ws.cell(row=row, column=4).fill = subheader_fill
    for c in range(5, 8):
        ws.cell(row=row, column=c).fill = subheader_fill
    row += 1
    for lbl, col in [("Session", 4), ("Trades", 5), ("WR", 6), ("Net PnL", 7)]:
        ws.cell(row=row, column=col, value=lbl).font = Font(bold=True)
    row += 1
    session_stats = {}
    for t in stats.trades:
        s = t.session
        if s not in session_stats:
            session_stats[s] = {"w": 0, "l": 0, "p": 0.0}
        if t.result == TradeResult.WIN:
            session_stats[s]["w"] += 1
        else:
            session_stats[s]["l"] += 1
        session_stats[s]["p"] += t.profit_usd
    for sess, d in sorted(session_stats.items(), key=lambda x: -x[1]["p"]):
        total = d["w"] + d["l"]
        wr = d["w"] / total * 100 if total > 0 else 0
        ws.cell(row=row, column=4, value=sess)
        ws.cell(row=row, column=5, value=total)
        ws.cell(row=row, column=6, value=f"{wr:.1f}%")
        ws.cell(row=row, column=7, value=f"${d['p']:,.2f}")
        ws.cell(row=row, column=7).font = Font(color="006100" if d["p"] >= 0 else "9C0006")
        row += 1

    # SMC Component Analysis
    row += 1
    ws.cell(row=row, column=4, value="SMC Component Analysis")
    ws.cell(row=row, column=4).font = subheader_font
    ws.cell(row=row, column=4).fill = subheader_fill
    for c in range(5, 8):
        ws.cell(row=row, column=c).fill = subheader_fill
    row += 1
    for lbl, col in [("Component", 4), ("Trades", 5), ("WR", 6), ("Net PnL", 7)]:
        ws.cell(row=row, column=col, value=lbl).font = Font(bold=True)
    row += 1
    for comp_name, attr in [("BOS", "has_bos"), ("CHoCH", "has_choch"), ("FVG", "has_fvg"), ("OB", "has_ob")]:
        ct = [t for t in stats.trades if getattr(t, attr)]
        cw = sum(1 for t in ct if t.result == TradeResult.WIN)
        cp = sum(t.profit_usd for t in ct)
        cwr = cw / len(ct) * 100 if ct else 0
        ws.cell(row=row, column=4, value=comp_name)
        ws.cell(row=row, column=5, value=len(ct))
        ws.cell(row=row, column=6, value=f"{cwr:.1f}%")
        ws.cell(row=row, column=7, value=f"${cp:,.2f}")
        row += 1

    col_widths = {4: 28, 5: 10, 6: 12, 7: 14}
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # Trade Log
    ws2 = wb.create_sheet("Trade Log")
    ws2.sheet_properties.tabColor = "2E75B6"
    headers = [
        "Ticket", "Entry Time", "Exit Time", "Dir", "Entry", "Exit", "SL", "TP",
        "Lot", "Profit ($)", "Pips", "Result", "Exit Reason", "SMC Conf",
        "Regime", "Session", "Signal", "BOS", "CHoCH", "FVG", "OB", "ATR", "RR",
        "Mode", "Stoch K", "Stoch D",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for ri, t in enumerate(stats.trades, 2):
        vals = [
            t.ticket, t.entry_time.strftime("%Y-%m-%d %H:%M"), t.exit_time.strftime("%Y-%m-%d %H:%M"),
            t.direction, t.entry_price, t.exit_price, t.stop_loss, t.take_profit,
            t.lot_size, round(t.profit_usd, 2), round(t.profit_pips, 1), t.result.value,
            t.exit_reason.value, round(t.smc_confidence, 2), t.regime, t.session, t.signal_reason,
            "Y" if t.has_bos else "", "Y" if t.has_choch else "", "Y" if t.has_fvg else "",
            "Y" if t.has_ob else "", round(t.atr_at_entry, 2), round(t.rr_ratio, 2), t.trading_mode,
            round(t.stoch_k, 1), round(t.stoch_d, 1),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.border = border
            if ci == 10 and isinstance(v, (int, float)):
                cell.fill = win_fill if v > 0 else (loss_fill if v < 0 else PatternFill())
            if ci == 12:
                cell.fill = win_fill if v == "WIN" else (loss_fill if v == "LOSS" else PatternFill())
    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = max(11, len(headers[col - 1]) + 3)

    # Equity Curve
    ws3 = wb.create_sheet("Equity Curve")
    ws3.sheet_properties.tabColor = "548235"
    for c, h in enumerate(["Trade #", "Equity", "Drawdown ($)"], 1):
        ws3.cell(row=1, column=c, value=h).font = header_font
        ws3.cell(row=1, column=c).fill = header_fill
    peak = stats.equity_curve[0] if stats.equity_curve else 5000
    for idx, eq in enumerate(stats.equity_curve):
        if eq > peak:
            peak = eq
        ws3.cell(row=idx + 2, column=1, value=idx)
        ws3.cell(row=idx + 2, column=2, value=round(eq, 2))
        ws3.cell(row=idx + 2, column=3, value=round(peak - eq, 2))
    if len(stats.equity_curve) > 1:
        chart = LineChart()
        chart.title = "Equity Curve"
        chart.style = 10
        chart.y_axis.title = "Equity ($)"
        chart.x_axis.title = "Trade #"
        chart.width = 30
        chart.height = 15
        data = Reference(ws3, min_col=2, min_row=1, max_row=len(stats.equity_curve) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.series[0].graphicalProperties.line.width = 20000
        ws3.add_chart(chart, "E2")

    # Daily PnL
    ws4 = wb.create_sheet("Daily PnL")
    ws4.sheet_properties.tabColor = "BF8F00"
    daily_pnl = {}
    for t in stats.trades:
        day = t.entry_time.strftime("%Y-%m-%d")
        if day not in daily_pnl:
            daily_pnl[day] = {"trades": 0, "wins": 0, "profit": 0.0}
        daily_pnl[day]["trades"] += 1
        if t.result == TradeResult.WIN:
            daily_pnl[day]["wins"] += 1
        daily_pnl[day]["profit"] += t.profit_usd
    for c, h in enumerate(["Date", "Trades", "Wins", "WR", "Net PnL", "Cumulative"], 1):
        ws4.cell(row=1, column=c, value=h).font = header_font
        ws4.cell(row=1, column=c).fill = header_fill
    cum = 0.0
    for ri, (day, d) in enumerate(sorted(daily_pnl.items()), 2):
        wr = d["wins"] / d["trades"] * 100 if d["trades"] > 0 else 0
        cum += d["profit"]
        ws4.cell(row=ri, column=1, value=day)
        ws4.cell(row=ri, column=2, value=d["trades"])
        ws4.cell(row=ri, column=3, value=d["wins"])
        ws4.cell(row=ri, column=4, value=f"{wr:.0f}%")
        ws4.cell(row=ri, column=5, value=round(d["profit"], 2))
        ws4.cell(row=ri, column=6, value=round(cum, 2))
        ws4.cell(row=ri, column=5).fill = win_fill if d["profit"] >= 0 else loss_fill
    for c in range(1, 7):
        ws4.column_dimensions[get_column_letter(c)].width = 16

    wb.save(filepath)
    print(f"\n  Report saved: {filepath}")


# ─── Log Generator ─────────────────────────────────────────────

def generate_log(stats: BacktestStats, filepath: str, start_date, end_date):
    net_pnl = stats.total_profit - stats.total_loss
    lines = []
    lines.append("=" * 80)
    lines.append("XAUBOT AI — SMC + Stoch + Sell + Broker SL Only Exit Backtest Log")
    lines.append("=" * 80)
    lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
    lines.append(f"Strategy: SMC-Only v4 + Stochastic (K={STOCH_K_PERIOD}) + Sell Filter (ML >= {SELL_FILTER_MIN_ML_CONF:.0%}) + Broker SL Only Exit")
    lines.append("")
    lines.append("--- FILTER STATS ---")
    lines.append(f"  Stochastic Blocked:       {stats.stoch_filtered}")
    lines.append(f"    BUY (K>{STOCH_OVERBOUGHT}):           {stats.stoch_filtered_buy_overbought}")
    lines.append(f"    SELL (K<{STOCH_OVERSOLD}):           {stats.stoch_filtered_sell_oversold}")
    lines.append(f"  Sell Filter Blocked:      {stats.sell_filtered}")
    lines.append(f"    ML disagree:            {stats.sell_filtered_no_ml_agree}")
    lines.append(f"    Low ML conf:            {stats.sell_filtered_low_conf}")
    lines.append(f"  Combined blocked:         {stats.stoch_filtered + stats.sell_filtered}")
    lines.append("")
    lines.append("--- PERFORMANCE SUMMARY ---")
    lines.append(f"  Total Trades:    {stats.total_trades}")
    lines.append(f"  Wins:            {stats.wins}")
    lines.append(f"  Losses:          {stats.losses}")
    lines.append(f"  Win Rate:        {stats.win_rate:.1f}%")
    lines.append(f"  Total Profit:    ${stats.total_profit:,.2f}")
    lines.append(f"  Total Loss:      ${stats.total_loss:,.2f}")
    lines.append(f"  Net PnL:         ${net_pnl:,.2f}")
    lines.append(f"  Profit Factor:   {stats.profit_factor:.2f}")
    lines.append(f"  Max Drawdown:    {stats.max_drawdown:.1f}% (${stats.max_drawdown_usd:,.2f})")
    lines.append(f"  Avg Win:         ${stats.avg_win:,.2f}")
    lines.append(f"  Avg Loss:        ${stats.avg_loss:,.2f}")
    lines.append(f"  Expectancy:      ${stats.expectancy:,.2f}")
    lines.append(f"  Sharpe Ratio:    {stats.sharpe_ratio:.2f}")
    lines.append(f"  Avoided (AVOID): {stats.avoided_signals}")
    lines.append(f"  Recovery Trades: {stats.recovery_mode_trades}")
    lines.append(f"  Daily Stops:     {stats.daily_limit_stops}")
    lines.append("")

    lines.append("--- EXIT REASON BREAKDOWN ---")
    exit_counts = {}
    for t in stats.trades:
        r = t.exit_reason.value
        exit_counts[r] = exit_counts.get(r, 0) + 1
    for reason, count in sorted(exit_counts.items(), key=lambda x: -x[1]):
        pct = count / stats.total_trades * 100 if stats.total_trades > 0 else 0
        lines.append(f"  {reason:20s}: {count:4d} ({pct:5.1f}%)")
    lines.append("")

    lines.append("--- DIRECTION BREAKDOWN ---")
    for d in ["BUY", "SELL"]:
        dt = [t for t in stats.trades if t.direction == d]
        dw = sum(1 for t in dt if t.result == TradeResult.WIN)
        dp = sum(t.profit_usd for t in dt)
        dwr = dw / len(dt) * 100 if dt else 0
        lines.append(f"  {d}: {len(dt)} trades, {dwr:.1f}% WR, ${dp:,.2f}")
    lines.append("")

    lines.append("--- SESSION BREAKDOWN ---")
    ss = {}
    for t in stats.trades:
        if t.session not in ss:
            ss[t.session] = {"w": 0, "l": 0, "p": 0.0}
        if t.result == TradeResult.WIN:
            ss[t.session]["w"] += 1
        else:
            ss[t.session]["l"] += 1
        ss[t.session]["p"] += t.profit_usd
    for s, d in sorted(ss.items(), key=lambda x: -x[1]["p"]):
        total = d["w"] + d["l"]
        wr = d["w"] / total * 100 if total > 0 else 0
        lines.append(f"  {s:30s}: {total:3d} trades, {wr:5.1f}% WR, ${d['p']:>8,.2f}")
    lines.append("")

    lines.append("--- SMC COMPONENT ANALYSIS ---")
    for cn, attr in [("BOS", "has_bos"), ("CHoCH", "has_choch"), ("FVG", "has_fvg"), ("OB", "has_ob")]:
        ct = [t for t in stats.trades if getattr(t, attr)]
        cw = sum(1 for t in ct if t.result == TradeResult.WIN)
        cp = sum(t.profit_usd for t in ct)
        cwr = cw / len(ct) * 100 if ct else 0
        lines.append(f"  {cn:6s}: {len(ct):3d} trades, {cwr:5.1f}% WR, ${cp:>8,.2f}")
    lines.append("")

    lines.append("--- TRADE LOG ---")
    lines.append(f"{'#':>4} {'Entry Time':>16} {'Dir':>4} {'Entry':>10} {'Exit':>10} {'P/L($)':>8} {'Result':>6} {'Exit Reason':>18} {'StochK':>7} {'Session':>20}")
    lines.append("-" * 140)
    for idx, t in enumerate(stats.trades, 1):
        lines.append(
            f"{idx:4d} {t.entry_time.strftime('%Y-%m-%d %H:%M'):>16} {t.direction:>4} "
            f"{t.entry_price:>10.2f} {t.exit_price:>10.2f} {t.profit_usd:>8.2f} "
            f"{t.result.value:>6} {t.exit_reason.value:>18} {t.stoch_k:>7.1f} "
            f"{t.session:>20}"
        )
    lines.append("\n" + "=" * 80)
    lines.append("END OF REPORT")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Log saved: {filepath}")


# ─── Main ──────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("XAUBOT AI — SMC + Stoch + Sell + Broker SL Only Exit Backtest")
    print("Entry: SMC-Only v4 + Stochastic + Sell Filter")
    print(f"Filter 1: Stochastic (K={STOCH_K_PERIOD}, OB>{STOCH_OVERBOUGHT} block BUY, OS<{STOCH_OVERSOLD} block SELL)")
    print(f"Filter 2: Sell Filter (SELL requires ML agree + conf >= {SELL_FILTER_MIN_ML_CONF:.0%})")
    print("Exit: Broker SL Only (simplified — no breakeven, wider trail, 12h max)")
    print("=" * 70)

    config = get_config()
    mt5 = MT5Connector(
        login=config.mt5_login, password=config.mt5_password,
        server=config.mt5_server, path=config.mt5_path,
    )
    mt5.connect()
    print(f"\nConnected to MT5")

    print("Fetching XAUUSD M15 historical data...")
    df = mt5.get_market_data(symbol="XAUUSD", timeframe="M15", count=50000)

    if len(df) == 0:
        print("ERROR: No data received")
        mt5.disconnect()
        return

    print(f"  Received {len(df)} bars")
    times = df["time"].to_list()
    print(f"  Data range: {times[0]} to {times[-1]}")

    end_date = datetime.now()
    start_date = datetime(2025, 8, 1)

    data_start = times[0]
    if hasattr(data_start, 'replace') and data_start.tzinfo:
        start_date = start_date.replace(tzinfo=data_start.tzinfo)
        end_date = end_date.replace(tzinfo=data_start.tzinfo)

    if data_start > start_date:
        start_date = data_start + timedelta(days=5)
        print(f"  [INFO] Adjusted start: {start_date}")

    print(f"\n  Backtest period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

    print("\nCalculating indicators...")
    features = FeatureEngineer()
    smc = SMCAnalyzer(swing_length=config.smc.swing_length, ob_lookback=config.smc.ob_lookback)

    df = features.calculate_all(df, include_ml_features=True)
    df = smc.calculate_all(df)

    print("  Calculating Stochastic Oscillator...")
    df = calculate_stochastic(df, k_period=STOCH_K_PERIOD, d_period=STOCH_D_PERIOD)

    regime_detector = MarketRegimeDetector(model_path="models/hmm_regime.pkl")
    try:
        regime_detector.load()
        df = regime_detector.predict(df)
        print("  HMM regime loaded")
    except Exception:
        print("  [WARN] HMM not available")

    print("  Indicators calculated")

    backtest = StochSellBrokerSLBacktest(
        capital=5000.0,
        max_daily_loss_percent=5.0,
        max_loss_per_trade_percent=1.0,
        base_lot_size=0.01,
        max_lot_size=0.02,
        recovery_lot_size=0.01,
        trade_cooldown_bars=10,
    )

    stats = backtest.run(df=df, start_date=start_date, end_date=end_date, initial_capital=5000.0)

    net_pnl = stats.total_profit - stats.total_loss

    print("\n" + "=" * 70)
    print("SMC + STOCH + SELL + BROKER SL ONLY — RESULTS")
    print("=" * 70)

    print(f"\n  Filter Stats:")
    print(f"    Stochastic blocked:  {stats.stoch_filtered}")
    print(f"      BUY overbought:   {stats.stoch_filtered_buy_overbought}")
    print(f"      SELL oversold:    {stats.stoch_filtered_sell_oversold}")
    print(f"    Sell Filter blocked: {stats.sell_filtered}")
    print(f"      ML disagree:      {stats.sell_filtered_no_ml_agree}")
    print(f"      Low ML conf:      {stats.sell_filtered_low_conf}")
    print(f"    Combined blocked:    {stats.stoch_filtered + stats.sell_filtered}")

    print(f"\n  Performance:")
    print(f"    Total Trades:     {stats.total_trades}")
    print(f"    Wins:             {stats.wins}")
    print(f"    Losses:           {stats.losses}")
    print(f"    Win Rate:         {stats.win_rate:.1f}%")

    print(f"\n  Profit/Loss:")
    print(f"    Total Profit:     ${stats.total_profit:,.2f}")
    print(f"    Total Loss:       ${stats.total_loss:,.2f}")
    print(f"    Net PnL:          ${net_pnl:,.2f}")
    print(f"    Profit Factor:    {stats.profit_factor:.2f}")

    print(f"\n  Risk Metrics:")
    print(f"    Max Drawdown:     {stats.max_drawdown:.1f}% (${stats.max_drawdown_usd:,.2f})")
    print(f"    Avg Win:          ${stats.avg_win:,.2f}")
    print(f"    Avg Loss:         ${stats.avg_loss:,.2f}")
    print(f"    Expectancy:       ${stats.expectancy:,.2f}")
    print(f"    Sharpe Ratio:     {stats.sharpe_ratio:.2f}")

    print(f"\n  Sync Metrics:")
    print(f"    Avoided (AVOID):  {stats.avoided_signals}")
    print(f"    Recovery Trades:  {stats.recovery_mode_trades}")
    print(f"    Daily Limit Stops:{stats.daily_limit_stops}")

    print(f"\n  Exit Reasons:")
    exit_counts = {}
    for t in stats.trades:
        r = t.exit_reason.value
        exit_counts[r] = exit_counts.get(r, 0) + 1
    for reason, count in sorted(exit_counts.items(), key=lambda x: -x[1]):
        pct = count / stats.total_trades * 100 if stats.total_trades > 0 else 0
        print(f"    {reason:20s}: {count} ({pct:.1f}%)")

    print(f"\n  Direction:")
    for d in ["BUY", "SELL"]:
        dt = [t for t in stats.trades if t.direction == d]
        dw = sum(1 for t in dt if t.result == TradeResult.WIN)
        dp = sum(t.profit_usd for t in dt)
        dwr = dw / len(dt) * 100 if dt else 0
        print(f"    {d}: {len(dt)} trades, {dwr:.1f}% WR, ${dp:,.2f}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "12_stoch_sell_broker_sl_results")
    os.makedirs(output_dir, exist_ok=True)

    log_path = os.path.join(output_dir, f"stoch_sell_broker_sl_{timestamp}.log")
    xlsx_path = os.path.join(output_dir, f"stoch_sell_broker_sl_{timestamp}.xlsx")

    generate_log(stats, log_path, start_date, end_date)
    generate_xlsx_report(stats, xlsx_path, start_date, end_date)

    mt5.disconnect()

    print("\n" + "=" * 70)
    print(f"Output: {output_dir}")
    print(f"  Log:    {os.path.basename(log_path)}")
    print(f"  Report: {os.path.basename(xlsx_path)}")
    print("=" * 70)
    print("Backtest complete!")


if __name__ == "__main__":
    main()
