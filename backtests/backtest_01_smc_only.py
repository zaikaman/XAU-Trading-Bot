"""
Backtest SMC-Only — 100% Synced with main_live.py Signal Logic v4
==================================================================
All 3 exit systems replicated:
  A) SmartPositionManager  — breakeven, trailing SL, peak drawdown, market close
  B) SmartRiskManager      — momentum TP, early cut, stall, daily limit, recovery mode
  C) Time/Trend exit       — timeout 4h/6h/8h, ATR trend reversal

Entry: SMC-Only (no ML gate, no persistence, no pullback filter)
Filters: DynamicConfidence AVOID, Regime CRISIS, Session filter, Weekend

Usage:
    python backtests/backtest_smc_only.py
"""

import polars as pl
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, field
from enum import Enum
import sys
import os
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.mt5_connector import MT5Connector
from src.smc_polars import SMCAnalyzer, SMCSignal
from src.feature_eng import FeatureEngineer
from src.regime_detector import MarketRegimeDetector, MarketRegime
from src.ml_model import TradingModel
from src.config import get_config
from src.dynamic_confidence import DynamicConfidenceManager, create_dynamic_confidence, MarketQuality
from loguru import logger

logger.remove()
logger.add(sys.stderr, level="WARNING")

WIB = ZoneInfo("Asia/Jakarta")


# ─── Enums & Dataclasses ──────────────────────────────────────

class TradeResult(Enum):
    WIN = "WIN"
    LOSS = "LOSS"
    BREAKEVEN = "BREAKEVEN"


class ExitReason(Enum):
    TAKE_PROFIT = "take_profit"
    SMART_TP = "smart_tp"              # Momentum-based TP (SmartRiskManager)
    PEAK_PROTECT = "peak_protect"      # Peak profit protection
    EARLY_EXIT = "early_exit"          # Small profit + reversal signal
    EARLY_CUT = "early_cut"            # Loss + negative momentum
    MAX_LOSS = "max_loss"              # 50% of max_loss_per_trade ($25)
    STALL = "stall"                    # Price stalled with loss
    TREND_REVERSAL = "trend_reversal"  # ATR momentum + ML reversal
    TIMEOUT = "timeout"                # 4h/6h/8h smart timeout
    WEEKEND_CLOSE = "weekend_close"    # Near weekend close
    TRAILING_SL = "trailing_sl"        # Hit trailing SL
    BREAKEVEN_EXIT = "breakeven_exit"  # Hit breakeven SL
    DAILY_LIMIT = "daily_limit"        # Daily loss limit hit
    REGIME_DANGER = "regime_danger"    # Regime change to crisis/high_vol
    MARKET_SIGNAL = "market_signal"    # RSI/trend opposite signal


class TradingMode(Enum):
    NORMAL = "normal"
    RECOVERY = "recovery"
    PROTECTED = "protected"
    STOPPED = "stopped"


@dataclass
class SimulatedTrade:
    ticket: int
    entry_time: datetime
    exit_time: datetime
    direction: str
    entry_price: float
    exit_price: float
    stop_loss: float
    take_profit: float
    lot_size: float
    profit_usd: float
    profit_pips: float
    result: TradeResult
    exit_reason: ExitReason
    smc_confidence: float
    regime: str
    session: str
    signal_reason: str
    has_bos: bool = False
    has_choch: bool = False
    has_fvg: bool = False
    has_ob: bool = False
    atr_at_entry: float = 0.0
    rr_ratio: float = 0.0
    trading_mode: str = "normal"


@dataclass
class BacktestStats:
    total_trades: int = 0
    wins: int = 0
    losses: int = 0
    total_profit: float = 0.0
    total_loss: float = 0.0
    max_drawdown: float = 0.0
    max_drawdown_usd: float = 0.0
    win_rate: float = 0.0
    profit_factor: float = 0.0
    avg_win: float = 0.0
    avg_loss: float = 0.0
    avg_trade: float = 0.0
    expectancy: float = 0.0
    sharpe_ratio: float = 0.0
    trades: List[SimulatedTrade] = field(default_factory=list)
    equity_curve: List[float] = field(default_factory=list)
    avoided_signals: int = 0       # Signals blocked by AVOID filter
    daily_limit_stops: int = 0     # Days stopped by daily loss limit
    recovery_mode_trades: int = 0  # Trades in RECOVERY mode


# ─── SMC-Only Backtest (100% Synced) ──────────────────────────

class SMCOnlyBacktest:
    """100% synced with main_live.py Signal Logic v4 + all exit systems."""

    def __init__(
        self,
        capital: float = 5000.0,
        # SmartRiskManager params (synced)
        max_daily_loss_percent: float = 5.0,
        max_loss_per_trade_percent: float = 1.0,
        base_lot_size: float = 0.01,
        max_lot_size: float = 0.02,   # reduced from 0.03
        recovery_lot_size: float = 0.01,
        trend_reversal_threshold: float = 0.75,
        max_concurrent_positions: int = 2,
        # SmartPositionManager params (synced with main_live.py init)
        breakeven_pips: float = 30.0,      # $3 profit
        trail_start_pips: float = 50.0,    # $5 profit
        trail_step_pips: float = 30.0,     # $3 trail distance
        min_profit_to_protect: float = 5.0,
        max_drawdown_from_peak: float = 50.0,  # 50% drawdown
        # Other
        trade_cooldown_bars: int = 10,
        trend_reversal_mult: float = 0.6,
    ):
        self.capital = capital
        self.max_daily_loss_usd = capital * (max_daily_loss_percent / 100)
        self.max_loss_per_trade = capital * (max_loss_per_trade_percent / 100)
        self.base_lot_size = base_lot_size
        self.max_lot_size = max_lot_size
        self.recovery_lot_size = recovery_lot_size
        self.trend_reversal_threshold = trend_reversal_threshold
        self.max_concurrent_positions = max_concurrent_positions
        self.breakeven_pips = breakeven_pips
        self.trail_start_pips = trail_start_pips
        self.trail_step_pips = trail_step_pips
        self.min_profit_to_protect = min_profit_to_protect
        self.max_drawdown_from_peak = max_drawdown_from_peak
        self.trade_cooldown_bars = trade_cooldown_bars
        self.trend_reversal_mult = trend_reversal_mult

        config = get_config()
        self.smc = SMCAnalyzer(
            swing_length=config.smc.swing_length,
            ob_lookback=config.smc.ob_lookback,
        )
        self.features = FeatureEngineer()
        self.dynamic_confidence = create_dynamic_confidence()

        # ML model for exit evaluation (synced: ML used for exits even in SMC-only)
        self.ml_model = TradingModel(model_path="models/xgboost_model.pkl")
        try:
            self.ml_model.load()
            print("  ML model loaded (for exit evaluation)")
        except Exception:
            print("  [WARN] ML model not loaded — exit ML checks disabled")

        self.regime_detector = MarketRegimeDetector(model_path="models/hmm_regime.pkl")
        try:
            self.regime_detector.load()
        except Exception:
            print("  [WARN] HMM model not loaded")

        self._ticket_counter = 2000000

    # ── Session filter (synced) ──

    def _get_session_from_time(self, dt: datetime) -> Tuple[str, bool, float]:
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib_time = dt.astimezone(WIB)
        hour = wib_time.hour
        if 6 <= hour < 15:
            return "Sydney-Tokyo", True, 0.5
        elif 15 <= hour < 16:
            return "Tokyo-London Overlap", True, 0.75
        elif 16 <= hour < 19:
            return "London Early", True, 0.8
        elif 19 <= hour < 24:
            return "London-NY Overlap (Golden)", True, 1.0
        elif 0 <= hour < 4:
            return "NY Session", True, 0.9
        else:
            return "Off Hours", False, 0.0

    def _hours_to_golden(self, dt: datetime) -> float:
        """Hours until golden time (19:00 WIB). Returns 0 if already in golden."""
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib = dt.astimezone(WIB)
        if 19 <= wib.hour < 24:
            return 0
        target = wib.replace(hour=19, minute=0, second=0, microsecond=0)
        if wib.hour >= 19:
            target += timedelta(days=1)
        return max(0, (target - wib).total_seconds() / 3600)

    def _is_near_weekend_close(self, dt: datetime) -> bool:
        """Check if near weekend market close (Saturday 04:30+ WIB)."""
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib = dt.astimezone(WIB)
        if wib.weekday() == 5 and wib.hour >= 4 and wib.minute >= 30:
            return True
        # Friday night very late (after midnight = Saturday early)
        return False

    # ── SmartRiskManager: Lot sizing with RECOVERY mode (synced) ──

    def _calculate_lot_size(
        self,
        confidence: float,
        regime: str,
        trading_mode: TradingMode,
        session_mult: float,
    ) -> float:
        """Synced with SmartRiskManager.calculate_lot_size()"""
        if trading_mode == TradingMode.STOPPED:
            return 0

        lot = self.base_lot_size

        if trading_mode in (TradingMode.RECOVERY, TradingMode.PROTECTED):
            lot = self.recovery_lot_size
        else:
            # ML confidence-based sizing (using SMC confidence as proxy)
            if confidence >= 0.65:
                lot = self.max_lot_size
            elif confidence >= 0.55:
                lot = self.base_lot_size
            else:
                lot = self.recovery_lot_size

        # Regime override
        if regime.lower() in ["high_volatility", "crisis"]:
            lot = self.recovery_lot_size

        # Session multiplier
        lot = max(0.01, lot * session_mult)
        return round(lot, 2)

    # ── Full exit simulation (all 3 systems) ──

    def _simulate_trade_exit(
        self,
        df: pl.DataFrame,
        entry_idx: int,
        direction: str,
        entry_price: float,
        take_profit: float,
        stop_loss: float,
        lot_size: float,
        daily_loss_so_far: float,
        feature_cols: list,
        max_bars: int = 100,
    ) -> Tuple[float, float, ExitReason, int, float]:
        """
        Simulate trade exit with ALL 3 exit systems synced with main_live.py:
        A) SmartPositionManager (breakeven, trailing, peak protect, market signal)
        B) SmartRiskManager (smart TP, early cut, stall, daily limit, reversal)
        C) Time/Trend exit (4h/6h/8h timeout, ATR momentum)
        """
        pip_value = 10  # XAUUSD: 1 pip = $10 per lot

        highs = df["high"].to_list()
        lows = df["low"].to_list()
        closes = df["close"].to_list()
        times = df["time"].to_list()

        # ATR at entry
        atr = 12.0
        if "atr" in df.columns:
            atr_list = df["atr"].to_list()
            if entry_idx < len(atr_list) and atr_list[entry_idx] is not None:
                atr = atr_list[entry_idx]

        reversal_momentum_threshold = atr * self.trend_reversal_mult
        min_loss_for_reversal_exit = atr * 0.8

        # ── State tracking (simulating SmartRiskManager PositionGuard) ──
        profit_history = []
        price_history = []
        peak_profit = 0.0
        stall_count = 0
        reversal_warnings = 0

        # SmartPositionManager state
        current_sl = stop_loss  # broker SL (mutable via trailing)
        breakeven_moved = False

        # Target TP profit for probability estimation
        if direction == "BUY":
            target_tp_profit = (take_profit - entry_price) / 0.1 * pip_value * lot_size
        else:
            target_tp_profit = (entry_price - take_profit) / 0.1 * pip_value * lot_size

        # ML prediction cache (evaluate every 4 bars like live)
        cached_ml_signal = ""
        cached_ml_confidence = 0.5

        for i in range(entry_idx + 1, min(entry_idx + max_bars, len(df))):
            high = highs[i]
            low = lows[i]
            close = closes[i]
            current_time = times[i]

            # Current P/L
            if direction == "BUY":
                current_pips = (close - entry_price) / 0.1
                pip_profit_from_entry = (close - entry_price) / 0.1
            else:
                current_pips = (entry_price - close) / 0.1
                pip_profit_from_entry = (entry_price - close) / 0.1
            current_profit = current_pips * pip_value * lot_size

            # Track history
            profit_history.append(current_profit)
            price_history.append(close)
            if current_profit > peak_profit:
                peak_profit = current_profit

            bars_since_entry = i - entry_idx

            # ── ML prediction (every 4 bars, synced with live) ──
            if bars_since_entry % 4 == 0 and self.ml_model.fitted:
                try:
                    df_slice = df.head(i + 1)
                    ml_pred = self.ml_model.predict(df_slice, feature_cols)
                    cached_ml_signal = ml_pred.signal
                    cached_ml_confidence = ml_pred.confidence
                except Exception:
                    pass

            # ── Momentum calculation (synced with PositionGuard.calculate_momentum) ──
            momentum = 0.0
            if len(profit_history) >= 3:
                recent = profit_history[-5:] if len(profit_history) >= 5 else profit_history
                profit_change = recent[-1] - recent[0]
                momentum = max(-100, min(100, (profit_change / 10) * 50))

            profit_growing = momentum > 0

            # ════════════════════════════════════════════════
            # A) SmartPositionManager checks (every bar)
            # ════════════════════════════════════════════════

            # A.0 TP hit by price action (high/low)
            if direction == "BUY" and high >= take_profit:
                pips = (take_profit - entry_price) / 0.1
                profit = pips * pip_value * lot_size
                return profit, pips, ExitReason.TAKE_PROFIT, i, take_profit
            elif direction == "SELL" and low <= take_profit:
                pips = (entry_price - take_profit) / 0.1
                profit = pips * pip_value * lot_size
                return profit, pips, ExitReason.TAKE_PROFIT, i, take_profit

            # A.0b Trailing SL hit check
            if breakeven_moved and current_sl > 0:
                if direction == "BUY" and low <= current_sl:
                    pips = (current_sl - entry_price) / 0.1
                    profit = pips * pip_value * lot_size
                    reason = ExitReason.TRAILING_SL if pip_profit_from_entry >= self.trail_start_pips else ExitReason.BREAKEVEN_EXIT
                    return profit, pips, reason, i, current_sl
                elif direction == "SELL" and high >= current_sl:
                    pips = (entry_price - current_sl) / 0.1
                    profit = pips * pip_value * lot_size
                    reason = ExitReason.TRAILING_SL if pip_profit_from_entry >= self.trail_start_pips else ExitReason.BREAKEVEN_EXIT
                    return profit, pips, reason, i, current_sl

            # A.1 Breakeven move (after 30 pips / $3 profit)
            if pip_profit_from_entry >= self.breakeven_pips and not breakeven_moved:
                if direction == "BUY":
                    current_sl = entry_price + 2  # 2 points buffer
                else:
                    current_sl = entry_price - 2
                breakeven_moved = True

            # A.2 Trailing SL (after 50 pips / $5 profit)
            if pip_profit_from_entry >= self.trail_start_pips:
                trail_distance = self.trail_step_pips * 0.1
                if direction == "BUY":
                    new_trail_sl = close - trail_distance
                    if new_trail_sl > current_sl:
                        current_sl = new_trail_sl
                else:
                    new_trail_sl = close + trail_distance
                    if current_sl == 0 or new_trail_sl < current_sl:
                        current_sl = new_trail_sl

            # A.3 Peak profit drawdown protection (50% drawdown from peak for $5+ profit)
            if peak_profit > self.min_profit_to_protect:
                drawdown_pct = ((peak_profit - current_profit) / peak_profit) * 100 if peak_profit > 0 else 0
                if drawdown_pct > self.max_drawdown_from_peak:
                    return current_profit, current_pips, ExitReason.PEAK_PROTECT, i, close

            # A.4 Market analysis: trend + momentum + RSI (synced with position_manager)
            if bars_since_entry % 5 == 0 and bars_since_entry >= 5:
                # Trend analysis (5-bar vs 20-bar MA)
                if i >= 20:
                    ma_fast = np.mean(closes[i-4:i+1])
                    ma_slow = np.mean(closes[i-19:i+1])
                    trend = "NEUTRAL"
                    if ma_fast > ma_slow * 1.001:
                        trend = "BULLISH"
                    elif ma_fast < ma_slow * 0.999:
                        trend = "BEARISH"

                    # ROC momentum
                    roc = (closes[i] / closes[max(0,i-4)] - 1) * 100
                    mom_dir = "BULLISH" if roc > 0.3 else ("BEARISH" if roc < -0.3 else "NEUTRAL")

                    # RSI check
                    rsi_val = None
                    if "rsi" in df.columns:
                        rsi_list = df["rsi"].to_list()
                        if i < len(rsi_list):
                            rsi_val = rsi_list[i]

                    urgency = 0
                    should_exit = False

                    # Strong ML opposite signal
                    if cached_ml_confidence > 0.75:
                        if direction == "BUY" and cached_ml_signal == "SELL":
                            should_exit = True
                            urgency += 2
                        elif direction == "SELL" and cached_ml_signal == "BUY":
                            should_exit = True
                            urgency += 2

                    # RSI extremes
                    if rsi_val:
                        if rsi_val > 75 and direction == "BUY":
                            should_exit = True
                            urgency += 2
                        elif rsi_val < 25 and direction == "SELL":
                            should_exit = True
                            urgency += 2

                    # Trend + momentum reversal
                    if direction == "BUY" and trend == "BEARISH" and mom_dir == "BEARISH":
                        should_exit = True
                        urgency += 3
                    elif direction == "SELL" and trend == "BULLISH" and mom_dir == "BULLISH":
                        should_exit = True
                        urgency += 3

                    # Close on strong opposite signal with profit (synced)
                    if should_exit and current_profit > self.min_profit_to_protect / 2:
                        return current_profit, current_pips, ExitReason.MARKET_SIGNAL, i, close

                    # High urgency with any profit
                    if urgency >= 7 and current_profit > 0:
                        return current_profit, current_pips, ExitReason.MARKET_SIGNAL, i, close

            # A.5 Weekend close check
            if self._is_near_weekend_close(current_time):
                if current_profit > 0:
                    return current_profit, current_pips, ExitReason.WEEKEND_CLOSE, i, close
                elif current_profit > -10:
                    return current_profit, current_pips, ExitReason.WEEKEND_CLOSE, i, close

            # ════════════════════════════════════════════════
            # B) SmartRiskManager checks
            # ════════════════════════════════════════════════

            # B.1 Smart TP ($15+ with momentum analysis — synced evaluate_position CHECK 1)
            if current_profit >= 15:
                # Hard TP at $40
                if current_profit >= 40:
                    return current_profit, current_pips, ExitReason.SMART_TP, i, close

                # Momentum-based TP: profit $25+ but momentum dropping
                if current_profit >= 25 and momentum < -30:
                    return current_profit, current_pips, ExitReason.SMART_TP, i, close

                # Peak protection: profit turun ke 60% dari peak
                if peak_profit > 30 and current_profit < peak_profit * 0.6:
                    return current_profit, current_pips, ExitReason.PEAK_PROTECT, i, close

                # Low TP probability: profit $20+ tapi kemungkinan TP rendah
                if current_profit >= 20:
                    # Simplified TP probability (synced with PositionGuard.get_tp_probability)
                    progress = (current_profit / target_tp_profit) * 100 if target_tp_profit > 0 else 0
                    progress_score = min(40, max(0, progress * 0.4))
                    momentum_score = ((momentum + 100) / 200) * 30
                    time_penalty = min(10, bars_since_entry / 4 * 2)  # 2 points per hour
                    tp_probability = progress_score + momentum_score + 10 - time_penalty
                    if tp_probability < 25:
                        return current_profit, current_pips, ExitReason.SMART_TP, i, close

            # B.2 Smart Early Exit ($5-15 profit + reversal, synced CHECK 2)
            if 5 <= current_profit < 15:
                if momentum < -50 and cached_ml_confidence >= 0.65:
                    is_reversal = (
                        (direction == "BUY" and cached_ml_signal == "SELL") or
                        (direction == "SELL" and cached_ml_signal == "BUY")
                    )
                    if is_reversal:
                        return current_profit, current_pips, ExitReason.EARLY_EXIT, i, close

            # B.3 Early cut: loss significant + momentum negative (synced CHECK 3)
            if current_profit < 0:
                loss_percent_of_max = abs(current_profit) / self.max_loss_per_trade * 100
                if momentum < -30 and loss_percent_of_max >= 30:
                    return current_profit, current_pips, ExitReason.EARLY_CUT, i, close

            # B.4 Trend Reversal: ML 75%+ opposite (synced CHECK 4)
            is_ml_reversal = False
            if direction == "BUY" and cached_ml_signal == "SELL" and cached_ml_confidence >= self.trend_reversal_threshold:
                is_ml_reversal = True
                reversal_warnings += 1
            elif direction == "SELL" and cached_ml_signal == "BUY" and cached_ml_confidence >= self.trend_reversal_threshold:
                is_ml_reversal = True
                reversal_warnings += 1

            loss_moderate = abs(current_profit) > (self.max_loss_per_trade * 0.4)
            if is_ml_reversal and current_profit < -8 and loss_moderate:
                return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close

            if reversal_warnings >= 3 and current_profit < -10:
                return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close

            # B.5 Max loss per trade — 50% of max (synced CHECK 5)
            if current_profit <= -(self.max_loss_per_trade * 0.50):
                # Last chance hold if golden time very close (synced)
                htg = self._hours_to_golden(current_time)
                if htg <= 1 and htg > 0 and momentum > -40:
                    pass  # Hold — last chance for recovery
                else:
                    return current_profit, current_pips, ExitReason.MAX_LOSS, i, close

            # B.6 Stall detection (synced CHECK 5b)
            if len(profit_history) >= 10:
                recent_range = max(profit_history[-10:]) - min(profit_history[-10:])
                if recent_range < 3 and current_profit < -15:
                    stall_count += 1
                    if stall_count >= 5:
                        return current_profit, current_pips, ExitReason.STALL, i, close

            # B.7 Daily loss limit (synced CHECK 6)
            potential_daily_loss = daily_loss_so_far + abs(min(0, current_profit))
            if potential_daily_loss >= self.max_daily_loss_usd:
                return current_profit, current_pips, ExitReason.DAILY_LIMIT, i, close

            # ════════════════════════════════════════════════
            # C) Time-based exit (synced CHECK 8)
            # ════════════════════════════════════════════════

            # Check ML agreement for timeout decision
            ml_agrees = (
                (direction == "BUY" and cached_ml_signal == "BUY") or
                (direction == "SELL" and cached_ml_signal == "SELL")
            )

            # 4+ hours: exit if stuck (synced)
            if bars_since_entry >= 16:
                if current_profit < 5 and not profit_growing:
                    if current_profit >= 0:
                        return current_profit, current_pips, ExitReason.TIMEOUT, i, close
                    elif current_profit > -15:
                        return current_profit, current_pips, ExitReason.TIMEOUT, i, close

            # 6+ hours: exit unless significantly profitable AND growing (synced)
            if bars_since_entry >= 24:
                if current_profit < 10 or not profit_growing:
                    return current_profit, current_pips, ExitReason.TIMEOUT, i, close

            # 8+ hours: hard max (synced)
            if bars_since_entry >= 32:
                return current_profit, current_pips, ExitReason.TIMEOUT, i, close

            # C.2 ATR trend reversal (synced with original backtest)
            if bars_since_entry > 10:
                recent_closes = closes[i-5:i+1]
                mom = recent_closes[-1] - recent_closes[0]
                if direction == "BUY" and mom < -reversal_momentum_threshold:
                    if current_profit < -min_loss_for_reversal_exit:
                        return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close
                elif direction == "SELL" and mom > reversal_momentum_threshold:
                    if current_profit < -min_loss_for_reversal_exit:
                        return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close

        # End of data — close at last price
        final_idx = min(entry_idx + max_bars - 1, len(df) - 1)
        final_price = closes[final_idx]
        if direction == "BUY":
            pips = (final_price - entry_price) / 0.1
        else:
            pips = (entry_price - final_price) / 0.1
        profit = pips * pip_value * lot_size
        return profit, pips, ExitReason.TIMEOUT, final_idx, final_price

    # ── Main backtest run ──

    def run(
        self,
        df: pl.DataFrame,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        initial_capital: float = 5000.0,
    ) -> BacktestStats:
        stats = BacktestStats()
        capital = initial_capital
        peak_capital = initial_capital
        stats.equity_curve.append(capital)

        # SmartRiskManager state tracking
        daily_loss = 0.0
        daily_profit = 0.0
        daily_trades = 0
        consecutive_losses = 0
        trading_mode = TradingMode.NORMAL
        current_date = None

        # Feature columns for ML predictions
        feature_cols = []
        if self.ml_model.fitted and self.ml_model.feature_names:
            feature_cols = [f for f in self.ml_model.feature_names if f in df.columns]

        times = df["time"].to_list()
        start_idx = next((i for i, t in enumerate(times) if t >= start_date), 100) if start_date else 100
        end_idx = next((i for i, t in enumerate(times) if t > end_date), len(df) - 100) if end_date else len(df) - 100

        last_trade_idx = -self.trade_cooldown_bars * 2

        print(f"\n  Running SMC-Only backtest (100% synced)...")
        print(f"  Date range: {times[start_idx]} to {times[end_idx - 1]}")
        print(f"  Total bars: {end_idx - start_idx}")

        for i in range(start_idx, end_idx):
            # Cooldown
            if i - last_trade_idx < self.trade_cooldown_bars:
                continue

            current_time = times[i]

            # ── Daily reset (synced with SmartRiskManager.check_new_day) ──
            trade_date = current_time.date() if hasattr(current_time, 'date') else current_time
            if current_date is None or trade_date != current_date:
                if daily_loss > 0 and current_date is not None:
                    pass  # Could log daily summary
                daily_loss = 0.0
                daily_profit = 0.0
                daily_trades = 0
                current_date = trade_date
                # Reset mode unless consecutive losses persist
                if consecutive_losses < 2:
                    trading_mode = TradingMode.NORMAL

            # ── Trading mode check (synced) ──
            if trading_mode == TradingMode.STOPPED:
                continue

            # Session filter
            session_name, can_trade, lot_mult = self._get_session_from_time(current_time)
            if not can_trade:
                continue

            # Skip weekends
            if hasattr(current_time, 'weekday') and current_time.weekday() >= 5:
                continue

            df_slice = df.head(i + 1)

            # Regime check — CRISIS and SLEEP (synced)
            regime = "normal"
            regime_state = None
            try:
                if self.regime_detector.fitted:
                    regime_state = self.regime_detector.get_current_state(df_slice)
                    if regime_state:
                        regime = regime_state.regime.value
                        if regime_state.regime == MarketRegime.CRISIS:
                            continue
                        if regime_state.recommendation == "SLEEP":
                            continue
            except Exception:
                pass

            # ═══ DYNAMIC CONFIDENCE — AVOID filter (synced with _combine_signals) ═══
            try:
                # Get ML prediction for dynamic confidence analysis
                ml_signal = ""
                ml_confidence = 0.5
                if self.ml_model.fitted and feature_cols:
                    ml_pred = self.ml_model.predict(df_slice, feature_cols)
                    ml_signal = ml_pred.signal
                    ml_confidence = ml_pred.confidence

                market_analysis = self.dynamic_confidence.analyze_market(
                    session=session_name,
                    regime=regime,
                    volatility="medium",
                    trend_direction=regime,
                    has_smc_signal=True,
                    ml_signal=ml_signal,
                    ml_confidence=ml_confidence,
                )

                if market_analysis.quality == MarketQuality.AVOID:
                    stats.avoided_signals += 1
                    continue
            except Exception:
                pass

            # ═══ SMC SIGNAL ═══
            try:
                smc_signal = self.smc.generate_signal(df_slice)
            except Exception:
                continue

            if smc_signal is None:
                continue

            # ═══ NO ML gate, NO persistence, NO pullback — SMC-Only v4 ═══

            # SMC details
            recent_df = df_slice.tail(10)
            recent_bos = recent_df["bos"].to_list() if "bos" in df_slice.columns else []
            recent_choch = recent_df["choch"].to_list() if "choch" in df_slice.columns else []
            recent_fvg_bull = recent_df["is_fvg_bull"].to_list() if "is_fvg_bull" in df_slice.columns else []
            recent_fvg_bear = recent_df["is_fvg_bear"].to_list() if "is_fvg_bear" in df_slice.columns else []
            recent_obs = recent_df["ob"].to_list() if "ob" in df_slice.columns else []

            has_bos = 1 in recent_bos or -1 in recent_bos
            has_choch = 1 in recent_choch or -1 in recent_choch
            has_fvg = any(recent_fvg_bull) or any(recent_fvg_bear)
            has_ob = 1 in recent_obs or -1 in recent_obs

            atr_at_entry = 12.0
            if "atr" in df_slice.columns:
                atr_val = df_slice.tail(1)["atr"].item()
                if atr_val is not None and atr_val > 0:
                    atr_at_entry = atr_val

            # ═══ Confidence (synced with _combine_signals) ═══
            confidence = smc_signal.confidence
            # ML agrees → average confidence (synced)
            ml_agrees = (
                (smc_signal.signal_type == "BUY" and ml_signal == "BUY") or
                (smc_signal.signal_type == "SELL" and ml_signal == "SELL")
            )
            if ml_agrees:
                confidence = (smc_signal.confidence + ml_confidence) / 2

            # High vol adjustment (synced)
            if regime == "high_volatility":
                confidence *= 0.9

            # ═══ Lot size with RECOVERY mode (synced) ═══
            lot_size = self._calculate_lot_size(confidence, regime, trading_mode, lot_mult)
            if lot_size <= 0:
                continue

            if trading_mode == TradingMode.RECOVERY:
                stats.recovery_mode_trades += 1

            # ═══ Execute trade ═══
            entry_price = smc_signal.entry_price
            take_profit_price = smc_signal.take_profit
            stop_loss_price = smc_signal.stop_loss
            risk = abs(entry_price - stop_loss_price)
            rr = abs(take_profit_price - entry_price) / risk if risk > 0 else 0

            profit, pips, exit_reason, exit_idx, exit_price = self._simulate_trade_exit(
                df=df,
                entry_idx=i,
                direction=smc_signal.signal_type,
                entry_price=entry_price,
                take_profit=take_profit_price,
                stop_loss=stop_loss_price,
                lot_size=lot_size,
                daily_loss_so_far=daily_loss,
                feature_cols=feature_cols,
            )

            # Record trade
            self._ticket_counter += 1
            result = TradeResult.WIN if profit > 0 else (TradeResult.LOSS if profit < 0 else TradeResult.BREAKEVEN)

            trade = SimulatedTrade(
                ticket=self._ticket_counter,
                entry_time=current_time,
                exit_time=times[exit_idx] if exit_idx < len(times) else times[-1],
                direction=smc_signal.signal_type,
                entry_price=entry_price,
                exit_price=exit_price,
                stop_loss=stop_loss_price,
                take_profit=take_profit_price,
                lot_size=lot_size,
                profit_usd=profit,
                profit_pips=pips,
                result=result,
                exit_reason=exit_reason,
                smc_confidence=confidence,
                regime=regime,
                session=session_name,
                signal_reason=smc_signal.reason,
                has_bos=has_bos,
                has_choch=has_choch,
                has_fvg=has_fvg,
                has_ob=has_ob,
                atr_at_entry=atr_at_entry,
                rr_ratio=rr,
                trading_mode=trading_mode.value,
            )
            stats.trades.append(trade)

            # ── Update SmartRiskManager state (synced record_trade_result) ──
            stats.total_trades += 1
            daily_trades += 1
            capital += profit

            if profit > 0:
                stats.wins += 1
                stats.total_profit += profit
                daily_profit += profit
                consecutive_losses = 0
                if trading_mode == TradingMode.RECOVERY:
                    trading_mode = TradingMode.NORMAL
            else:
                stats.losses += 1
                stats.total_loss += abs(profit)
                daily_loss += abs(profit)
                consecutive_losses += 1

            # Mode transitions (synced with SmartRiskManager._update_state)
            if daily_loss >= self.max_daily_loss_usd:
                trading_mode = TradingMode.STOPPED
                stats.daily_limit_stops += 1
            elif consecutive_losses >= 3 or daily_loss >= self.max_daily_loss_usd * 0.6:
                trading_mode = TradingMode.PROTECTED
            elif consecutive_losses >= 2:
                trading_mode = TradingMode.RECOVERY

            # Drawdown
            if capital > peak_capital:
                peak_capital = capital
            drawdown_pct = (peak_capital - capital) / peak_capital * 100
            drawdown_usd = peak_capital - capital
            if drawdown_pct > stats.max_drawdown:
                stats.max_drawdown = drawdown_pct
                stats.max_drawdown_usd = drawdown_usd

            stats.equity_curve.append(capital)
            last_trade_idx = exit_idx

            if stats.total_trades % 100 == 0:
                print(f"  {stats.total_trades} trades processed...")

        # Final statistics
        if stats.total_trades > 0:
            stats.win_rate = stats.wins / stats.total_trades * 100
            stats.avg_win = stats.total_profit / stats.wins if stats.wins > 0 else 0
            stats.avg_loss = stats.total_loss / stats.losses if stats.losses > 0 else 0
            stats.avg_trade = (stats.total_profit - stats.total_loss) / stats.total_trades
            stats.profit_factor = stats.total_profit / stats.total_loss if stats.total_loss > 0 else float("inf")

            win_prob = stats.wins / stats.total_trades
            loss_prob = stats.losses / stats.total_trades
            stats.expectancy = (win_prob * stats.avg_win) - (loss_prob * stats.avg_loss)

            returns = [t.profit_usd for t in stats.trades]
            if len(returns) > 1:
                avg_return = np.mean(returns)
                std_return = np.std(returns)
                stats.sharpe_ratio = (avg_return / std_return) * np.sqrt(252) if std_return > 0 else 0

        return stats


# ─── XLSX Report ───────────────────────────────────────────────

def generate_xlsx_report(stats: BacktestStats, filepath: str, start_date: datetime, end_date: datetime):
    wb = Workbook()
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_font = Font(name="Calibri", bold=True, size=10)
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    net_pnl = stats.total_profit - stats.total_loss

    # ═══ SHEET 1: SUMMARY ═══
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1F4E79"
    ws.merge_cells("A1:F1")
    ws["A1"] = "XAUBot AI — SMC-Only Backtest Report (100% Synced)"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws["A2"] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"].font = Font(name="Calibri", size=10, italic=True)

    summary_data = [
        ("Performance Metrics", "", True),
        ("Total Trades", stats.total_trades, False),
        ("Wins", stats.wins, False),
        ("Losses", stats.losses, False),
        ("Win Rate", f"{stats.win_rate:.1f}%", False),
        ("Avoided (AVOID filter)", stats.avoided_signals, False),
        ("Recovery Mode Trades", stats.recovery_mode_trades, False),
        ("Daily Limit Stops", stats.daily_limit_stops, False),
        ("", "", False),
        ("Profit - Loss", "", True),
        ("Total Profit", f"${stats.total_profit:,.2f}", False),
        ("Total Loss", f"${stats.total_loss:,.2f}", False),
        ("Net PnL", f"${net_pnl:,.2f}", False),
        ("Profit Factor", f"{stats.profit_factor:.2f}", False),
        ("", "", False),
        ("Risk Metrics", "", True),
        ("Max Drawdown", f"{stats.max_drawdown:.1f}%", False),
        ("Max Drawdown ($)", f"${stats.max_drawdown_usd:,.2f}", False),
        ("Avg Win", f"${stats.avg_win:,.2f}", False),
        ("Avg Loss", f"${stats.avg_loss:,.2f}", False),
        ("Avg Trade", f"${stats.avg_trade:,.2f}", False),
        ("Expectancy", f"${stats.expectancy:,.2f}", False),
        ("Sharpe Ratio", f"{stats.sharpe_ratio:.2f}", False),
    ]

    row = 5
    for label, value, is_header in summary_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if is_header:
            ws.cell(row=row, column=1).font = subheader_font
            ws.cell(row=row, column=1).fill = subheader_fill
            ws.cell(row=row, column=2).fill = subheader_fill
        if label == "Net PnL":
            ws.cell(row=row, column=2).font = Font(bold=True, color="006100" if net_pnl > 0 else "9C0006")
        row += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18

    # Exit Reason Breakdown
    exit_counts = {}
    for t in stats.trades:
        reason = t.exit_reason.value
        exit_counts[reason] = exit_counts.get(reason, 0) + 1

    ws.cell(row=5, column=4, value="Exit Reasons")
    ws.cell(row=5, column=4).font = subheader_font
    ws.cell(row=5, column=4).fill = subheader_fill
    ws.cell(row=5, column=5).fill = subheader_fill
    ws.cell(row=5, column=6).fill = subheader_fill
    row = 6
    for reason, count in sorted(exit_counts.items(), key=lambda x: -x[1]):
        pct = count / stats.total_trades * 100 if stats.total_trades > 0 else 0
        ws.cell(row=row, column=4, value=reason)
        ws.cell(row=row, column=5, value=count)
        ws.cell(row=row, column=6, value=f"{pct:.1f}%")
        row += 1

    # Session Breakdown
    row += 1
    ws.cell(row=row, column=4, value="Session Performance")
    ws.cell(row=row, column=4).font = subheader_font
    ws.cell(row=row, column=4).fill = subheader_fill
    for c in range(5, 8):
        ws.cell(row=row, column=c).fill = subheader_fill
    row += 1
    for lbl, col in [("Session", 4), ("Trades", 5), ("WR", 6), ("Net PnL", 7)]:
        ws.cell(row=row, column=col, value=lbl).font = Font(bold=True)
    row += 1
    session_stats = {}
    for t in stats.trades:
        s = t.session
        if s not in session_stats:
            session_stats[s] = {"w": 0, "l": 0, "p": 0.0}
        if t.result == TradeResult.WIN:
            session_stats[s]["w"] += 1
        else:
            session_stats[s]["l"] += 1
        session_stats[s]["p"] += t.profit_usd
    for sess, d in sorted(session_stats.items(), key=lambda x: -x[1]["p"]):
        total = d["w"] + d["l"]
        wr = d["w"] / total * 100 if total > 0 else 0
        ws.cell(row=row, column=4, value=sess)
        ws.cell(row=row, column=5, value=total)
        ws.cell(row=row, column=6, value=f"{wr:.1f}%")
        ws.cell(row=row, column=7, value=f"${d['p']:,.2f}")
        ws.cell(row=row, column=7).font = Font(color="006100" if d["p"] >= 0 else "9C0006")
        row += 1

    # SMC Component Analysis
    row += 1
    ws.cell(row=row, column=4, value="SMC Component Analysis")
    ws.cell(row=row, column=4).font = subheader_font
    ws.cell(row=row, column=4).fill = subheader_fill
    for c in range(5, 8):
        ws.cell(row=row, column=c).fill = subheader_fill
    row += 1
    for lbl, col in [("Component", 4), ("Trades", 5), ("WR", 6), ("Net PnL", 7)]:
        ws.cell(row=row, column=col, value=lbl).font = Font(bold=True)
    row += 1
    for comp_name, attr in [("BOS", "has_bos"), ("CHoCH", "has_choch"), ("FVG", "has_fvg"), ("OB", "has_ob")]:
        ct = [t for t in stats.trades if getattr(t, attr)]
        cw = sum(1 for t in ct if t.result == TradeResult.WIN)
        cp = sum(t.profit_usd for t in ct)
        cwr = cw / len(ct) * 100 if ct else 0
        ws.cell(row=row, column=4, value=comp_name)
        ws.cell(row=row, column=5, value=len(ct))
        ws.cell(row=row, column=6, value=f"{cwr:.1f}%")
        ws.cell(row=row, column=7, value=f"${cp:,.2f}")
        row += 1

    col_widths = {4: 28, 5: 10, 6: 12, 7: 14}
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # ═══ SHEET 2: TRADE LOG ═══
    ws2 = wb.create_sheet("Trade Log")
    ws2.sheet_properties.tabColor = "2E75B6"
    headers = [
        "Ticket", "Entry Time", "Exit Time", "Dir", "Entry", "Exit", "SL", "TP",
        "Lot", "Profit ($)", "Pips", "Result", "Exit Reason", "SMC Conf",
        "Regime", "Session", "Signal", "BOS", "CHoCH", "FVG", "OB", "ATR", "RR", "Mode",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for ri, t in enumerate(stats.trades, 2):
        vals = [
            t.ticket, t.entry_time.strftime("%Y-%m-%d %H:%M"), t.exit_time.strftime("%Y-%m-%d %H:%M"),
            t.direction, t.entry_price, t.exit_price, t.stop_loss, t.take_profit,
            t.lot_size, round(t.profit_usd, 2), round(t.profit_pips, 1), t.result.value,
            t.exit_reason.value, round(t.smc_confidence, 2), t.regime, t.session, t.signal_reason,
            "Y" if t.has_bos else "", "Y" if t.has_choch else "", "Y" if t.has_fvg else "",
            "Y" if t.has_ob else "", round(t.atr_at_entry, 2), round(t.rr_ratio, 2), t.trading_mode,
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.border = border
            if ci == 10 and isinstance(v, (int, float)):
                cell.fill = win_fill if v > 0 else (loss_fill if v < 0 else PatternFill())
            if ci == 12:
                cell.fill = win_fill if v == "WIN" else (loss_fill if v == "LOSS" else PatternFill())

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = max(11, len(headers[col - 1]) + 3)

    # ═══ SHEET 3: EQUITY CURVE ═══
    ws3 = wb.create_sheet("Equity Curve")
    ws3.sheet_properties.tabColor = "548235"
    for c, h in enumerate(["Trade #", "Equity", "Drawdown ($)"], 1):
        ws3.cell(row=1, column=c, value=h).font = header_font
        ws3.cell(row=1, column=c).fill = header_fill

    peak = stats.equity_curve[0] if stats.equity_curve else 5000
    for idx, eq in enumerate(stats.equity_curve):
        if eq > peak:
            peak = eq
        ws3.cell(row=idx + 2, column=1, value=idx)
        ws3.cell(row=idx + 2, column=2, value=round(eq, 2))
        ws3.cell(row=idx + 2, column=3, value=round(peak - eq, 2))

    if len(stats.equity_curve) > 1:
        chart = LineChart()
        chart.title = "Equity Curve"
        chart.style = 10
        chart.y_axis.title = "Equity ($)"
        chart.x_axis.title = "Trade #"
        chart.width = 30
        chart.height = 15
        data = Reference(ws3, min_col=2, min_row=1, max_row=len(stats.equity_curve) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.series[0].graphicalProperties.line.width = 20000
        ws3.add_chart(chart, "E2")

    # ═══ SHEET 4: DAILY PnL ═══
    ws4 = wb.create_sheet("Daily PnL")
    ws4.sheet_properties.tabColor = "BF8F00"
    daily_pnl = {}
    for t in stats.trades:
        day = t.entry_time.strftime("%Y-%m-%d")
        if day not in daily_pnl:
            daily_pnl[day] = {"trades": 0, "wins": 0, "profit": 0.0}
        daily_pnl[day]["trades"] += 1
        if t.result == TradeResult.WIN:
            daily_pnl[day]["wins"] += 1
        daily_pnl[day]["profit"] += t.profit_usd

    for c, h in enumerate(["Date", "Trades", "Wins", "WR", "Net PnL", "Cumulative"], 1):
        ws4.cell(row=1, column=c, value=h).font = header_font
        ws4.cell(row=1, column=c).fill = header_fill

    cum = 0.0
    for ri, (day, d) in enumerate(sorted(daily_pnl.items()), 2):
        wr = d["wins"] / d["trades"] * 100 if d["trades"] > 0 else 0
        cum += d["profit"]
        ws4.cell(row=ri, column=1, value=day)
        ws4.cell(row=ri, column=2, value=d["trades"])
        ws4.cell(row=ri, column=3, value=d["wins"])
        ws4.cell(row=ri, column=4, value=f"{wr:.0f}%")
        ws4.cell(row=ri, column=5, value=round(d["profit"], 2))
        ws4.cell(row=ri, column=6, value=round(cum, 2))
        ws4.cell(row=ri, column=5).fill = win_fill if d["profit"] >= 0 else loss_fill

    for c in range(1, 7):
        ws4.column_dimensions[get_column_letter(c)].width = 16

    wb.save(filepath)
    print(f"\n  Report saved: {filepath}")


# ─── Log Generator ─────────────────────────────────────────────

def generate_log(stats: BacktestStats, filepath: str, start_date: datetime, end_date: datetime):
    net_pnl = stats.total_profit - stats.total_loss
    lines = []
    lines.append("=" * 80)
    lines.append("XAUBOT AI — SMC-Only Backtest Log (100% Synced with main_live.py)")
    lines.append("=" * 80)
    lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
    lines.append(f"Strategy: SMC-Only v4 + SmartRiskManager + SmartPositionManager")
    lines.append("")
    lines.append("--- PERFORMANCE SUMMARY ---")
    lines.append(f"  Total Trades:    {stats.total_trades}")
    lines.append(f"  Wins:            {stats.wins}")
    lines.append(f"  Losses:          {stats.losses}")
    lines.append(f"  Win Rate:        {stats.win_rate:.1f}%")
    lines.append(f"  Total Profit:    ${stats.total_profit:,.2f}")
    lines.append(f"  Total Loss:      ${stats.total_loss:,.2f}")
    lines.append(f"  Net PnL:         ${net_pnl:,.2f}")
    lines.append(f"  Profit Factor:   {stats.profit_factor:.2f}")
    lines.append(f"  Max Drawdown:    {stats.max_drawdown:.1f}% (${stats.max_drawdown_usd:,.2f})")
    lines.append(f"  Avg Win:         ${stats.avg_win:,.2f}")
    lines.append(f"  Avg Loss:        ${stats.avg_loss:,.2f}")
    lines.append(f"  Expectancy:      ${stats.expectancy:,.2f}")
    lines.append(f"  Sharpe Ratio:    {stats.sharpe_ratio:.2f}")
    lines.append(f"  Avoided (AVOID): {stats.avoided_signals}")
    lines.append(f"  Recovery Trades: {stats.recovery_mode_trades}")
    lines.append(f"  Daily Stops:     {stats.daily_limit_stops}")
    lines.append("")

    lines.append("--- EXIT REASON BREAKDOWN ---")
    exit_counts = {}
    for t in stats.trades:
        r = t.exit_reason.value
        exit_counts[r] = exit_counts.get(r, 0) + 1
    for reason, count in sorted(exit_counts.items(), key=lambda x: -x[1]):
        pct = count / stats.total_trades * 100 if stats.total_trades > 0 else 0
        lines.append(f"  {reason:20s}: {count:4d} ({pct:5.1f}%)")
    lines.append("")

    lines.append("--- DIRECTION BREAKDOWN ---")
    for d in ["BUY", "SELL"]:
        dt = [t for t in stats.trades if t.direction == d]
        dw = sum(1 for t in dt if t.result == TradeResult.WIN)
        dp = sum(t.profit_usd for t in dt)
        dwr = dw / len(dt) * 100 if dt else 0
        lines.append(f"  {d}: {len(dt)} trades, {dwr:.1f}% WR, ${dp:,.2f}")
    lines.append("")

    lines.append("--- SESSION BREAKDOWN ---")
    ss = {}
    for t in stats.trades:
        if t.session not in ss:
            ss[t.session] = {"w": 0, "l": 0, "p": 0.0}
        if t.result == TradeResult.WIN:
            ss[t.session]["w"] += 1
        else:
            ss[t.session]["l"] += 1
        ss[t.session]["p"] += t.profit_usd
    for s, d in sorted(ss.items(), key=lambda x: -x[1]["p"]):
        total = d["w"] + d["l"]
        wr = d["w"] / total * 100 if total > 0 else 0
        lines.append(f"  {s:30s}: {total:3d} trades, {wr:5.1f}% WR, ${d['p']:>8,.2f}")
    lines.append("")

    lines.append("--- SMC COMPONENT ANALYSIS ---")
    for cn, attr in [("BOS", "has_bos"), ("CHoCH", "has_choch"), ("FVG", "has_fvg"), ("OB", "has_ob")]:
        ct = [t for t in stats.trades if getattr(t, attr)]
        cw = sum(1 for t in ct if t.result == TradeResult.WIN)
        cp = sum(t.profit_usd for t in ct)
        cwr = cw / len(ct) * 100 if ct else 0
        lines.append(f"  {cn:6s}: {len(ct):3d} trades, {cwr:5.1f}% WR, ${cp:>8,.2f}")
    lines.append("")

    lines.append("--- TRADE LOG ---")
    lines.append(f"{'#':>4} {'Entry Time':>16} {'Dir':>4} {'Entry':>10} {'Exit':>10} {'P/L($)':>8} {'Result':>6} {'Exit Reason':>18} {'Conf':>5} {'Mode':>10} {'Session':>20}")
    lines.append("-" * 140)
    for idx, t in enumerate(stats.trades, 1):
        lines.append(
            f"{idx:4d} {t.entry_time.strftime('%Y-%m-%d %H:%M'):>16} {t.direction:>4} "
            f"{t.entry_price:>10.2f} {t.exit_price:>10.2f} {t.profit_usd:>8.2f} "
            f"{t.result.value:>6} {t.exit_reason.value:>18} {t.smc_confidence:>5.0%} "
            f"{t.trading_mode:>10} {t.session:>20}"
        )
    lines.append("\n" + "=" * 80)
    lines.append("END OF REPORT")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Log saved: {filepath}")


# ─── Main ──────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("XAUBOT AI — SMC-Only Backtest (100% Synced)")
    print("All 3 systems: SmartPositionManager + SmartRiskManager + Time/Trend")
    print("=" * 70)

    config = get_config()
    mt5 = MT5Connector(
        login=config.mt5_login,
        password=config.mt5_password,
        server=config.mt5_server,
        path=config.mt5_path,
    )
    mt5.connect()
    print(f"\nConnected to MT5")

    print("Fetching XAUUSD M15 historical data...")
    df = mt5.get_market_data(symbol="XAUUSD", timeframe="M15", count=50000)

    if len(df) == 0:
        print("ERROR: No data received")
        mt5.disconnect()
        return

    print(f"  Received {len(df)} bars")
    times = df["time"].to_list()
    print(f"  Data range: {times[0]} to {times[-1]}")

    end_date = datetime.now()
    start_date = datetime(2025, 8, 1)

    data_start = times[0]
    if hasattr(data_start, 'replace') and data_start.tzinfo:
        start_date = start_date.replace(tzinfo=data_start.tzinfo)
        end_date = end_date.replace(tzinfo=data_start.tzinfo)

    if data_start > start_date:
        start_date = data_start + timedelta(days=5)
        print(f"  [INFO] Adjusted start: {start_date}")

    print(f"\n  Backtest period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

    print("\nCalculating indicators...")
    features = FeatureEngineer()
    smc = SMCAnalyzer(swing_length=config.smc.swing_length, ob_lookback=config.smc.ob_lookback)

    df = features.calculate_all(df, include_ml_features=True)
    df = smc.calculate_all(df)

    regime_detector = MarketRegimeDetector(model_path="models/hmm_regime.pkl")
    try:
        regime_detector.load()
        df = regime_detector.predict(df)
        print("  HMM regime loaded")
    except Exception:
        print("  [WARN] HMM not available")

    print("  Indicators calculated")

    backtest = SMCOnlyBacktest(
        capital=5000.0,
        max_daily_loss_percent=5.0,
        max_loss_per_trade_percent=1.0,
        base_lot_size=0.01,
        max_lot_size=0.02,
        recovery_lot_size=0.01,
        breakeven_pips=30.0,
        trail_start_pips=50.0,
        trail_step_pips=30.0,
        min_profit_to_protect=5.0,
        max_drawdown_from_peak=50.0,
        trade_cooldown_bars=10,
        trend_reversal_mult=0.6,
    )

    stats = backtest.run(df=df, start_date=start_date, end_date=end_date, initial_capital=5000.0)

    net_pnl = stats.total_profit - stats.total_loss

    print("\n" + "=" * 70)
    print("SMC-ONLY BACKTEST RESULTS (100% Synced)")
    print("=" * 70)
    print(f"\n  Strategy:  SMC-Only v4 + all 3 exit systems")
    print(f"  Synced:    SmartPositionManager + SmartRiskManager + DynamicConfidence")

    print(f"\n  Performance:")
    print(f"    Total Trades:     {stats.total_trades}")
    print(f"    Wins:             {stats.wins}")
    print(f"    Losses:           {stats.losses}")
    print(f"    Win Rate:         {stats.win_rate:.1f}%")

    print(f"\n  Profit/Loss:")
    print(f"    Total Profit:     ${stats.total_profit:,.2f}")
    print(f"    Total Loss:       ${stats.total_loss:,.2f}")
    print(f"    Net PnL:          ${net_pnl:,.2f}")
    print(f"    Profit Factor:    {stats.profit_factor:.2f}")

    print(f"\n  Risk Metrics:")
    print(f"    Max Drawdown:     {stats.max_drawdown:.1f}% (${stats.max_drawdown_usd:,.2f})")
    print(f"    Avg Win:          ${stats.avg_win:,.2f}")
    print(f"    Avg Loss:         ${stats.avg_loss:,.2f}")
    print(f"    Expectancy:       ${stats.expectancy:,.2f}")
    print(f"    Sharpe Ratio:     {stats.sharpe_ratio:.2f}")

    print(f"\n  Sync Metrics:")
    print(f"    Avoided (AVOID):  {stats.avoided_signals}")
    print(f"    Recovery Trades:  {stats.recovery_mode_trades}")
    print(f"    Daily Limit Stops:{stats.daily_limit_stops}")

    print(f"\n  Exit Reasons:")
    exit_counts = {}
    for t in stats.trades:
        r = t.exit_reason.value
        exit_counts[r] = exit_counts.get(r, 0) + 1
    for reason, count in sorted(exit_counts.items(), key=lambda x: -x[1]):
        pct = count / stats.total_trades * 100 if stats.total_trades > 0 else 0
        print(f"    {reason:20s}: {count} ({pct:.1f}%)")

    print(f"\n  Direction:")
    for d in ["BUY", "SELL"]:
        dt = [t for t in stats.trades if t.direction == d]
        dw = sum(1 for t in dt if t.result == TradeResult.WIN)
        dp = sum(t.profit_usd for t in dt)
        dwr = dw / len(dt) * 100 if dt else 0
        print(f"    {d}: {len(dt)} trades, {dwr:.1f}% WR, ${dp:,.2f}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "01_smc_only_results")
    os.makedirs(output_dir, exist_ok=True)

    log_path = os.path.join(output_dir, f"smc_only_synced_{timestamp}.log")
    xlsx_path = os.path.join(output_dir, f"smc_only_synced_{timestamp}.xlsx")

    generate_log(stats, log_path, start_date, end_date)
    generate_xlsx_report(stats, xlsx_path, start_date, end_date)

    mt5.disconnect()

    print("\n" + "=" * 70)
    print(f"Output: {output_dir}")
    print(f"  Log:    {os.path.basename(log_path)}")
    print(f"  Report: {os.path.basename(xlsx_path)}")
    print("=" * 70)
    print("Backtest complete!")


if __name__ == "__main__":
    main()
