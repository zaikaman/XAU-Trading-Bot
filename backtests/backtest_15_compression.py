"""
Backtest #15 — SMC + RTM Compression Filter
==================================================================
Base: SMC-Only v4 (#1 Baseline)
Added: RTM Compression detection as entry filter

Compression = price ranges narrowing + more base candles → zone approach is validated
When price compresses into OB/FVG zone, the zone is more likely to hold.

Usage:
    python backtests/backtest_15_compression.py
"""

import polars as pl
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, field
from enum import Enum
import sys
import os
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.mt5_connector import MT5Connector
from src.smc_polars import SMCAnalyzer, SMCSignal
from src.feature_eng import FeatureEngineer
from src.regime_detector import MarketRegimeDetector, MarketRegime
from src.ml_model import TradingModel
from src.config import get_config
from src.dynamic_confidence import DynamicConfidenceManager, create_dynamic_confidence, MarketQuality
from loguru import logger

logger.remove()
logger.add(sys.stderr, level="WARNING")

WIB = ZoneInfo("Asia/Jakarta")


# ─── Enums & Dataclasses ──────────────────────────────────────

class TradeResult(Enum):
    WIN = "WIN"
    LOSS = "LOSS"
    BREAKEVEN = "BREAKEVEN"


class ExitReason(Enum):
    TAKE_PROFIT = "take_profit"
    SMART_TP = "smart_tp"
    PEAK_PROTECT = "peak_protect"
    EARLY_EXIT = "early_exit"
    EARLY_CUT = "early_cut"
    MAX_LOSS = "max_loss"
    STALL = "stall"
    TREND_REVERSAL = "trend_reversal"
    TIMEOUT = "timeout"
    WEEKEND_CLOSE = "weekend_close"
    TRAILING_SL = "trailing_sl"
    BREAKEVEN_EXIT = "breakeven_exit"
    DAILY_LIMIT = "daily_limit"
    REGIME_DANGER = "regime_danger"
    MARKET_SIGNAL = "market_signal"


class TradingMode(Enum):
    NORMAL = "normal"
    RECOVERY = "recovery"
    PROTECTED = "protected"
    STOPPED = "stopped"


@dataclass
class SimulatedTrade:
    ticket: int
    entry_time: datetime
    exit_time: datetime
    direction: str
    entry_price: float
    exit_price: float
    stop_loss: float
    take_profit: float
    lot_size: float
    profit_usd: float
    profit_pips: float
    result: TradeResult
    exit_reason: ExitReason
    smc_confidence: float
    regime: str
    session: str
    signal_reason: str
    has_bos: bool = False
    has_choch: bool = False
    has_fvg: bool = False
    has_ob: bool = False
    atr_at_entry: float = 0.0
    rr_ratio: float = 0.0
    trading_mode: str = "normal"
    compression_score: float = 0.0


@dataclass
class BacktestStats:
    total_trades: int = 0
    wins: int = 0
    losses: int = 0
    total_profit: float = 0.0
    total_loss: float = 0.0
    max_drawdown: float = 0.0
    max_drawdown_usd: float = 0.0
    win_rate: float = 0.0
    profit_factor: float = 0.0
    avg_win: float = 0.0
    avg_loss: float = 0.0
    avg_trade: float = 0.0
    expectancy: float = 0.0
    sharpe_ratio: float = 0.0
    trades: List[SimulatedTrade] = field(default_factory=list)
    equity_curve: List[float] = field(default_factory=list)
    avoided_signals: int = 0
    daily_limit_stops: int = 0
    recovery_mode_trades: int = 0
    # Compression filter stats
    compression_blocked: int = 0
    compression_passed: int = 0
    avg_compression_score: float = 0.0


# ─── SMC + Compression Backtest ──────────────────────────

class CompressionBacktest:
    """SMC-Only v4 + RTM Compression entry filter."""

    def __init__(
        self,
        capital: float = 5000.0,
        max_daily_loss_percent: float = 5.0,
        max_loss_per_trade_percent: float = 1.0,
        base_lot_size: float = 0.01,
        max_lot_size: float = 0.02,
        recovery_lot_size: float = 0.01,
        trend_reversal_threshold: float = 0.75,
        max_concurrent_positions: int = 2,
        breakeven_pips: float = 30.0,
        trail_start_pips: float = 50.0,
        trail_step_pips: float = 30.0,
        min_profit_to_protect: float = 5.0,
        max_drawdown_from_peak: float = 50.0,
        trade_cooldown_bars: int = 10,
        trend_reversal_mult: float = 0.6,
        # ═══ Compression filter params ═══
        compression_lookback: int = 8,
        compression_min_score: float = 40.0,
    ):
        self.capital = capital
        self.max_daily_loss_usd = capital * (max_daily_loss_percent / 100)
        self.max_loss_per_trade = capital * (max_loss_per_trade_percent / 100)
        self.base_lot_size = base_lot_size
        self.max_lot_size = max_lot_size
        self.recovery_lot_size = recovery_lot_size
        self.trend_reversal_threshold = trend_reversal_threshold
        self.max_concurrent_positions = max_concurrent_positions
        self.breakeven_pips = breakeven_pips
        self.trail_start_pips = trail_start_pips
        self.trail_step_pips = trail_step_pips
        self.min_profit_to_protect = min_profit_to_protect
        self.max_drawdown_from_peak = max_drawdown_from_peak
        self.trade_cooldown_bars = trade_cooldown_bars
        self.trend_reversal_mult = trend_reversal_mult
        self.compression_lookback = compression_lookback
        self.compression_min_score = compression_min_score

        config = get_config()
        self.smc = SMCAnalyzer(
            swing_length=config.smc.swing_length,
            ob_lookback=config.smc.ob_lookback,
        )
        self.features = FeatureEngineer()
        self.dynamic_confidence = create_dynamic_confidence()

        self.ml_model = TradingModel(model_path="models/xgboost_model.pkl")
        try:
            self.ml_model.load()
            print("  ML model loaded (for exit evaluation)")
        except Exception:
            print("  [WARN] ML model not loaded — exit ML checks disabled")

        self.regime_detector = MarketRegimeDetector(model_path="models/hmm_regime.pkl")
        try:
            self.regime_detector.load()
        except Exception:
            print("  [WARN] HMM model not loaded")

        self._ticket_counter = 2000000

    # ═══ RTM COMPRESSION DETECTION ═══

    def _detect_compression(self, df_slice: pl.DataFrame, lookback: int = None) -> Tuple[bool, float]:
        """
        Detect RTM Compression: price ranges narrowing + base candles increasing.

        Compression scoring (0-100):
          - Range narrowing (recent vs older): up to 40 points
          - Base candle ratio (body < 50% of range): up to 30 points
          - ATR declining: up to 30 points

        Returns: (is_compressed, score)
        """
        if lookback is None:
            lookback = self.compression_lookback

        n = len(df_slice)
        if n < lookback + 5:
            return False, 0.0

        highs = df_slice["high"].to_list()
        lows = df_slice["low"].to_list()
        opens = df_slice["open"].to_list()
        closes = df_slice["close"].to_list()

        half = lookback // 2

        # 1. Range narrowing: compare recent half vs older half
        recent_ranges = []
        older_ranges = []
        for j in range(half):
            recent_ranges.append(highs[n - 1 - j] - lows[n - 1 - j])
            older_ranges.append(highs[n - 1 - half - j] - lows[n - 1 - half - j])

        avg_recent = sum(recent_ranges) / len(recent_ranges) if recent_ranges else 1
        avg_older = sum(older_ranges) / len(older_ranges) if older_ranges else 1

        if avg_older <= 0:
            return False, 0.0

        range_ratio = avg_recent / avg_older  # < 1.0 = narrowing

        # 2. Base candle count in lookback window
        base_count = 0
        for j in range(lookback):
            idx = n - 1 - j
            if idx < 0:
                continue
            body = abs(closes[idx] - opens[idx])
            candle_range = highs[idx] - lows[idx]
            if candle_range > 0 and body / candle_range < 0.50:
                base_count += 1

        base_ratio = base_count / lookback

        # 3. ATR declining
        atr_score = 0.0
        if "atr" in df_slice.columns:
            atr_list = df_slice["atr"].to_list()
            current_atr = atr_list[-1]
            older_idx = max(0, n - lookback - 1)
            older_atr = atr_list[older_idx]
            if current_atr and older_atr and older_atr > 0:
                atr_ratio = current_atr / older_atr
                if atr_ratio < 0.80:
                    atr_score = 30
                elif atr_ratio < 0.90:
                    atr_score = 20
                elif atr_ratio < 0.95:
                    atr_score = 10

        # Score calculation
        score = 0.0

        # Range narrowing score (max 40)
        if range_ratio < 0.60:
            score += 40
        elif range_ratio < 0.70:
            score += 30
        elif range_ratio < 0.80:
            score += 20
        elif range_ratio < 0.90:
            score += 10

        # Base candle score (max 30)
        if base_ratio >= 0.625:  # 5/8
            score += 30
        elif base_ratio >= 0.50:  # 4/8
            score += 20
        elif base_ratio >= 0.375:  # 3/8
            score += 10

        # ATR score (max 30)
        score += atr_score

        is_compressed = score >= self.compression_min_score
        return is_compressed, score

    # ── Session filter (synced) ──

    def _get_session_from_time(self, dt: datetime) -> Tuple[str, bool, float]:
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib_time = dt.astimezone(WIB)
        hour = wib_time.hour
        if 6 <= hour < 15:
            return "Sydney-Tokyo", True, 0.5
        elif 15 <= hour < 16:
            return "Tokyo-London Overlap", True, 0.75
        elif 16 <= hour < 19:
            return "London Early", True, 0.8
        elif 19 <= hour < 24:
            return "London-NY Overlap (Golden)", True, 1.0
        elif 0 <= hour < 4:
            return "NY Session", True, 0.9
        else:
            return "Off Hours", False, 0.0

    def _hours_to_golden(self, dt: datetime) -> float:
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib = dt.astimezone(WIB)
        if 19 <= wib.hour < 24:
            return 0
        target = wib.replace(hour=19, minute=0, second=0, microsecond=0)
        if wib.hour >= 19:
            target += timedelta(days=1)
        return max(0, (target - wib).total_seconds() / 3600)

    def _is_near_weekend_close(self, dt: datetime) -> bool:
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        wib = dt.astimezone(WIB)
        if wib.weekday() == 5 and wib.hour >= 4 and wib.minute >= 30:
            return True
        return False

    # ── Lot sizing (synced) ──

    def _calculate_lot_size(self, confidence, regime, trading_mode, session_mult):
        if trading_mode == TradingMode.STOPPED:
            return 0
        lot = self.base_lot_size
        if trading_mode in (TradingMode.RECOVERY, TradingMode.PROTECTED):
            lot = self.recovery_lot_size
        else:
            if confidence >= 0.65:
                lot = self.max_lot_size
            elif confidence >= 0.55:
                lot = self.base_lot_size
            else:
                lot = self.recovery_lot_size
        if regime.lower() in ["high_volatility", "crisis"]:
            lot = self.recovery_lot_size
        lot = max(0.01, lot * session_mult)
        return round(lot, 2)

    # ── Full exit simulation (all 3 systems — same as baseline) ──

    def _simulate_trade_exit(
        self, df, entry_idx, direction, entry_price, take_profit, stop_loss,
        lot_size, daily_loss_so_far, feature_cols, max_bars=100,
    ) -> Tuple[float, float, ExitReason, int, float]:
        pip_value = 10
        highs = df["high"].to_list()
        lows = df["low"].to_list()
        closes = df["close"].to_list()
        times = df["time"].to_list()

        atr = 12.0
        if "atr" in df.columns:
            atr_list = df["atr"].to_list()
            if entry_idx < len(atr_list) and atr_list[entry_idx] is not None:
                atr = atr_list[entry_idx]

        reversal_momentum_threshold = atr * self.trend_reversal_mult
        min_loss_for_reversal_exit = atr * 0.8

        profit_history = []
        price_history = []
        peak_profit = 0.0
        stall_count = 0
        reversal_warnings = 0
        current_sl = stop_loss
        breakeven_moved = False

        if direction == "BUY":
            target_tp_profit = (take_profit - entry_price) / 0.1 * pip_value * lot_size
        else:
            target_tp_profit = (entry_price - take_profit) / 0.1 * pip_value * lot_size

        cached_ml_signal = ""
        cached_ml_confidence = 0.5

        for i in range(entry_idx + 1, min(entry_idx + max_bars, len(df))):
            high = highs[i]
            low = lows[i]
            close = closes[i]
            current_time = times[i]

            if direction == "BUY":
                current_pips = (close - entry_price) / 0.1
                pip_profit_from_entry = (close - entry_price) / 0.1
            else:
                current_pips = (entry_price - close) / 0.1
                pip_profit_from_entry = (entry_price - close) / 0.1
            current_profit = current_pips * pip_value * lot_size

            profit_history.append(current_profit)
            price_history.append(close)
            if current_profit > peak_profit:
                peak_profit = current_profit

            bars_since_entry = i - entry_idx

            if bars_since_entry % 4 == 0 and self.ml_model.fitted:
                try:
                    df_slice = df.head(i + 1)
                    ml_pred = self.ml_model.predict(df_slice, feature_cols)
                    cached_ml_signal = ml_pred.signal
                    cached_ml_confidence = ml_pred.confidence
                except Exception:
                    pass

            momentum = 0.0
            if len(profit_history) >= 3:
                recent = profit_history[-5:] if len(profit_history) >= 5 else profit_history
                profit_change = recent[-1] - recent[0]
                momentum = max(-100, min(100, (profit_change / 10) * 50))

            profit_growing = momentum > 0

            # A) SmartPositionManager
            if direction == "BUY" and high >= take_profit:
                pips = (take_profit - entry_price) / 0.1
                profit = pips * pip_value * lot_size
                return profit, pips, ExitReason.TAKE_PROFIT, i, take_profit
            elif direction == "SELL" and low <= take_profit:
                pips = (entry_price - take_profit) / 0.1
                profit = pips * pip_value * lot_size
                return profit, pips, ExitReason.TAKE_PROFIT, i, take_profit

            if breakeven_moved and current_sl > 0:
                if direction == "BUY" and low <= current_sl:
                    pips = (current_sl - entry_price) / 0.1
                    profit = pips * pip_value * lot_size
                    reason = ExitReason.TRAILING_SL if pip_profit_from_entry >= self.trail_start_pips else ExitReason.BREAKEVEN_EXIT
                    return profit, pips, reason, i, current_sl
                elif direction == "SELL" and high >= current_sl:
                    pips = (entry_price - current_sl) / 0.1
                    profit = pips * pip_value * lot_size
                    reason = ExitReason.TRAILING_SL if pip_profit_from_entry >= self.trail_start_pips else ExitReason.BREAKEVEN_EXIT
                    return profit, pips, reason, i, current_sl

            if pip_profit_from_entry >= self.breakeven_pips and not breakeven_moved:
                if direction == "BUY":
                    current_sl = entry_price + 2
                else:
                    current_sl = entry_price - 2
                breakeven_moved = True

            if pip_profit_from_entry >= self.trail_start_pips:
                trail_distance = self.trail_step_pips * 0.1
                if direction == "BUY":
                    new_trail_sl = close - trail_distance
                    if new_trail_sl > current_sl:
                        current_sl = new_trail_sl
                else:
                    new_trail_sl = close + trail_distance
                    if current_sl == 0 or new_trail_sl < current_sl:
                        current_sl = new_trail_sl

            if peak_profit > self.min_profit_to_protect:
                drawdown_pct = ((peak_profit - current_profit) / peak_profit) * 100 if peak_profit > 0 else 0
                if drawdown_pct > self.max_drawdown_from_peak:
                    return current_profit, current_pips, ExitReason.PEAK_PROTECT, i, close

            if bars_since_entry % 5 == 0 and bars_since_entry >= 5:
                if i >= 20:
                    ma_fast = np.mean(closes[i-4:i+1])
                    ma_slow = np.mean(closes[i-19:i+1])
                    trend = "NEUTRAL"
                    if ma_fast > ma_slow * 1.001:
                        trend = "BULLISH"
                    elif ma_fast < ma_slow * 0.999:
                        trend = "BEARISH"

                    roc = (closes[i] / closes[max(0,i-4)] - 1) * 100
                    mom_dir = "BULLISH" if roc > 0.3 else ("BEARISH" if roc < -0.3 else "NEUTRAL")

                    rsi_val = None
                    if "rsi" in df.columns:
                        rsi_list = df["rsi"].to_list()
                        if i < len(rsi_list):
                            rsi_val = rsi_list[i]

                    urgency = 0
                    should_exit = False

                    if cached_ml_confidence > 0.75:
                        if direction == "BUY" and cached_ml_signal == "SELL":
                            should_exit = True
                            urgency += 2
                        elif direction == "SELL" and cached_ml_signal == "BUY":
                            should_exit = True
                            urgency += 2

                    if rsi_val:
                        if rsi_val > 75 and direction == "BUY":
                            should_exit = True
                            urgency += 2
                        elif rsi_val < 25 and direction == "SELL":
                            should_exit = True
                            urgency += 2

                    if direction == "BUY" and trend == "BEARISH" and mom_dir == "BEARISH":
                        should_exit = True
                        urgency += 3
                    elif direction == "SELL" and trend == "BULLISH" and mom_dir == "BULLISH":
                        should_exit = True
                        urgency += 3

                    if should_exit and current_profit > self.min_profit_to_protect / 2:
                        return current_profit, current_pips, ExitReason.MARKET_SIGNAL, i, close
                    if urgency >= 7 and current_profit > 0:
                        return current_profit, current_pips, ExitReason.MARKET_SIGNAL, i, close

            if self._is_near_weekend_close(current_time):
                if current_profit > 0:
                    return current_profit, current_pips, ExitReason.WEEKEND_CLOSE, i, close
                elif current_profit > -10:
                    return current_profit, current_pips, ExitReason.WEEKEND_CLOSE, i, close

            # B) SmartRiskManager
            if current_profit >= 15:
                if current_profit >= 40:
                    return current_profit, current_pips, ExitReason.SMART_TP, i, close
                if current_profit >= 25 and momentum < -30:
                    return current_profit, current_pips, ExitReason.SMART_TP, i, close
                if peak_profit > 30 and current_profit < peak_profit * 0.6:
                    return current_profit, current_pips, ExitReason.PEAK_PROTECT, i, close
                if current_profit >= 20:
                    progress = (current_profit / target_tp_profit) * 100 if target_tp_profit > 0 else 0
                    progress_score = min(40, max(0, progress * 0.4))
                    momentum_score = ((momentum + 100) / 200) * 30
                    time_penalty = min(10, bars_since_entry / 4 * 2)
                    tp_probability = progress_score + momentum_score + 10 - time_penalty
                    if tp_probability < 25:
                        return current_profit, current_pips, ExitReason.SMART_TP, i, close

            if 5 <= current_profit < 15:
                if momentum < -50 and cached_ml_confidence >= 0.65:
                    is_reversal = (
                        (direction == "BUY" and cached_ml_signal == "SELL") or
                        (direction == "SELL" and cached_ml_signal == "BUY")
                    )
                    if is_reversal:
                        return current_profit, current_pips, ExitReason.EARLY_EXIT, i, close

            if current_profit < 0:
                loss_percent_of_max = abs(current_profit) / self.max_loss_per_trade * 100
                if momentum < -30 and loss_percent_of_max >= 30:
                    return current_profit, current_pips, ExitReason.EARLY_CUT, i, close

            is_ml_reversal = False
            if direction == "BUY" and cached_ml_signal == "SELL" and cached_ml_confidence >= self.trend_reversal_threshold:
                is_ml_reversal = True
                reversal_warnings += 1
            elif direction == "SELL" and cached_ml_signal == "BUY" and cached_ml_confidence >= self.trend_reversal_threshold:
                is_ml_reversal = True
                reversal_warnings += 1

            loss_moderate = abs(current_profit) > (self.max_loss_per_trade * 0.4)
            if is_ml_reversal and current_profit < -8 and loss_moderate:
                return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close
            if reversal_warnings >= 3 and current_profit < -10:
                return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close

            if current_profit <= -(self.max_loss_per_trade * 0.50):
                htg = self._hours_to_golden(current_time)
                if htg <= 1 and htg > 0 and momentum > -40:
                    pass
                else:
                    return current_profit, current_pips, ExitReason.MAX_LOSS, i, close

            if len(profit_history) >= 10:
                recent_range = max(profit_history[-10:]) - min(profit_history[-10:])
                if recent_range < 3 and current_profit < -15:
                    stall_count += 1
                    if stall_count >= 5:
                        return current_profit, current_pips, ExitReason.STALL, i, close

            potential_daily_loss = daily_loss_so_far + abs(min(0, current_profit))
            if potential_daily_loss >= self.max_daily_loss_usd:
                return current_profit, current_pips, ExitReason.DAILY_LIMIT, i, close

            # C) Time-based exit
            ml_agrees = (
                (direction == "BUY" and cached_ml_signal == "BUY") or
                (direction == "SELL" and cached_ml_signal == "SELL")
            )

            if bars_since_entry >= 16:
                if current_profit < 5 and not profit_growing:
                    if current_profit >= 0:
                        return current_profit, current_pips, ExitReason.TIMEOUT, i, close
                    elif current_profit > -15:
                        return current_profit, current_pips, ExitReason.TIMEOUT, i, close

            if bars_since_entry >= 24:
                if current_profit < 10 or not profit_growing:
                    return current_profit, current_pips, ExitReason.TIMEOUT, i, close

            if bars_since_entry >= 32:
                return current_profit, current_pips, ExitReason.TIMEOUT, i, close

            if bars_since_entry > 10:
                recent_closes = closes[i-5:i+1]
                mom = recent_closes[-1] - recent_closes[0]
                if direction == "BUY" and mom < -reversal_momentum_threshold:
                    if current_profit < -min_loss_for_reversal_exit:
                        return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close
                elif direction == "SELL" and mom > reversal_momentum_threshold:
                    if current_profit < -min_loss_for_reversal_exit:
                        return current_profit, current_pips, ExitReason.TREND_REVERSAL, i, close

        final_idx = min(entry_idx + max_bars - 1, len(df) - 1)
        final_price = closes[final_idx]
        if direction == "BUY":
            pips = (final_price - entry_price) / 0.1
        else:
            pips = (entry_price - final_price) / 0.1
        profit = pips * pip_value * lot_size
        return profit, pips, ExitReason.TIMEOUT, final_idx, final_price

    # ── Main backtest run ──

    def run(self, df, start_date=None, end_date=None, initial_capital=5000.0):
        stats = BacktestStats()
        capital = initial_capital
        peak_capital = initial_capital
        stats.equity_curve.append(capital)

        daily_loss = 0.0
        daily_profit = 0.0
        daily_trades = 0
        consecutive_losses = 0
        trading_mode = TradingMode.NORMAL
        current_date = None

        feature_cols = []
        if self.ml_model.fitted and self.ml_model.feature_names:
            feature_cols = [f for f in self.ml_model.feature_names if f in df.columns]

        times = df["time"].to_list()
        start_idx = next((i for i, t in enumerate(times) if t >= start_date), 100) if start_date else 100
        end_idx = next((i for i, t in enumerate(times) if t > end_date), len(df) - 100) if end_date else len(df) - 100

        last_trade_idx = -self.trade_cooldown_bars * 2
        compression_scores = []

        print(f"\n  Running SMC + Compression Filter backtest...")
        print(f"  Compression: lookback={self.compression_lookback}, min_score={self.compression_min_score}")
        print(f"  Date range: {times[start_idx]} to {times[end_idx - 1]}")
        print(f"  Total bars: {end_idx - start_idx}")

        for i in range(start_idx, end_idx):
            if i - last_trade_idx < self.trade_cooldown_bars:
                continue

            current_time = times[i]

            trade_date = current_time.date() if hasattr(current_time, 'date') else current_time
            if current_date is None or trade_date != current_date:
                daily_loss = 0.0
                daily_profit = 0.0
                daily_trades = 0
                current_date = trade_date
                if consecutive_losses < 2:
                    trading_mode = TradingMode.NORMAL

            if trading_mode == TradingMode.STOPPED:
                continue

            session_name, can_trade, lot_mult = self._get_session_from_time(current_time)
            if not can_trade:
                continue

            if hasattr(current_time, 'weekday') and current_time.weekday() >= 5:
                continue

            df_slice = df.head(i + 1)

            regime = "normal"
            try:
                if self.regime_detector.fitted:
                    regime_state = self.regime_detector.get_current_state(df_slice)
                    if regime_state:
                        regime = regime_state.regime.value
                        if regime_state.regime == MarketRegime.CRISIS:
                            continue
                        if regime_state.recommendation == "SLEEP":
                            continue
            except Exception:
                pass

            # Dynamic confidence AVOID
            try:
                ml_signal = ""
                ml_confidence = 0.5
                if self.ml_model.fitted and feature_cols:
                    ml_pred = self.ml_model.predict(df_slice, feature_cols)
                    ml_signal = ml_pred.signal
                    ml_confidence = ml_pred.confidence

                market_analysis = self.dynamic_confidence.analyze_market(
                    session=session_name, regime=regime, volatility="medium",
                    trend_direction=regime, has_smc_signal=True,
                    ml_signal=ml_signal, ml_confidence=ml_confidence,
                )
                if market_analysis.quality == MarketQuality.AVOID:
                    stats.avoided_signals += 1
                    continue
            except Exception:
                pass

            # SMC Signal
            try:
                smc_signal = self.smc.generate_signal(df_slice)
            except Exception:
                continue

            if smc_signal is None:
                continue

            # ═══ COMPRESSION FILTER (RTM) ═══
            is_compressed, comp_score = self._detect_compression(df_slice)
            compression_scores.append(comp_score)

            if not is_compressed:
                stats.compression_blocked += 1
                continue

            stats.compression_passed += 1

            # SMC details
            recent_df = df_slice.tail(10)
            recent_bos = recent_df["bos"].to_list() if "bos" in df_slice.columns else []
            recent_choch = recent_df["choch"].to_list() if "choch" in df_slice.columns else []
            recent_fvg_bull = recent_df["is_fvg_bull"].to_list() if "is_fvg_bull" in df_slice.columns else []
            recent_fvg_bear = recent_df["is_fvg_bear"].to_list() if "is_fvg_bear" in df_slice.columns else []
            recent_obs = recent_df["ob"].to_list() if "ob" in df_slice.columns else []

            has_bos = 1 in recent_bos or -1 in recent_bos
            has_choch = 1 in recent_choch or -1 in recent_choch
            has_fvg = any(recent_fvg_bull) or any(recent_fvg_bear)
            has_ob = 1 in recent_obs or -1 in recent_obs

            atr_at_entry = 12.0
            if "atr" in df_slice.columns:
                atr_val = df_slice.tail(1)["atr"].item()
                if atr_val is not None and atr_val > 0:
                    atr_at_entry = atr_val

            confidence = smc_signal.confidence
            ml_agrees = (
                (smc_signal.signal_type == "BUY" and ml_signal == "BUY") or
                (smc_signal.signal_type == "SELL" and ml_signal == "SELL")
            )
            if ml_agrees:
                confidence = (smc_signal.confidence + ml_confidence) / 2
            if regime == "high_volatility":
                confidence *= 0.9

            lot_size = self._calculate_lot_size(confidence, regime, trading_mode, lot_mult)
            if lot_size <= 0:
                continue

            if trading_mode == TradingMode.RECOVERY:
                stats.recovery_mode_trades += 1

            entry_price = smc_signal.entry_price
            take_profit_price = smc_signal.take_profit
            stop_loss_price = smc_signal.stop_loss
            risk = abs(entry_price - stop_loss_price)
            rr = abs(take_profit_price - entry_price) / risk if risk > 0 else 0

            profit, pips, exit_reason, exit_idx, exit_price = self._simulate_trade_exit(
                df=df, entry_idx=i, direction=smc_signal.signal_type,
                entry_price=entry_price, take_profit=take_profit_price,
                stop_loss=stop_loss_price, lot_size=lot_size,
                daily_loss_so_far=daily_loss, feature_cols=feature_cols,
            )

            self._ticket_counter += 1
            result = TradeResult.WIN if profit > 0 else (TradeResult.LOSS if profit < 0 else TradeResult.BREAKEVEN)

            trade = SimulatedTrade(
                ticket=self._ticket_counter, entry_time=current_time,
                exit_time=times[exit_idx] if exit_idx < len(times) else times[-1],
                direction=smc_signal.signal_type, entry_price=entry_price,
                exit_price=exit_price, stop_loss=stop_loss_price,
                take_profit=take_profit_price, lot_size=lot_size,
                profit_usd=profit, profit_pips=pips, result=result,
                exit_reason=exit_reason, smc_confidence=confidence,
                regime=regime, session=session_name,
                signal_reason=smc_signal.reason,
                has_bos=has_bos, has_choch=has_choch, has_fvg=has_fvg,
                has_ob=has_ob, atr_at_entry=atr_at_entry, rr_ratio=rr,
                trading_mode=trading_mode.value, compression_score=comp_score,
            )
            stats.trades.append(trade)

            stats.total_trades += 1
            daily_trades += 1
            capital += profit

            if profit > 0:
                stats.wins += 1
                stats.total_profit += profit
                daily_profit += profit
                consecutive_losses = 0
                if trading_mode == TradingMode.RECOVERY:
                    trading_mode = TradingMode.NORMAL
            else:
                stats.losses += 1
                stats.total_loss += abs(profit)
                daily_loss += abs(profit)
                consecutive_losses += 1

            if daily_loss >= self.max_daily_loss_usd:
                trading_mode = TradingMode.STOPPED
                stats.daily_limit_stops += 1
            elif consecutive_losses >= 3 or daily_loss >= self.max_daily_loss_usd * 0.6:
                trading_mode = TradingMode.PROTECTED
            elif consecutive_losses >= 2:
                trading_mode = TradingMode.RECOVERY

            if capital > peak_capital:
                peak_capital = capital
            drawdown_pct = (peak_capital - capital) / peak_capital * 100
            drawdown_usd = peak_capital - capital
            if drawdown_pct > stats.max_drawdown:
                stats.max_drawdown = drawdown_pct
                stats.max_drawdown_usd = drawdown_usd

            stats.equity_curve.append(capital)
            last_trade_idx = exit_idx

            if stats.total_trades % 100 == 0:
                print(f"  {stats.total_trades} trades processed...")

        # Final statistics
        if stats.total_trades > 0:
            stats.win_rate = stats.wins / stats.total_trades * 100
            stats.avg_win = stats.total_profit / stats.wins if stats.wins > 0 else 0
            stats.avg_loss = stats.total_loss / stats.losses if stats.losses > 0 else 0
            stats.avg_trade = (stats.total_profit - stats.total_loss) / stats.total_trades
            stats.profit_factor = stats.total_profit / stats.total_loss if stats.total_loss > 0 else float("inf")
            win_prob = stats.wins / stats.total_trades
            loss_prob = stats.losses / stats.total_trades
            stats.expectancy = (win_prob * stats.avg_win) - (loss_prob * stats.avg_loss)
            returns = [t.profit_usd for t in stats.trades]
            if len(returns) > 1:
                avg_return = np.mean(returns)
                std_return = np.std(returns)
                stats.sharpe_ratio = (avg_return / std_return) * np.sqrt(252) if std_return > 0 else 0

        if compression_scores:
            stats.avg_compression_score = sum(compression_scores) / len(compression_scores)

        return stats


# ─── XLSX Report ───────────────────────────────────────────────

def generate_xlsx_report(stats, filepath, start_date, end_date):
    wb = Workbook()
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_font = Font(name="Calibri", bold=True, size=10)
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    net_pnl = stats.total_profit - stats.total_loss

    ws = wb.active
    ws.title = "Summary"
    ws.merge_cells("A1:F1")
    ws["A1"] = "XAUBot AI — #15 SMC + Compression Filter Report"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws["A2"] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    summary_data = [
        ("Performance Metrics", "", True),
        ("Total Trades", stats.total_trades, False),
        ("Wins", stats.wins, False),
        ("Losses", stats.losses, False),
        ("Win Rate", f"{stats.win_rate:.1f}%", False),
        ("Compression Blocked", stats.compression_blocked, False),
        ("Compression Passed", stats.compression_passed, False),
        ("Avg Compression Score", f"{stats.avg_compression_score:.1f}", False),
        ("Avoided (AVOID)", stats.avoided_signals, False),
        ("Recovery Trades", stats.recovery_mode_trades, False),
        ("", "", False),
        ("Profit - Loss", "", True),
        ("Total Profit", f"${stats.total_profit:,.2f}", False),
        ("Total Loss", f"${stats.total_loss:,.2f}", False),
        ("Net PnL", f"${net_pnl:,.2f}", False),
        ("Profit Factor", f"{stats.profit_factor:.2f}", False),
        ("", "", False),
        ("Risk Metrics", "", True),
        ("Max Drawdown", f"{stats.max_drawdown:.1f}%", False),
        ("Max Drawdown ($)", f"${stats.max_drawdown_usd:,.2f}", False),
        ("Avg Win", f"${stats.avg_win:,.2f}", False),
        ("Avg Loss", f"${stats.avg_loss:,.2f}", False),
        ("Expectancy", f"${stats.expectancy:,.2f}", False),
        ("Sharpe Ratio", f"{stats.sharpe_ratio:.2f}", False),
    ]

    row = 5
    for label, value, is_header in summary_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if is_header:
            ws.cell(row=row, column=1).font = subheader_font
            ws.cell(row=row, column=1).fill = subheader_fill
            ws.cell(row=row, column=2).fill = subheader_fill
        row += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18

    # Exit reasons
    exit_counts = {}
    for t in stats.trades:
        r = t.exit_reason.value
        exit_counts[r] = exit_counts.get(r, 0) + 1

    ws.cell(row=5, column=4, value="Exit Reasons").font = subheader_font
    ws.cell(row=5, column=4).fill = subheader_fill
    row = 6
    for reason, count in sorted(exit_counts.items(), key=lambda x: -x[1]):
        pct = count / stats.total_trades * 100 if stats.total_trades > 0 else 0
        ws.cell(row=row, column=4, value=reason)
        ws.cell(row=row, column=5, value=count)
        ws.cell(row=row, column=6, value=f"{pct:.1f}%")
        row += 1

    # Trade Log sheet
    ws2 = wb.create_sheet("Trade Log")
    headers = [
        "Ticket", "Entry Time", "Exit Time", "Dir", "Entry", "Exit", "SL", "TP",
        "Lot", "Profit ($)", "Pips", "Result", "Exit Reason", "Comp Score",
        "Regime", "Session", "Signal", "BOS", "CHoCH", "FVG", "OB", "ATR", "RR", "Mode",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for ri, t in enumerate(stats.trades, 2):
        vals = [
            t.ticket, t.entry_time.strftime("%Y-%m-%d %H:%M"), t.exit_time.strftime("%Y-%m-%d %H:%M"),
            t.direction, t.entry_price, t.exit_price, t.stop_loss, t.take_profit,
            t.lot_size, round(t.profit_usd, 2), round(t.profit_pips, 1), t.result.value,
            t.exit_reason.value, round(t.compression_score, 1),
            t.regime, t.session, t.signal_reason,
            "Y" if t.has_bos else "", "Y" if t.has_choch else "", "Y" if t.has_fvg else "",
            "Y" if t.has_ob else "", round(t.atr_at_entry, 2), round(t.rr_ratio, 2), t.trading_mode,
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.border = border
            if ci == 10 and isinstance(v, (int, float)):
                cell.fill = win_fill if v > 0 else (loss_fill if v < 0 else PatternFill())

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = max(11, len(headers[col-1]) + 3)

    # Equity Curve sheet
    ws3 = wb.create_sheet("Equity Curve")
    for c, h in enumerate(["Trade #", "Equity", "Drawdown ($)"], 1):
        ws3.cell(row=1, column=c, value=h).font = header_font
        ws3.cell(row=1, column=c).fill = header_fill

    peak = stats.equity_curve[0] if stats.equity_curve else 5000
    for idx, eq in enumerate(stats.equity_curve):
        if eq > peak: peak = eq
        ws3.cell(row=idx+2, column=1, value=idx)
        ws3.cell(row=idx+2, column=2, value=round(eq, 2))
        ws3.cell(row=idx+2, column=3, value=round(peak - eq, 2))

    if len(stats.equity_curve) > 1:
        chart = LineChart()
        chart.title = "Equity Curve"
        chart.style = 10
        chart.y_axis.title = "Equity ($)"
        chart.width = 30
        chart.height = 15
        data = Reference(ws3, min_col=2, min_row=1, max_row=len(stats.equity_curve)+1)
        chart.add_data(data, titles_from_data=True)
        chart.series[0].graphicalProperties.line.width = 20000
        ws3.add_chart(chart, "E2")

    wb.save(filepath)
    print(f"\n  Report saved: {filepath}")


# ─── Main ──────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("XAUBOT AI — #15 SMC + RTM Compression Filter")
    print("Base: SMC-Only v4 | Added: RTM Compression entry filter")
    print("=" * 70)

    config = get_config()
    mt5 = MT5Connector(
        login=config.mt5_login, password=config.mt5_password,
        server=config.mt5_server, path=config.mt5_path,
    )
    mt5.connect()
    print(f"\nConnected to MT5")

    print("Fetching XAUUSD M15 historical data...")
    df = mt5.get_market_data(symbol="XAUUSD", timeframe="M15", count=50000)

    if len(df) == 0:
        print("ERROR: No data received")
        mt5.disconnect()
        return

    print(f"  Received {len(df)} bars")
    times = df["time"].to_list()
    print(f"  Data range: {times[0]} to {times[-1]}")

    end_date = datetime.now()
    start_date = datetime(2025, 8, 1)

    data_start = times[0]
    if hasattr(data_start, 'replace') and data_start.tzinfo:
        start_date = start_date.replace(tzinfo=data_start.tzinfo)
        end_date = end_date.replace(tzinfo=data_start.tzinfo)

    if data_start > start_date:
        start_date = data_start + timedelta(days=5)

    print(f"\n  Backtest period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

    print("\nCalculating indicators...")
    features = FeatureEngineer()
    smc = SMCAnalyzer(swing_length=config.smc.swing_length, ob_lookback=config.smc.ob_lookback)

    df = features.calculate_all(df, include_ml_features=True)
    df = smc.calculate_all(df)

    regime_detector = MarketRegimeDetector(model_path="models/hmm_regime.pkl")
    try:
        regime_detector.load()
        df = regime_detector.predict(df)
        print("  HMM regime loaded")
    except Exception:
        print("  [WARN] HMM not available")

    print("  Indicators calculated")

    backtest = CompressionBacktest(
        capital=5000.0,
        max_daily_loss_percent=5.0,
        max_loss_per_trade_percent=1.0,
        base_lot_size=0.01,
        max_lot_size=0.02,
        recovery_lot_size=0.01,
        breakeven_pips=30.0,
        trail_start_pips=50.0,
        trail_step_pips=30.0,
        min_profit_to_protect=5.0,
        max_drawdown_from_peak=50.0,
        trade_cooldown_bars=10,
        trend_reversal_mult=0.6,
        # Compression params
        compression_lookback=8,
        compression_min_score=40.0,
    )

    stats = backtest.run(df=df, start_date=start_date, end_date=end_date, initial_capital=5000.0)

    net_pnl = stats.total_profit - stats.total_loss
    total_signals = stats.compression_passed + stats.compression_blocked

    print("\n" + "=" * 70)
    print("#15 SMC + COMPRESSION FILTER — RESULTS")
    print("=" * 70)

    print(f"\n  Compression Filter:")
    print(f"    Total SMC signals:  {total_signals}")
    print(f"    Blocked (no comp):  {stats.compression_blocked}")
    print(f"    Passed (compressed):{stats.compression_passed}")
    filter_rate = stats.compression_blocked / total_signals * 100 if total_signals > 0 else 0
    print(f"    Filter rate:        {filter_rate:.1f}%")
    print(f"    Avg comp score:     {stats.avg_compression_score:.1f}")

    print(f"\n  Performance:")
    print(f"    Total Trades:     {stats.total_trades}")
    print(f"    Wins:             {stats.wins}")
    print(f"    Losses:           {stats.losses}")
    print(f"    Win Rate:         {stats.win_rate:.1f}%")

    print(f"\n  Profit/Loss:")
    print(f"    Total Profit:     ${stats.total_profit:,.2f}")
    print(f"    Total Loss:       ${stats.total_loss:,.2f}")
    print(f"    Net PnL:          ${net_pnl:,.2f}")
    print(f"    Profit Factor:    {stats.profit_factor:.2f}")

    print(f"\n  Risk Metrics:")
    print(f"    Max Drawdown:     {stats.max_drawdown:.1f}% (${stats.max_drawdown_usd:,.2f})")
    print(f"    Avg Win:          ${stats.avg_win:,.2f}")
    print(f"    Avg Loss:         ${stats.avg_loss:,.2f}")
    print(f"    Expectancy:       ${stats.expectancy:,.2f}")
    print(f"    Sharpe Ratio:     {stats.sharpe_ratio:.2f}")

    print(f"\n  Sync Metrics:")
    print(f"    Avoided (AVOID):  {stats.avoided_signals}")
    print(f"    Recovery Trades:  {stats.recovery_mode_trades}")
    print(f"    Daily Limit Stops:{stats.daily_limit_stops}")

    print(f"\n  vs BASELINE (#1): ${net_pnl - 1449.86:+,.2f}")

    print(f"\n  Exit Reasons:")
    exit_counts = {}
    for t in stats.trades:
        r = t.exit_reason.value
        exit_counts[r] = exit_counts.get(r, 0) + 1
    for reason, count in sorted(exit_counts.items(), key=lambda x: -x[1]):
        pct = count / stats.total_trades * 100 if stats.total_trades > 0 else 0
        print(f"    {reason:20s}: {count} ({pct:.1f}%)")

    print(f"\n  Direction:")
    for d in ["BUY", "SELL"]:
        dt = [t for t in stats.trades if t.direction == d]
        dw = sum(1 for t in dt if t.result == TradeResult.WIN)
        dp = sum(t.profit_usd for t in dt)
        dwr = dw / len(dt) * 100 if dt else 0
        print(f"    {d}: {len(dt)} trades, {dwr:.1f}% WR, ${dp:,.2f}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "15_compression_results")
    os.makedirs(output_dir, exist_ok=True)

    log_path = os.path.join(output_dir, f"compression_{timestamp}.log")
    xlsx_path = os.path.join(output_dir, f"compression_{timestamp}.xlsx")

    # Simple log
    lines = []
    lines.append("=" * 80)
    lines.append("#15 SMC + RTM Compression Filter — Backtest Log")
    lines.append("=" * 80)
    lines.append(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
    lines.append(f"Compression: lookback=8, min_score=40")
    lines.append(f"")
    lines.append(f"Compression Filter: {stats.compression_blocked} blocked, {stats.compression_passed} passed ({filter_rate:.1f}% filtered)")
    lines.append(f"Avg Compression Score: {stats.avg_compression_score:.1f}")
    lines.append(f"")
    lines.append(f"Trades: {stats.total_trades} | WR: {stats.win_rate:.1f}% | Net: ${net_pnl:,.2f}")
    lines.append(f"PF: {stats.profit_factor:.2f} | DD: {stats.max_drawdown:.1f}% | Sharpe: {stats.sharpe_ratio:.2f}")
    lines.append(f"Avg Win: ${stats.avg_win:,.2f} | Avg Loss: ${stats.avg_loss:,.2f}")
    lines.append(f"vs Baseline: ${net_pnl - 1449.86:+,.2f}")
    lines.append("")
    lines.append("--- TRADE LOG ---")
    lines.append(f"{'#':>4} {'Entry Time':>16} {'Dir':>4} {'Entry':>10} {'Exit':>10} {'P/L($)':>8} {'Result':>6} {'Exit Reason':>18} {'CmpScore':>8}")
    lines.append("-" * 110)
    for idx, t in enumerate(stats.trades, 1):
        lines.append(
            f"{idx:4d} {t.entry_time.strftime('%Y-%m-%d %H:%M'):>16} {t.direction:>4} "
            f"{t.entry_price:>10.2f} {t.exit_price:>10.2f} {t.profit_usd:>8.2f} "
            f"{t.result.value:>6} {t.exit_reason.value:>18} {t.compression_score:>8.1f}"
        )

    with open(log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Log saved: {log_path}")

    generate_xlsx_report(stats, xlsx_path, start_date, end_date)

    mt5.disconnect()

    print("\n" + "=" * 70)
    print(f"Output: {output_dir}")
    print("=" * 70)
    print("Backtest complete!")


if __name__ == "__main__":
    main()
